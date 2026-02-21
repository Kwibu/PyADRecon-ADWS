#!/usr/bin/env python3
"""
PyADRecon-ADWS # Python Active Directory Reconnaissance Tool using ADWS
A Python3 implementation of PyADRecon using ADWS instead of LDAP for Pentesters, Red and Blue Teams

This version uses ADWS (port 9389) instead of LDAP (port 389/636) for AD enumeration.
ADWS is often less monitored (AV/EDR) than traditional LDAP in enterprise environments.

Author: LRVT - https://github.com/l4rm4nd
License: MIT
"""

import argparse
import csv
import os
import sys
import socket
import struct
import ssl
import re
import json
import time
import random
from datetime import datetime, timedelta, timezone
from typing import Dict, List, Optional, Any, Tuple
from dataclasses import dataclass, field
from pathlib import Path
import logging
import threading
import subprocess
import tempfile
import traceback

# ADWS imports from helpers module
ADWS_AVAILABLE = False
try:
    from helpers.adws_wrapper import ADWSServer, ADWSConnection, ADWSEntry, ADWSAttribute
    ADWS_AVAILABLE = True
except ImportError as e:
    print("[!] ADWS helpers not available")
    print(f"[!] Error: {e}")
    print("[!] Make sure the helpers directory is present in the same directory as this script")
    sys.exit(1)

# Impacket imports (still needed for data type handling)
try:
    from impacket.krb5.kerberosv5 import getKerberosTGT, getKerberosTGS
    from impacket.krb5.types import Principal, KerberosTime, Ticket
    from impacket.krb5 import constants
    from impacket.krb5.asn1 import TGS_REP, EncTGSRepPart, EncTicketPart
    from impacket.krb5.ccache import CCache
    from impacket.smbconnection import SMBConnection
    from impacket.nmb import NetBIOSError
    from impacket.ldap import ldaptypes
    IMPACKET_AVAILABLE = True
except ImportError:
    IMPACKET_AVAILABLE = False
    print("[*] impacket not available - SMB features disabled")
    # Fallback class if impacket not available
    class ldaptypes:
        pass

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("[*] openpyxl not available - Excel export disabled")

# Try to import dashboard generator
DASHBOARD_AVAILABLE = False
try:
    from dashboard_generator import generate_dashboard
    DASHBOARD_AVAILABLE = True
except ImportError:
    DASHBOARD_AVAILABLE = False

# Constants
VERSION = "v0.4.12"  # Automatically updated by CI/CD pipeline during release
BANNER = f"""
╔═══════════════════════════════════════════════════════════════════
║  PyADRecon-ADWS {VERSION} - Python AD Reconnaissance Tool (ADWS)
║  A Python implementation inspired by ADRecon
║  -------------------------------------------------------
║  Author: LRVT - https://github.com/l4rm4nd/PyADRecon-ADWS              
╚═══════════════════════════════════════════════════════════════════
"""

# AD Constants
SAM_USER_OBJECT = 805306368
SAM_COMPUTER_OBJECT = 805306369
SAM_GROUP_OBJECT = 268435456
SAM_ALIAS_OBJECT = 536870912

# UserAccountControl flags
UAC_FLAGS = {
    0x0001: "SCRIPT",
    0x0002: "ACCOUNTDISABLE",
    0x0008: "HOMEDIR_REQUIRED",
    0x0010: "LOCKOUT",
    0x0020: "PASSWD_NOTREQD",
    0x0040: "PASSWD_CANT_CHANGE",
    0x0080: "ENCRYPTED_TEXT_PWD_ALLOWED",
    0x0100: "TEMP_DUPLICATE_ACCOUNT",
    0x0200: "NORMAL_ACCOUNT",
    0x0800: "INTERDOMAIN_TRUST_ACCOUNT",
    0x1000: "WORKSTATION_TRUST_ACCOUNT",
    0x2000: "SERVER_TRUST_ACCOUNT",
    0x10000: "DONT_EXPIRE_PASSWORD",
    0x20000: "MNS_LOGON_ACCOUNT",
    0x40000: "SMARTCARD_REQUIRED",
    0x80000: "TRUSTED_FOR_DELEGATION",
    0x100000: "NOT_DELEGATED",
    0x200000: "USE_DES_KEY_ONLY",
    0x400000: "DONT_REQ_PREAUTH",
    0x800000: "PASSWORD_EXPIRED",
    0x1000000: "TRUSTED_TO_AUTH_FOR_DELEGATION",
    0x4000000: "PARTIAL_SECRETS_ACCOUNT",
}

# Kerberos encryption types
KERB_ENC_FLAGS = {
    0x01: "DES_CBC_CRC",
    0x02: "DES_CBC_MD5",
    0x04: "RC4_HMAC",
    0x08: "AES128_CTS_HMAC_SHA1_96",
    0x10: "AES256_CTS_HMAC_SHA1_96",
}

# Trust directions
TRUST_DIRECTION = {
    0: "Disabled",
    1: "Inbound",
    2: "Outbound",
    3: "Bidirectional",
}

# Trust types
TRUST_TYPE = {
    1: "Downlevel",
    2: "Uplevel",
    3: "MIT",
    4: "DCE",
}

# Domain functional levels
DOMAIN_FUNCTIONAL_LEVELS = {
    0: "Windows2000",
    1: "Windows2003Mixed",
    2: "Windows2003",
    3: "Windows2008",
    4: "Windows2008R2",
    5: "Windows2012",
    6: "Windows2012R2",
    7: "Windows2016",
}

# Group types
GROUP_TYPE = {
    2: "Global Distribution",
    4: "Domain Local Distribution",
    8: "Universal Distribution",
    -2147483646: "Global Security",
    -2147483644: "Domain Local Security",
    -2147483640: "Universal Security",
}

# Logging setup - level will be configured in main() based on -v flag
logger = logging.getLogger('PyADRecon')


def windows_timestamp_to_datetime(timestamp: int) -> Optional[datetime]:
    """Convert Windows FILETIME to Python datetime."""
    if timestamp is None or timestamp == 0 or timestamp == 9223372036854775807:
        return None
    try:
        # Windows FILETIME is 100-nanosecond intervals since January 1, 1601
        return datetime(1601, 1, 1) + timedelta(microseconds=timestamp // 10)
    except (ValueError, OverflowError):
        return None


def generalized_time_to_datetime(time_str: str) -> Optional[datetime]:
    """Convert LDAP Generalized Time to Python datetime."""
    if not time_str:
        return None
    try:
        # Handle various formats
        if time_str.endswith('Z'):
            time_str = time_str[:-1]
        if '.' in time_str:
            return datetime.strptime(time_str, "%Y%m%d%H%M%S.%f")
        return datetime.strptime(time_str, "%Y%m%d%H%M%S")
    except ValueError:
        return None


def format_datetime(dt) -> str:
    """Format datetime in UTC (M/D/YYYY H:MM:SS AM/PM).
    All timestamps are displayed in UTC for consistency and timezone independence.
    """
    if dt is None:
        return ""

    if isinstance(dt, datetime):
        # Linux/macOS support %-m etc. Windows does not.
        if os.name == "nt":
            # Windows: %# removes leading zeros on most builds.
            try:
                return dt.strftime("%#m/%#d/%Y %#I:%M:%S %p")
            except ValueError:
                # Fallback: build without platform-specific flags
                m = dt.month
                d = dt.day
                y = dt.year
                hour12 = dt.hour % 12 or 12
                ampm = "AM" if dt.hour < 12 else "PM"
                return f"{m}/{d}/{y} {hour12}:{dt.minute:02d}:{dt.second:02d} {ampm}"
        else:
            return dt.strftime("%-m/%-d/%Y %-I:%M:%S %p")

    return ""


def _extract_adws_value(attr):
    """Extract primitive value from ADWS Attribute object."""
    if hasattr(attr, 'values'):
        # It's an ADWS Attribute object - use .values to get all values (returns list)
        return attr.values
    return attr


def safe_int(val, default=0):
    """Safely convert a value to int, handling ADWS Attribute objects."""
    if val is None:
        return default
    # Handle ADWS Attribute objects
    if hasattr(val, 'value'):
        val = val.value
    if val is None:
        return default
    try:
        return int(val)
    except (ValueError, TypeError):
        return default


def safe_str(val, default=''):
    """Safely convert a value to string, handling ADWS Attribute objects."""
    if val is None:
        return default
    # Handle ADWS Attribute objects
    if hasattr(val, 'value'):
        val = val.value
    if val is None:
        return default
    return str(val)


def get_attr(entry, attr_name: str, default=None):
    """Safely get attribute value from ADWS entry."""
    try:
        if attr_name in entry:
            attr = entry[attr_name]
            if attr is not None:
                # Extract value from ADWS Attribute if needed
                val = _extract_adws_value(attr)
                if val is not None:
                    if isinstance(val, list):
                        return val[0] if len(val) == 1 else val
                    return val
    except (IndexError, KeyError, AttributeError):
        pass
    return default


def get_attr_list(entry, attr_name: str) -> List:
    """Get attribute as list from ADWS entry."""
    try:
        if attr_name in entry:
            attr = entry[attr_name]
            if attr is not None and hasattr(attr, 'values'):
                return attr.values
            val = _extract_adws_value(attr)
            if val is not None:
                return [val] if not isinstance(val, list) else val
    except (IndexError, KeyError, AttributeError):
        pass
    return []


def dn_to_fqdn(dn: str) -> str:
    """Convert DN to FQDN (e.g., DC=example,DC=com -> example.com)."""
    if not dn:
        return ""
    parts = []
    for part in dn.split(','):
        if part.strip().upper().startswith('DC='):
            parts.append(part.split('=', 1)[1].strip())
    return '.'.join(parts)


def sid_to_str(sid_bytes: bytes) -> str:
    """Convert binary SID to string format.
    Handles both raw bytes and base64-encoded strings."""
    try:
        # Handle base64-encoded strings
        if isinstance(sid_bytes, str):
            import base64
            sid_bytes = base64.b64decode(sid_bytes)
        
        if IMPACKET_AVAILABLE:
            return ldaptypes.LDAP_SID(sid_bytes).formatCanonical()
        # Fallback manual parsing
        if len(sid_bytes) < 8:
            return ""
        revision = sid_bytes[0]
        sub_auth_count = sid_bytes[1]
        authority = struct.unpack('>Q', b'\x00\x00' + sid_bytes[2:8])[0]
        sub_auths = []
        for i in range(sub_auth_count):
            offset = 8 + (i * 4)
            sub_auth = struct.unpack('<I', sid_bytes[offset:offset + 4])[0]
            sub_auths.append(str(sub_auth))
        return f"S-{revision}-{authority}-" + '-'.join(sub_auths)
    except Exception:
        return ""


def parse_uac_flags(uac: int) -> List[str]:
    """Parse UserAccountControl flags into list of flag names."""
    flags = []
    for flag_val, flag_name in UAC_FLAGS.items():
        if uac & flag_val:
            flags.append(flag_name)
    return flags


def parse_kerb_enc_flags(enc: int) -> List[str]:
    """Parse Kerberos encryption flags into list of encryption types."""
    flags = []
    for flag_val, flag_name in KERB_ENC_FLAGS.items():
        if enc & flag_val:
            flags.append(flag_name)
    return flags


def parse_uac(uac) -> Dict[str, bool]:
    """Parse UserAccountControl value into individual flags."""
    if uac is None:
        return {}
    uac = safe_int(uac, 0)
    result = {}
    result['Enabled'] = not bool(uac & 0x0002)
    result['PasswordNotRequired'] = bool(uac & 0x0020)
    result['PasswordCantChange'] = bool(uac & 0x0040)
    result['ReversibleEncryption'] = bool(uac & 0x0080)
    result['PasswordNeverExpires'] = bool(uac & 0x10000)
    result['SmartcardRequired'] = bool(uac & 0x40000)
    result['TrustedForDelegation'] = bool(uac & 0x80000)
    result['NotDelegated'] = bool(uac & 0x100000)
    result['UseDESKeyOnly'] = bool(uac & 0x200000)
    result['DoesNotRequirePreAuth'] = bool(uac & 0x400000)
    result['PasswordExpired'] = bool(uac & 0x800000)
    result['TrustedToAuthForDelegation'] = bool(uac & 0x1000000)
    result['AccountLockedOut'] = bool(uac & 0x0010)
    return result


def parse_kerb_enc_types(enc_types) -> Dict[str, str]:
    """Parse msDS-SupportedEncryptionTypes into individual encryption types.
    
    If msDS-SupportedEncryptionTypes is not set (None/0), the DC uses defaults:
    - Windows Server 2008+: RC4, AES128, AES256 (all supported)
    - Older versions: RC4, DES (deprecated)
    
    Returns dict with 'Supported', 'Not Supported', or 'Default' for each type.
    """
    if enc_types is None or safe_int(enc_types, 0) == 0:
        # Attribute not set - system uses defaults (typically all modern algorithms)
        return {
            'RC4': 'Default',
            'AES128': 'Default',
            'AES256': 'Default',
        }
    
    enc_types = safe_int(enc_types, 0)
    return {
        'RC4': 'Supported' if (enc_types & 0x04) else 'Not Supported',
        'AES128': 'Supported' if (enc_types & 0x08) else 'Not Supported',
        'AES256': 'Supported' if (enc_types & 0x10) else 'Not Supported',
    }


def check_cannot_change_password(entry) -> bool:
    """
    Check if user cannot change password by examining ACL.
    This checks if SELF and Everyone are denied the 'Change Password' right.
    
    Change Password GUID: ab721a53-1e2f-11d0-9819-00aa0040529b (user-Change-Password)
    Access mask for ExtendedRight: 0x00000100 (ADS_RIGHT_DS_CONTROL_ACCESS)
    """
    try:
        # Get the security descriptor
        sd_attr = None
        if hasattr(entry, 'ntSecurityDescriptor'):
            if hasattr(entry.ntSecurityDescriptor, 'raw_values'):
                if entry.ntSecurityDescriptor.raw_values:
                    sd_attr = entry.ntSecurityDescriptor.raw_values[0]
        
        if not sd_attr or not isinstance(sd_attr, bytes) or len(sd_attr) < 20:
            return False
        
        # Parse security descriptor structure
        control = struct.unpack('<H', sd_attr[2:4])[0]
        dacl_offset = struct.unpack('<I', sd_attr[16:20])[0]
        
        # Check if DACL is present (SE_DACL_PRESENT = 0x0004)
        if dacl_offset == 0 or (control & 0x0004) == 0:
            return False
        
        if dacl_offset >= len(sd_attr):
            return False
        
        # Parse DACL
        dacl = sd_attr[dacl_offset:]
        if len(dacl) < 8:
            return False
        
        # DACL header: AclRevision(1) + Sbz1(1) + AclSize(2) + AceCount(2) + Sbz2(2)
        ace_count = struct.unpack('<H', dacl[4:6])[0]
        
        # Change Password GUID (user-Change-Password extended right)
        change_pwd_guid = bytes.fromhex('531a72ab2f1ed011981900aa0040529b')  # Little-endian format
        
        # Well-known SIDs in binary format
        self_sid = b'\x01\x01\x00\x00\x00\x00\x00\x05\x0a\x00\x00\x00'  # S-1-5-10 (SELF)
        everyone_sid = b'\x01\x01\x00\x00\x00\x00\x00\x01\x00\x00\x00\x00'  # S-1-1-0 (Everyone)
        
        # Parse ACEs
        offset = 8  # Start after DACL header
        for _ in range(ace_count):
            if offset + 4 > len(dacl):
                break
            
            ace_type = dacl[offset]
            ace_size = struct.unpack('<H', dacl[offset+2:offset+4])[0]
            
            if offset + ace_size > len(dacl) or ace_size < 16:
                break
            
            # Check for ACCESS_DENIED_OBJECT_ACE_TYPE (0x06) or ACCESS_DENIED_ACE_TYPE (0x01)
            if ace_type == 0x06:  # ACCESS_DENIED_OBJECT_ACE_TYPE
                # Parse ACCESS_DENIED_OBJECT_ACE
                if ace_size < 36:
                    offset += ace_size
                    continue
                
                access_mask = struct.unpack('<I', dacl[offset+4:offset+8])[0]
                flags = struct.unpack('<I', dacl[offset+8:offset+12])[0]
                
                # Check if this ACE applies to the Change Password extended right
                # ACE_OBJECT_TYPE_PRESENT = 0x1
                if flags & 0x1:
                    object_type = dacl[offset+12:offset+28]
                    
                    # Check if this is for Change Password permission
                    if object_type == change_pwd_guid:
                        # Find the SID offset
                        sid_offset = offset + 12 + 16
                        if flags & 0x2:  # ACE_INHERITED_OBJECT_TYPE_PRESENT
                            sid_offset += 16
                        
                        if sid_offset + 12 <= offset + ace_size:
                            sid = dacl[sid_offset:offset+ace_size]
                            
                            # Check if SID is SELF or Everyone
                            if len(sid) >= 12:
                                if sid[:12] == self_sid or sid[:12] == everyone_sid:
                                    return True
            
            offset += ace_size
        
        return False
        
    except Exception as e:
        # If we can't parse the security descriptor, return False
        return False


def calculate_user_stats(users_data: List[Dict], password_age_days: int = 180, dormant_days: int = 90) -> Dict[str, Any]:
    """Calculate user account statistics for the User Stats tab."""
    if not users_data:
        return {}
    
    stats = {
        'enabled_count': 0,
        'disabled_count': 0,
        'total_count': len(users_data),
        'categories': {},
        'password_age_days': password_age_days,
        'dormant_days': dormant_days
    }
    
    # Initialize category counters with dynamic thresholds
    categories = [
        'Must Change Password at Logon',
        'Cannot Change Password',
        'Password Never Expires',
        'Reversible Password Encryption',
        'Smartcard Logon Required',
        'Delegation Permitted',
        'Kerberos DES Only',
        'Kerberos RC4',
        'Does Not Require Pre Auth',
        f'Password Age (> {password_age_days} days)',
        'Account Locked Out',
        'Never Logged in',
        f'Dormant (> {dormant_days} days)',
        'Password Not Required',
        'Unconstrained Delegation',
        'SIDHistory'
    ]
    
    for cat in categories:
        stats['categories'][cat] = {
            'enabled_count': 0,
            'disabled_count': 0,
            'total_count': 0
        }
    
    # Count statistics
    for user in users_data:
        is_enabled = user.get('Enabled', False)
        # Use explicit True comparison to handle both boolean and string values
        if is_enabled is True or is_enabled == 'True':
            stats['enabled_count'] += 1
        else:
            stats['disabled_count'] += 1
        
        # Check each category
        for cat in categories:
            is_match = False
            
            if cat == 'Must Change Password at Logon':
                val = user.get('Must Change Password at Logon', False)
                is_match = val is True or val == 'True'
            elif cat == 'Cannot Change Password':
                val = user.get('Cannot Change Password', False)
                is_match = val is True or val == 'True'
            elif cat == 'Password Never Expires':
                val = user.get('Password Never Expires', False)
                is_match = val is True or val == 'True'
            elif cat == 'Reversible Password Encryption':
                val = user.get('Reversible Password Encryption', False)
                is_match = val is True or val == 'True'
            elif cat == 'Smartcard Logon Required':
                val = user.get('Smartcard Logon Required', False)
                is_match = val is True or val == 'True'
            elif cat == 'Delegation Permitted':
                val = user.get('Delegation Permitted', False)
                is_match = val is True or val == 'True'
            elif cat == 'Kerberos DES Only':
                val = user.get('Kerberos DES Only', False)
                is_match = val is True or val == 'True'
            elif cat == 'Kerberos RC4':
                val = user.get('Kerberos RC4', '')
                is_match = bool(val) and val not in ('', 'False', False)
            elif cat == 'Does Not Require Pre Auth':
                val = user.get('Does Not Require Pre Auth', False)
                is_match = val is True or val == 'True'
            elif cat.startswith('Password Age (>'):
                pwd_age = user.get('Password Age (days)', '')
                # Handle both int/float and string values
                try:
                    pwd_age_num = float(pwd_age) if pwd_age not in ('', None) else -1
                    is_match = pwd_age_num > password_age_days
                except (ValueError, TypeError):
                    is_match = False
            elif cat == 'Account Locked Out':
                val = user.get('Account Locked Out', False)
                is_match = val is True or val == 'True'
            elif cat == 'Never Logged in':
                val = user.get('Never Logged in', False)
                is_match = val is True or val == 'True'
            elif cat.startswith('Dormant (>'):
                # Check for dormant field - it might have different keys based on config
                for key in user.keys():
                    if 'Dormant' in key:
                        val = user.get(key, False)
                        is_match = val is True or val == 'True'
                        break
            elif cat == 'Password Not Required':
                val = user.get('Password Not Required', False)
                is_match = val is True or val == 'True'
            elif cat == 'Unconstrained Delegation':
                is_match = user.get('Delegation Type', '') == 'Unconstrained'
            elif cat == 'SIDHistory':
                is_match = bool(user.get('SIDHistory', ''))
            
            if is_match:
                stats['categories'][cat]['total_count'] += 1
                if is_enabled is True or is_enabled == 'True':
                    stats['categories'][cat]['enabled_count'] += 1
                else:
                    stats['categories'][cat]['disabled_count'] += 1
    
    return stats


def calculate_computer_stats(computers_data: List[Dict], laps_data: List[Dict] = None, password_age_days: int = 30, dormant_days: int = 90) -> Dict[str, Any]:
    """Calculate computer account statistics for the Computer Stats tab."""
    if not computers_data:
        return {}
    
    # Build a set of computer names that have LAPS
    laps_computers = set()
    if laps_data:
        for laps_entry in laps_data:
            hostname = laps_entry.get('Hostname', '')
            stored = laps_entry.get('Stored', False)
            # Handle both boolean and string values
            stored_bool = stored is True or stored == 'True'
            if hostname and stored_bool:
                # Normalize hostname (remove domain suffix if present)
                hostname_parts = hostname.split('.')
                laps_computers.add(hostname_parts[0].lower())
    
    stats = {
        'enabled_count': 0,
        'disabled_count': 0,
        'total_count': len(computers_data),
        'categories': {},
        'password_age_days': password_age_days,
        'dormant_days': dormant_days
    }
    
    # Initialize category counters with dynamic thresholds
    categories = [
        'Unconstrained Delegation',
        'Constrained Delegation',
        'SIDHistory',
        f'Dormant (> {dormant_days} days)',
        f'Password Age (> {password_age_days} days)',
        'ms-ds-CreatorSid',
        'LAPS'
    ]
    
    for cat in categories:
        stats['categories'][cat] = {
            'enabled_count': 0,
            'disabled_count': 0,
            'total_count': 0
        }
    
    # Count statistics
    for computer in computers_data:
        is_enabled = computer.get('Enabled', False)
        # Use explicit True comparison to handle both boolean and string values
        if is_enabled is True or is_enabled == 'True':
            stats['enabled_count'] += 1
        else:
            stats['disabled_count'] += 1
        
        # Check each category
        for cat in categories:
            is_match = False
            
            if cat == 'Unconstrained Delegation':
                is_match = computer.get('Delegation Type', '') == 'Unconstrained'
            elif cat == 'Constrained Delegation':
                is_match = computer.get('Delegation Type', '') == 'Constrained'
            elif cat == 'SIDHistory':
                is_match = bool(computer.get('SIDHistory', ''))
            elif cat.startswith('Dormant (>'):
                # Check for dormant field - it might have different keys based on config
                for key in computer.keys():
                    if 'Dormant' in key:
                        val = computer.get(key, False)
                        is_match = val is True or val == 'True'
                        break
            elif cat.startswith('Password Age (>'):
                pwd_age = computer.get('Password Age (days)', '')
                # Handle both int/float and string values
                try:
                    pwd_age_num = float(pwd_age) if pwd_age not in ('', None) else -1
                    is_match = pwd_age_num > password_age_days
                except (ValueError, TypeError):
                    is_match = False
            elif cat == 'ms-ds-CreatorSid':
                creator_sid = computer.get('ms-ds-CreatorSid', '')
                is_match = bool(creator_sid)
            elif cat == 'LAPS':
                # Check if computer has LAPS by comparing hostname
                hostname = computer.get('DNSHostName', '') or computer.get('Name', '')
                if hostname:
                    hostname_parts = hostname.split('.')
                    is_match = hostname_parts[0].lower() in laps_computers
            
            if is_match:
                stats['categories'][cat]['total_count'] += 1
                if is_enabled is True or is_enabled == 'True':
                    stats['categories'][cat]['enabled_count'] += 1
                else:
                    stats['categories'][cat]['disabled_count'] += 1
    
    return stats


@dataclass
class ADReconConfig:
    """Configuration for AD Reconnaissance."""
    domain_controller: str
    domain: str = ""
    username: str = ""
    password: str = ""
    auth_method: str = "NTLM"  # NTLM only for ADWS
    port: int = 9389  # ADWS default port
    page_size: int = 256
    dormant_days: int = 90
    password_age_days: int = 180
    only_enabled: bool = False
    output_dir: str = "."
    workstation: Optional[str] = None  # NTLM workstation name
    spn: Optional[str] = None  # Kerberos SPN override
    
    # Collection flags
    collect_forest: bool = True
    collect_domain: bool = True
    collect_trusts: bool = True
    collect_sites: bool = True
    collect_subnets: bool = True
    collect_schema: bool = True
    collect_password_policy: bool = True
    collect_fgpp: bool = True
    collect_dcs: bool = True
    collect_users: bool = True
    collect_user_spns: bool = True
    collect_groups: bool = True
    collect_group_members: bool = True
    collect_ous: bool = True
    collect_gpos: bool = True
    collect_gplinks: bool = True
    collect_dns_zones: bool = True
    collect_dns_records: bool = True
    collect_printers: bool = True
    collect_computers: bool = True
    collect_computer_spns: bool = True
    collect_laps: bool = True
    collect_bitlocker: bool = True
    collect_gmsa: bool = True
    collect_dmsa: bool = True
    collect_adcs: bool = True
    collect_protected_groups: bool = True
    collect_krbtgt: bool = True
    collect_asrep_roastable: bool = True
    collect_kerberoastable: bool = True


class PyADRecon:
    """Main AD Reconnaissance class using ADWS."""

    def __init__(self, config: ADReconConfig):
        self.config = config
        self.conn: Optional[ADWSConnection] = None
        self.base_dn: str = ""
        self.config_dn: str = ""
        self.schema_dn: str = ""
        self.forest_root_dn: str = ""
        self.domain_sid: str = ""
        self.results: Dict[str, List] = {}
        self.start_time: datetime = datetime.now()
        self._sid_cache: Dict[str, str] = {}  # Cache for SID-to-name resolution
        self.domain_sid_to_name: Dict[str, str] = {}  # Mapping of domain SIDs to domain names from trusts

    def connect(self) -> bool:
        """Establish ADWS connection."""
        try:
            logger.info(f"Attempting ADWS connection to {self.config.domain_controller}:{self.config.port}...")
            
            # Create ADWS server and connection
            server = ADWSServer(
                host=self.config.domain_controller,
                domain=self.config.domain
            )

            # Create connection with credentials
            self.conn = ADWSConnection(
                server=server,
                user=self.config.username,
                password=self.config.password,
                authentication=self.config.auth_method,
                spn=self.config.spn,
                page_size=self.config.page_size
            )

            # Bind (authenticate)
            if not self.conn.bind():
                logger.error("ADWS bind failed")
                if hasattr(self.conn, 'result'):
                    logger.error(f"Error: {self.conn.result.get('description', 'Unknown error')}")
                return False

            logger.info("[+] ADWS connection established successfully")

            # Get root DN from server info
            self._get_root_dse()

            return True

        except Exception as e:
            logger.error(f"Connection failed: {e}")
            traceback.print_exc()
            return False

    def _get_root_dse(self):
        """Get root DSE information to determine base DNs."""
        try:
            # The ADWS wrapper should have populated server.info
            if self.conn.server and hasattr(self.conn.server, 'info'):
                default_nc = self.conn.server.info.other.get('defaultNamingContext', [])
                if default_nc:
                    self.base_dn = default_nc[0] if isinstance(default_nc, list) else default_nc

                # Get Configuration DN from RootDSE (critical for child domains - it's forest-wide)
                config_nc = self.conn.server.info.other.get('configurationNamingContext', [])
                if config_nc:
                    self.config_dn = config_nc[0] if isinstance(config_nc, list) else config_nc

                # Get Schema DN from RootDSE (also forest-wide)
                schema_nc = self.conn.server.info.other.get('schemaNamingContext', [])
                if schema_nc:
                    self.schema_dn = schema_nc[0] if isinstance(schema_nc, list) else schema_nc
                    
            # Fallback: Derive from base DN if not provided by RootDSE
            if not self.base_dn and self.config.domain:
                # Construct from domain name
                parts = self.config.domain.split('.')
                self.base_dn = ','.join([f'DC={part}' for part in parts])
            
            # For child domains, derive forest root DN (Configuration/Schema are forest-wide)
            forest_root_dn = self.base_dn
            if self.base_dn:
                dn_parts = [part.strip() for part in self.base_dn.split(',') if part.strip().startswith('DC=')]
                if len(dn_parts) >= 3:  # Child domain detected
                    forest_root_dn = ','.join(dn_parts[1:])
                    logger.info(f"[*] Child domain detected - using forest root DN: {forest_root_dn}")

            # Store forest root DN for use in forest-wide queries
            self.forest_root_dn = forest_root_dn

            # Fallback for config/schema if RootDSE didn't provide them
            if not hasattr(self, 'config_dn') or not self.config_dn:
                self.config_dn = f"CN=Configuration,{forest_root_dn}"
            if not hasattr(self, 'schema_dn') or not self.schema_dn:
                self.schema_dn = f"CN=Schema,{self.config_dn}"
            
            logger.info(f"[*] Base DN: {self.base_dn}")
            logger.info(f"[*] Config DN: {self.config_dn}")
            
        except Exception as e:
            logger.warning(f"Could not fully parse Root DSE: {e}")
            # Try to continue anyway

    def search(self, search_base: str, search_filter: str, attributes: List[str] = None,
               scope: str = 'SUBTREE') -> List:
        """Perform ADWS search operation.
        
        Args:
            search_base: Base DN to search from
            search_filter: LDAP filter string
            attributes: List of attributes to retrieve
            scope: Search scope (SUBTREE, BASE, etc.)
            
        Returns:
            List of entries matching the search
        """
        try:
            # ADWS wrapper handles the search
            self.conn.search(
                search_base=search_base,
                search_filter=search_filter,
                attributes=attributes or ['*']
            )
            result_count = len(self.conn.entries) if hasattr(self.conn, 'entries') else 0
            if result_count == 0:
                logger.debug(f"Search returned 0 results: Base={search_base}, Filter={search_filter}")
            return self.conn.entries if hasattr(self.conn, 'entries') else []
        except Exception as e:
            logger.error(f"Search failed: {e}")
            logger.debug(f"Filter: {search_filter}, Base: {search_base}")
            return []

    def _resolve_sid_to_name(self, sid: str) -> str:
        """Resolve a SID to a readable name (user/group) with caching.
        Handles cross-domain SIDs in forest environments."""
        # Check cache first
        if sid in self._sid_cache:
            return self._sid_cache[sid]
        
        # Well-known SIDs
        well_known_sids = {
            'S-1-5-11': 'Authenticated Users',
            'S-1-5-32-544': 'BUILTIN\\Administrators',
            'S-1-5-32-545': 'BUILTIN\\Users',
            'S-1-5-32-546': 'BUILTIN\\Guests',
            'S-1-1-0': 'Everyone',
            'S-1-5-18': 'NT AUTHORITY\\SYSTEM',
            'S-1-5-7': 'ANONYMOUS LOGON',
            'S-1-5-9': 'Enterprise Domain Controllers',
            'S-1-5-32-548': 'Account Operators',
            'S-1-5-32-549': 'Server Operators',
            'S-1-5-32-550': 'Print Operators',
            'S-1-5-32-551': 'Backup Operators',
        }
        
        if sid in well_known_sids:
            result = well_known_sids[sid]
            self._sid_cache[sid] = result
            return result
        
        # Well-known RIDs (work with any domain SID)
        well_known_rids = {
            '498': 'Enterprise Read-only Domain Controllers',
            '500': 'Administrator',
            '501': 'Guest',
            '502': 'KRBTGT',
            '512': 'Domain Admins',
            '513': 'Domain Users',
            '514': 'Domain Guests',
            '515': 'Domain Computers',
            '516': 'Domain Controllers',
            '517': 'Cert Publishers',
            '518': 'Schema Admins',
            '519': 'Enterprise Admins',
            '520': 'Group Policy Creator Owners',
            '521': 'Read-only Domain Controllers',
            '522': 'Cloneable Domain Controllers',
            '525': 'Protected Users',
            '526': 'Key Admins',
            '527': 'Enterprise Key Admins',
            '553': 'RAS and IAS Servers',
            '571': 'Allowed RODC Password Replication Group',
            '572': 'Denied RODC Password Replication Group',
        }
        
        # Check if this is a domain SID with a well-known RID
        if sid.startswith('S-1-5-21-'):
            parts = sid.split('-')
            if len(parts) >= 8:
                rid = parts[-1]
                if rid in well_known_rids:
                    # Check if it's from a different domain
                    sid_domain = '-'.join(parts[:7])  # S-1-5-21-xxxxxxxxxx-xxxxxxxxxx-xxxxxxxxxx
                    current_domain = self.domain_sid if hasattr(self, 'domain_sid') and self.domain_sid else None
                    
                    if current_domain and sid_domain != current_domain:
                        # Cross-domain SID - try to resolve the domain name
                        domain_name = self.domain_sid_to_name.get(sid_domain, "Foreign Domain")
                        result = f"[{domain_name}] {well_known_rids[rid]}"
                    else:
                        result = well_known_rids[rid]
                    
                    self._sid_cache[sid] = result
                    return result
        
        # Try to resolve from current domain
        try:
            # Search for the object by objectSid
            entries = self.search(
                self.base_dn,
                f"(objectSid={sid})",
                ['sAMAccountName', 'cn', 'distinguishedName', 'objectClass']
            )
            
            if entries:
                entry = entries[0]
                sam_name = get_attr(entry, 'sAMAccountName', '')
                cn_name = get_attr(entry, 'cn', '')
                obj_classes = get_attr_list(entry, 'objectClass')
                
                # Determine object type - check computer first since computers also have 'user' class
                if 'computer' in obj_classes:
                    result = f"[Computer] {sam_name or cn_name}"
                elif 'group' in obj_classes:
                    result = f"[Group] {sam_name or cn_name}"
                elif 'user' in obj_classes:
                    result = f"[User] {sam_name or cn_name}"
                else:
                    result = sam_name or cn_name
                
                self._sid_cache[sid] = result
                return result
        except Exception as e:
            logger.debug(f"Could not resolve SID {sid} in current domain: {e}")
        
        # If not found in current domain and we're in a child domain, try forest root
        if hasattr(self, 'forest_root_dn') and self.forest_root_dn != self.base_dn:
            try:
                entries = self.search(
                    self.forest_root_dn,
                    f"(objectSid={sid})",
                    ['sAMAccountName', 'cn', 'distinguishedName', 'objectClass']
                )
                
                if entries:
                    entry = entries[0]
                    sam_name = get_attr(entry, 'sAMAccountName', '')
                    cn_name = get_attr(entry, 'cn', '')
                    obj_classes = get_attr_list(entry, 'objectClass')
                    
                    if 'computer' in obj_classes:
                        result = f"[Computer] {sam_name or cn_name}"
                    elif 'group' in obj_classes:
                        result = f"[Group] {sam_name or cn_name}"
                    elif 'user' in obj_classes:
                        result = f"[User] {sam_name or cn_name}"
                    else:
                        result = sam_name or cn_name
                    
                    self._sid_cache[sid] = result
                    return result
            except Exception as e:
                logger.debug(f"Could not resolve SID {sid} in forest root: {e}")
        
        # Before returning just the SID, check if it belongs to a known domain
        if sid.startswith('S-1-5-21-'):
            parts = sid.split('-')
            if len(parts) >= 8:
                # Extract domain SID (everything except the RID)
                sid_domain = '-'.join(parts[:7])
                # Check if this domain is in our mapping
                domain_name = self.domain_sid_to_name.get(sid_domain)
                if domain_name:
                    # Extract the RID for display
                    rid = parts[-1]
                    result = f"[{domain_name} SID] {sid} (RID: {rid})"
                    self._sid_cache[sid] = result
                    return result
        
        # Return SID if resolution failed
        result = f"[SID] {sid}"
        self._sid_cache[sid] = result
        return result

    def _parse_write_permissions(self, entry) -> list:
        """Parse nTSecurityDescriptor to get principals with dangerous write permissions.
        Returns: list of principals with WriteDacl, WriteOwner, GenericWrite, or GenericAll rights.
        """
        write_principals = []
        
        if not IMPACKET_AVAILABLE:
            return write_principals
        
        try:
            # Get raw security descriptor bytes
            sd_data = get_attr(entry, 'nTSecurityDescriptor')
            if not sd_data or not isinstance(sd_data, bytes):
                return write_principals
            
            # Parse the security descriptor
            sd = ldaptypes.SR_SECURITY_DESCRIPTOR(data=sd_data)
            
            if not sd['Dacl']:
                return write_principals
            
            # Access rights
            # WRITE_DACL = 0x00040000 - can modify DACL (dangerous)
            # WRITE_OWNER = 0x00080000 - can take ownership (dangerous)
            # GENERIC_WRITE = 0x40000000 - can write any property (dangerous)
            # GENERIC_ALL = 0x10000000 - full control (dangerous)
            # ADS_RIGHT_DS_WRITE_PROP = 0x00000020 - property write (only dangerous if not object-specific)
            
            for ace in sd['Dacl'].aces:
                try:
                    sid = ace['Ace']['Sid'].formatCanonical()
                    ace_type = ace['AceType']
                    
                    # Only check ALLOW ACEs (0 = standard, 5 = object-specific)
                    if ace_type in [0, 5]:
                        mask = ace['Ace']['Mask']['Mask']
                        
                        # Check for dangerous write permissions
                        has_dangerous_write = bool(mask & (0x00040000 | 0x00080000 | 0x40000000 | 0x10000000))
                        
                        # For WriteProperty (0x00000020), only flag if it's NOT object-specific
                        has_generic_write_property = False
                        if mask & 0x00000020:  # ADS_RIGHT_DS_WRITE_PROP
                            if ace_type == 5:  # ACCESS_ALLOWED_OBJECT_ACE_TYPE
                                try:
                                    ace_data = ace['Ace']
                                    # Check if ObjectType is present (Flags is a bitfield)
                                    # Try to access Flags field directly
                                    try:
                                        flags = ace_data['Flags']
                                    except (KeyError, TypeError):
                                        flags = 0
                                    
                                    # If no ObjectType GUID (flags & 0x1 == 0), it applies to ALL properties
                                    if not (flags & 0x1):
                                        has_generic_write_property = True
                                except (KeyError, TypeError):
                                    # If we can't determine, assume it's generic (safer for detection)
                                    has_generic_write_property = True
                            else:
                                # Standard ACE with write property = applies to all properties
                                has_generic_write_property = True
                        
                        if has_dangerous_write or has_generic_write_property:
                            principal_name = self._resolve_sid_to_name(sid)
                            if principal_name not in write_principals:
                                write_principals.append(principal_name)
                                
                except Exception as e:
                    logger.debug(f"Error parsing ACE for write perms: {e}")
                    continue
                    
        except Exception as e:
            logger.debug(f"Error parsing write permissions: {e}")
            
        return write_principals

    def _is_well_known_high_privilege_principal(self, principal_name: str) -> bool:
        """Check if a principal is a well-known high-privilege group/user that should be filtered from LAPS readers.
        Returns True for default admin groups that are expected to have access.
        """
        principal_lower = principal_name.lower()
        
        # Well-known high-privilege groups and accounts that are expected to have LAPS access
        well_known_high_priv = {
            'nt authority\\system',
            'builtin\\administrators',
            '[group] domain admins',
            '[group] enterprise admins',
            '[group] administrators',
            'account operators',
            'backup operators',
            '[group] account operators',
            '[group] backup operators',
            '[group] server operators',
            'server operators',
            '[group] domain controllers',
            'enterprise domain controllers',
        }
        
        return principal_lower in well_known_high_priv
    
    def _get_schema_guid(self, attribute_name: str) -> Optional[str]:
        """Get the schemaIDGUID for a specific attribute.
        Returns: GUID string in canonical format (e.g., 'a740f3f1-...').
        """
        try:
            logger.debug(f"Looking up schema GUID for {attribute_name} in {self.schema_dn}")
            entries = self.search(
                search_base=self.schema_dn,
                search_filter=f"(name={attribute_name})",
                attributes=['schemaIDGUID']
            )
            
            logger.debug(f"Schema search returned {len(entries) if entries else 0} entries")
            
            if entries and len(entries) > 0:
                guid_attr = get_attr(entries[0], 'schemaIDGUID')
                logger.debug(f"Got schemaIDGUID: {guid_attr}, type: {type(guid_attr)}")
                
                # ADWS returns binary data as base64-encoded strings
                if guid_attr:
                    import base64
                    import uuid
                    
                    guid_bytes = None
                    if isinstance(guid_attr, str):
                        # Decode base64 string to bytes
                        try:
                            guid_bytes = base64.b64decode(guid_attr)
                            logger.debug(f"Decoded base64 to {len(guid_bytes)} bytes")
                        except Exception as e:
                            logger.debug(f"Failed to decode base64: {e}")
                    elif isinstance(guid_attr, bytes):
                        guid_bytes = guid_attr
                    
                    if guid_bytes and len(guid_bytes) == 16:
                        # AD stores schemaIDGUID in little-endian format
                        guid_str = str(uuid.UUID(bytes_le=guid_bytes)).lower()
                        logger.debug(f"Schema GUID for {attribute_name}: {guid_str}")
                        return guid_str
                    else:
                        logger.debug(f"schemaIDGUID not 16 bytes (got {len(guid_bytes) if guid_bytes else 0})")
                else:
                    logger.debug(f"schemaIDGUID attribute is empty for {attribute_name}")
        except Exception as e:
            logger.debug(f"Could not get schema GUID for {attribute_name}: {e}")
            import traceback
            logger.debug(traceback.format_exc())
        
        return None
    
    def _guid_bytes_to_str(self, guid_bytes: bytes) -> str:
        """Convert GUID bytes to canonical string format."""
        import uuid
        return str(uuid.UUID(bytes_le=guid_bytes)).lower()
    
    def _mask_to_rights(self, mask: int) -> list:
        """Decode access mask to list of right names."""
        ADS_RIGHT_DS_READ_PROP = 0x00000010
        ADS_RIGHT_DS_CONTROL_ACCESS = 0x00000100
        GENERIC_READ = 0x80000000
        GENERIC_ALL = 0x10000000
        WRITE_DAC = 0x00040000
        WRITE_OWNER = 0x00080000
        
        rights = []
        if mask & ADS_RIGHT_DS_READ_PROP:
            rights.append("READ_PROP")
        if mask & ADS_RIGHT_DS_CONTROL_ACCESS:
            rights.append("CONTROL_ACCESS")
        if mask & GENERIC_READ:
            rights.append("GENERIC_READ")
        if mask & GENERIC_ALL:
            rights.append("GENERIC_ALL")
        if mask & WRITE_DAC:
            rights.append("WRITE_DAC")
        if mask & WRITE_OWNER:
            rights.append("WRITE_OWNER")
        return rights
    
    def _parse_laps_readers(self, entry, laps_attr_guid: Optional[str] = None) -> list:
        """Parse nTSecurityDescriptor to get principals with read access to ms-Mcs-AdmPwd attribute.
        
        Args:
            entry: The ADWS entry (computer object)
            laps_attr_guid: The schemaIDGUID for ms-Mcs-AdmPwd (forest-specific)
        
        Returns: list of dicts with principal, SID, rights, and why information.
        """
        results = []
        
        if not IMPACKET_AVAILABLE:
            logger.debug("Impacket not available for ACL parsing")
            return results
        
        # Get raw security descriptor bytes
        sd_attr = get_attr(entry, 'nTSecurityDescriptor')
        if not sd_attr:
            logger.debug("No security descriptor data found for LAPS reader parsing")
            return results
        
        # ADWS may return binary data as base64-encoded string
        sd_data = None
        if isinstance(sd_attr, str):
            try:
                import base64
                sd_data = base64.b64decode(sd_attr)
                logger.debug(f"Decoded base64 nTSecurityDescriptor to {len(sd_data)} bytes")
            except Exception as e:
                logger.debug(f"Failed to decode base64 nTSecurityDescriptor: {e}")
                return results
        elif isinstance(sd_attr, (bytes, bytearray)):
            sd_data = sd_attr
        else:
            logger.debug(f"Unexpected nTSecurityDescriptor type: {type(sd_attr)}")
            return results
        
        if not sd_data:
            logger.debug("No security descriptor data after decoding")
            return results
        
        # If we don't have the LAPS GUID, we can't do proper attribute-specific checks
        if not laps_attr_guid:
            logger.debug("No LAPS attribute GUID provided, cannot perform accurate checks")
            return results
        
        try:
            # Parse the security descriptor using impacket
            sd = ldaptypes.SR_SECURITY_DESCRIPTOR(data=sd_data)
            
            # Check if DACL exists
            if not sd['Dacl']:
                logger.debug("No DACL in security descriptor")
                return results
            
            # Access right constants
            ADS_RIGHT_DS_CONTROL_ACCESS = 0x00000100
            GENERIC_ALL = 0x10000000
            WRITE_DAC = 0x00040000
            WRITE_OWNER = 0x00080000
            
            # ACE type constants
            ACCESS_ALLOWED_ACE = 0
            ACCESS_ALLOWED_OBJECT_ACE = 5
            ACE_OBJECT_TYPE_PRESENT = 0x1
            
            # Track seen principals to avoid duplicates
            seen = set()
            
            # Parse each ACE in the DACL
            for ace in sd['Dacl'].aces:
                try:
                    ace_type = ace['AceType']
                    ace_body = ace['Ace']
                    sid = ace_body['Sid'].formatCanonical()
                    mask = ace_body['Mask']['Mask']
                    try:
                        inherited = bool(ace['AceFlags'] & 0x10) if 'AceFlags' in ace.fields else False
                    except:
                        inherited = False
                    
                    # Quick wins: if someone has GENERIC_ALL / WRITE_DAC / WRITE_OWNER on the computer object,
                    # they can effectively get the LAPS value (or grant themselves rights)
                    if mask & (GENERIC_ALL | WRITE_DAC | WRITE_OWNER):
                        principal = self._resolve_sid_to_name(sid)
                        key = (principal, sid, "FULL_CONTROL_PATH")
                        if key not in seen:
                            seen.add(key)
                            rights_list = self._mask_to_rights(mask)
                            results.append({
                                "principal": principal,
                                "sid": sid,
                                "rights": rights_list,
                                "ace_type": ace_type,
                                "object_type": None,
                                "inherited": inherited,
                                "why": f"High privilege on computer ({', '.join(rights_list)})"
                            })
                        continue
                    
                    # Standard ACE (applies to whole object)
                    if ace_type == ACCESS_ALLOWED_ACE:
                        # For confidential attributes, CONTROL_ACCESS is key
                        if mask & ADS_RIGHT_DS_CONTROL_ACCESS:
                            principal = self._resolve_sid_to_name(sid)
                            key = (principal, sid, "CTRL_ACCESS_ALL")
                            if key not in seen:
                                seen.add(key)
                                rights_list = self._mask_to_rights(mask)
                                results.append({
                                    "principal": principal,
                                    "sid": sid,
                                    "rights": rights_list,
                                    "ace_type": ace_type,
                                    "object_type": None,
                                    "inherited": inherited,
                                    "why": "CONTROL_ACCESS on object (confidential attributes readable)"
                                })
                    
                    # Object-specific ACE
                    elif ace_type == ACCESS_ALLOWED_OBJECT_ACE:
                        try:
                            flags = int(ace_body['Flags']) if 'Flags' in ace_body.fields else 0
                        except:
                            flags = 0
                        object_type = None
                        
                        # Get the ObjectType GUID if present
                        if (flags & ACE_OBJECT_TYPE_PRESENT) and ('ObjectType' in ace_body.fields):
                            try:
                                ot = ace_body['ObjectType']
                                if isinstance(ot, bytes) and len(ot) == 16:
                                    object_type = self._guid_bytes_to_str(ot)
                                else:
                                    object_type = str(ot).lower()
                            except:
                                pass
                        
                        # Only count if ACE is specifically for ms-Mcs-AdmPwd (or if object_type absent -> applies to all)
                        applies_to_laps = (object_type is None) or (object_type == laps_attr_guid)
                        
                        if applies_to_laps and (mask & ADS_RIGHT_DS_CONTROL_ACCESS):
                            principal = self._resolve_sid_to_name(sid)
                            key = (principal, sid, object_type, mask)
                            if key not in seen:
                                seen.add(key)
                                rights_list = self._mask_to_rights(mask)
                                results.append({
                                    "principal": principal,
                                    "sid": sid,
                                    "rights": rights_list,
                                    "ace_type": ace_type,
                                    "object_type": object_type,
                                    "inherited": inherited,
                                    "why": f"Object-specific CONTROL_ACCESS for ms-Mcs-AdmPwd (GUID={object_type or 'all'})"
                                })
                                
                except Exception as e:
                    logger.debug(f"Error parsing ACE for LAPS readers: {e}")
                    continue
                    
        except Exception as e:
            logger.debug(f"Error parsing LAPS readers: {e}")
            import traceback
            logger.debug(traceback.format_exc())
            
        return results

    def _parse_enrollment_rights(self, entry) -> tuple:
        """Parse nTSecurityDescriptor to get principals with enrollment rights.
        Returns: (enrollment_principals list, auto_enrollment_principals list, owner_principal string)
        """
        enrollment_principals = []
        auto_enrollment_principals = []
        owner_principal = None
        
        if not IMPACKET_AVAILABLE:
            logger.debug("Impacket not available for ACL parsing")
            return (enrollment_principals, auto_enrollment_principals, owner_principal)
        
        try:
            # Certificate-Enrollment extended right GUID
            CERT_ENROLLMENT_GUID = "0e10c968-78fb-11d2-90d4-00c04f79dc55"
            # Certificate-AutoEnrollment extended right GUID  
            CERT_AUTO_ENROLLMENT_GUID = "a05b8cc2-17bc-4802-a710-e7c15ab866a2"
            
            # Get raw security descriptor bytes
            sd_data = get_attr(entry, 'nTSecurityDescriptor')
            if not sd_data or not isinstance(sd_data, bytes):
                logger.debug("No security descriptor data found or wrong type")
                return (enrollment_principals, auto_enrollment_principals, owner_principal)
            
            # Parse the security descriptor using impacket
            sd = ldaptypes.SR_SECURITY_DESCRIPTOR(data=sd_data)
            
            # Parse owner SID
            try:
                if sd['OwnerSid']:
                    owner_sid = sd['OwnerSid'].formatCanonical()
                    logger.debug(f"Owner SID: {owner_sid}")
                    owner_principal = self._resolve_sid_to_name(owner_sid)
                    logger.debug(f"Owner resolved to: {owner_principal}")
                else:
                    logger.debug("No OwnerSid in security descriptor")
            except Exception as e:
                logger.warning(f"Error parsing owner SID: {e}")
                import traceback
                logger.debug(traceback.format_exc())
            
            # Check if DACL exists
            if not sd['Dacl']:
                logger.debug("No DACL in security descriptor")
                return (enrollment_principals, auto_enrollment_principals, owner_principal)
            
            # Parse each ACE in the DACL
            logger.debug(f"Parsing {len(sd['Dacl'].aces)} ACEs from security descriptor")
            for ace in sd['Dacl'].aces:
                try:
                    # Get SID from ACE
                    sid = ace['Ace']['Sid'].formatCanonical()
                    logger.debug(f"Processing ACE with SID: {sid}")
                    
                    # Resolve SID to name
                    principal_name = self._resolve_sid_to_name(sid)
                    logger.debug(f"  Resolved to: {principal_name}")
                    
                    # Check if this is an ACCESS_ALLOWED_OBJECT_ACE
                    ace_type = ace['AceType']
                    
                    # ACCESS_ALLOWED_OBJECT_ACE_TYPE = 5
                    # ACCESS_ALLOWED_ACE_TYPE = 0
                    if ace_type in [0, 5]:  # Standard or Object ACE
                        mask = ace['Ace']['Mask']['Mask']
                        
                        # Check for extended right (0x100 = ADS_RIGHT_DS_CONTROL_ACCESS)
                        if mask & 0x100:
                            # For object ACEs, check the ObjectType GUID
                            if ace_type == 5:
                                # Get ObjectType GUID if present
                                ace_data = ace['Ace']
                                
                                # Check if Flags field exists and has ACE_OBJECT_TYPE_PRESENT
                                try:
                                    flags = ace_data['Flags']
                                except (KeyError, TypeError):
                                    flags = 0
                                
                                # ACE_OBJECT_TYPE_PRESENT = 0x1
                                if flags & 0x1:
                                    try:
                                        obj_type = ace_data['ObjectType']
                                        # Convert bytes to GUID string
                                        if isinstance(obj_type, bytes) and len(obj_type) == 16:
                                            import uuid
                                            # Parse as little-endian GUID
                                            obj_type_str = str(uuid.UUID(bytes_le=obj_type)).lower()
                                        else:
                                            obj_type_str = str(obj_type).lower()
                                        
                                        # Check if it's the enrollment GUID
                                        if obj_type_str == CERT_ENROLLMENT_GUID.lower():
                                            if principal_name not in enrollment_principals:
                                                enrollment_principals.append(principal_name)
                                        
                                        # Check if it's the auto-enrollment GUID
                                        if obj_type_str == CERT_AUTO_ENROLLMENT_GUID.lower():
                                            if principal_name not in auto_enrollment_principals:
                                                auto_enrollment_principals.append(principal_name)
                                    except (KeyError, TypeError, ValueError) as e:
                                        logger.debug(f"Could not get ObjectType: {e}")
                                else:
                                    # No ObjectType means applies to all extended rights
                                    if principal_name not in enrollment_principals:
                                        enrollment_principals.append(principal_name)
                            else:
                                # Standard ACE with extended right - applies to all
                                if principal_name not in enrollment_principals:
                                    enrollment_principals.append(principal_name)
                                    
                except Exception as e:
                    logger.debug(f"Error parsing ACE: {e}")
                    continue
                    
        except Exception as e:
            logger.debug(f"Error parsing security descriptor: {e}")
            import traceback
            logger.debug(traceback.format_exc())
        
        logger.debug(f"Found {len(enrollment_principals)} enrollment principals: {enrollment_principals}")
        logger.debug(f"Found {len(auto_enrollment_principals)} auto-enrollment principals: {auto_enrollment_principals}")
            
        return (enrollment_principals, auto_enrollment_principals, owner_principal)

    def _get_low_privilege_sids(self) -> set:
        """Get a set of well-known low-privilege SIDs including domain-specific ones."""
        low_priv_sids = {
            'S-1-1-0',       # Everyone
            'S-1-5-11',      # Authenticated Users
            'S-1-5-7',       # Anonymous Logon
            'S-1-5-32-545',  # BUILTIN\Users
            'S-1-5-32-546',  # BUILTIN\Guests
        }
        
        # Add domain-specific SIDs
        try:
            # Get domain SID - we already have it from domain info collection
            if hasattr(self, 'domain_sid') and self.domain_sid:
                domain_sid = self.domain_sid
                # Domain Users: <domain-sid>-513
                low_priv_sids.add(f"{domain_sid}-513")
                # Domain Computers: <domain-sid>-515
                low_priv_sids.add(f"{domain_sid}-515")
                # Domain Guests: <domain-sid>-514
                low_priv_sids.add(f"{domain_sid}-514")
        except Exception as e:
            logger.debug(f"Could not get domain SID for low-priv detection: {e}")
        
        return low_priv_sids

    def _is_low_privilege_principal(self, principal_name: str) -> bool:
        """Check if a principal is low-privileged based on SID resolution."""
        # Extract SID from principal name if it's in [SID] or [domain SID] format
        if '[SID]' in principal_name or 'SID]' in principal_name:
            # Extract SID from "[SID] S-1-..." or "[domain.name SID] S-1-..."
            sid_match = re.search(r'S-1-[0-9-]+', principal_name)
            if sid_match:
                sid = sid_match.group(0)
                low_priv_sids = self._get_low_privilege_sids()
                return sid in low_priv_sids
        
        # For resolved names, check against well-known low-privilege group names (exact match)
        principal_lower = principal_name.lower()
        low_priv_names = {
            'everyone',
            'authenticated users',
            'anonymous logon',
            'builtin\\users',
            'builtin\\guests',
            '[group] domain users',
            '[group] domain computers',
            '[group] domain guests',
        }
        
        if principal_lower in low_priv_names:
            return True
        
        # Check if this is an individual user account (not a well-known admin)
        # Individual users who are not privileged accounts should be considered low-priv for ESC detection
        if '[user]' in principal_lower:
            # Exclude well-known privileged user accounts
            admin_users = {
                '[user] administrator',
                '[user] krbtgt',
                '[user] guest'
            }
            if principal_lower not in admin_users:
                # This is a regular user account, consider it low-privilege
                return True
        
        return False

    def _initialize_sid_mappings(self):
        """Initialize SID to domain name mappings for cross-domain SID resolution.
        This runs silently before collections to populate domain_sid_to_name cache.
        """
        try:
            # Get current domain SID and name
            entries = self.search(
                search_base=self.base_dn,
                search_filter='(objectCategory=domainDNS)',
                attributes=['objectSid']
            )
            
            if entries:
                sid_bytes = get_attr(entries[0], 'objectSid')
                if sid_bytes:
                    self.domain_sid = sid_to_str(sid_bytes)
                    domain_name = dn_to_fqdn(self.base_dn)
                    self.domain_sid_to_name[self.domain_sid] = domain_name
                    logger.debug(f"Initialized current domain SID: {self.domain_sid} -> {domain_name}")
            
            # Get trusted domain SIDs and names
            trust_entries = self.search(
                search_base=self.base_dn,
                search_filter='(objectCategory=trustedDomain)',
                attributes=['trustPartner', 'securityIdentifier']
            )
            
            for entry in trust_entries:
                trusted_domain_sid_bytes = get_attr(entry, 'securityIdentifier')
                trusted_domain_name = get_attr(entry, 'trustPartner', '')
                
                if trusted_domain_sid_bytes and trusted_domain_name:
                    try:
                        trusted_domain_sid = sid_to_str(trusted_domain_sid_bytes)
                        self.domain_sid_to_name[trusted_domain_sid] = trusted_domain_name
                        logger.debug(f"Initialized trusted domain SID: {trusted_domain_sid} -> {trusted_domain_name}")
                    except Exception:
                        pass
                        
        except Exception as e:
            logger.debug(f"Error initializing SID mappings: {e}")
            # Non-fatal - continue even if this fails

    def collect_domain_info(self) -> List[Dict]:
        """Collect domain information."""
        logger.info("[-] Collecting Domain Information...")
        results = []

        try:
            # Query domain object
            # Using objectCategory=domainDNS instead of objectClass=domain
            entries = self.search(
                search_base=self.base_dn,
                search_filter='(objectCategory=domainDNS)',
                attributes=['msDS-Behavior-Version', 'objectSid', 'name', 'whenCreated', 'ms-DS-MachineAccountQuota', 'distinguishedName']
            )

            if entries:
                entry = entries[0]

                # Get domain functional level
                fl = get_attr(entry, 'msDS-Behavior-Version', 0)
                func_level = DOMAIN_FUNCTIONAL_LEVELS.get(safe_int(fl), "Unknown")

                # Get domain SID
                sid_bytes = get_attr(entry, 'objectSid')
                self.domain_sid = sid_to_str(sid_bytes)
                
                # Map current domain SID to its name for cross-domain resolution
                domain_name = dn_to_fqdn(self.base_dn)
                self.domain_sid_to_name[self.domain_sid] = domain_name

                results.append({"Category": "Name", "Value": domain_name})
                results.append({"Category": "NetBIOS", "Value": get_attr(entry, 'name', '')})
                results.append({"Category": "Functional Level", "Value": f"{func_level}Domain"})
                results.append({"Category": "DomainSID", "Value": self.domain_sid})
                results.append({"Category": "Creation Date", "Value": format_datetime(self._parse_timestamp(get_attr(entry, 'whenCreated')))})
                
                # Machine Account Quota - how many computers each user can add
                maq = get_attr(entry, 'ms-DS-MachineAccountQuota', '10')
                results.append({"Category": "ms-DS-MachineAccountQuota", "Value": str(maq)})

                # Get RID info
                rid_entries = self.search(
                    search_base=f"CN=RID Manager$,CN=System,{self.base_dn}",
                    search_filter='(objectClass=*)',
                    attributes=['rIDAvailablePool'],
                    scope='BASE'
                )
                if rid_entries:
                    rid_pool = get_attr(rid_entries[0], 'rIDAvailablePool')
                    if rid_pool:
                        rid_pool = safe_int(rid_pool)
                        total_sids = rid_pool >> 32
                        issued = rid_pool & 0xFFFFFFFF
                        remaining = total_sids - issued
                        results.append({"Category": "RIDs Issued", "Value": str(issued)})
                        results.append({"Category": "RIDs Remaining", "Value": str(remaining)})

        except Exception as e:
            logger.warning(f"Error collecting domain info: {e}")

        self.results['Domain'] = results
        logger.info(f"    Found {len(results)} domain properties")
        return results

    def collect_forest_info(self) -> List[Dict]:
        """Collect forest information."""
        logger.info("[-] Collecting Forest Information...")
        results = []

        try:
            # Get forest info from configuration partition
            entries = self.search(
                search_base=f"CN=Partitions,{self.config_dn}",
                search_filter='(objectClass=crossRefContainer)',
                attributes=['msDS-Behavior-Version', 'fSMORoleOwner', 'msDS-EnabledFeature', 'distinguishedName']
            )

            # Use forest root DN (not current domain's base_dn) for forest name
            forest_name = dn_to_fqdn(self.forest_root_dn if hasattr(self, 'forest_root_dn') else self.base_dn)
            forest_fl = 0
            
            if entries:
                entry = entries[0]
                forest_fl = get_attr(entry, 'msDS-Behavior-Version', 0)
                func_level = DOMAIN_FUNCTIONAL_LEVELS.get(safe_int(forest_fl), "Unknown")

                results.append({"Category": "Name", "Value": forest_name})
                results.append({"Category": "Functional Level", "Value": f"{func_level}Forest"})

            # Get FSMO role holders (we'll get these from Domain Controllers collection)
            # Domain Naming Master
            try:
                naming_entries = self.search(
                    search_base=f"CN=Partitions,{self.config_dn}",
                    search_filter='(objectClass=crossRefContainer)',
                    attributes=['fSMORoleOwner']
                )
                if naming_entries:
                    naming_owner = get_attr(naming_entries[0], 'fSMORoleOwner', '')
                    if naming_owner:
                        # Extract DC FQDN from DN: CN=NTDS Settings,CN=DC1,CN=Servers,CN=Site,... -> DC1.domain.local
                        # The DN contains domain components that tell us the actual domain
                        parts = str(naming_owner).split(',')
                        if len(parts) > 1:
                            dc_name = parts[1].replace('CN=', '').strip()
                            # Extract domain DN from the rest of the path
                            domain_dn_parts = [p for p in parts if p.strip().upper().startswith('DC=')]
                            if domain_dn_parts:
                                domain_fqdn = dn_to_fqdn(','.join(domain_dn_parts))
                                results.append({"Category": "Domain Naming Master", "Value": f"{dc_name.lower()}.{domain_fqdn}"})
                            else:
                                results.append({"Category": "Domain Naming Master", "Value": f"{dc_name.lower()}.{forest_name}"})
            except:
                pass
            
            # Schema Master
            try:
                schema_entries = self.search(
                    search_base=self.schema_dn,
                    search_filter='(objectClass=dMD)',
                    attributes=['fSMORoleOwner']
                )
                if schema_entries:
                    schema_owner = get_attr(schema_entries[0], 'fSMORoleOwner', '')
                    if schema_owner:
                        parts = str(schema_owner).split(',')
                        if len(parts) > 1:
                            dc_name = parts[1].replace('CN=', '').strip()
                            # Extract domain DN from the DN path
                            domain_dn_parts = [p for p in parts if p.strip().upper().startswith('DC=')]
                            if domain_dn_parts:
                                domain_fqdn = dn_to_fqdn(','.join(domain_dn_parts))
                                results.append({"Category": "Schema Master", "Value": f"{dc_name.lower()}.{domain_fqdn}"})
                            else:
                                results.append({"Category": "Schema Master", "Value": f"{dc_name.lower()}.{forest_name}"})
            except:
                pass

            # RootDomain
            results.append({"Category": "RootDomain", "Value": forest_name})

            # Domain Count - count cross-ref objects
            try:
                domain_entries = self.search(
                    search_base=f"CN=Partitions,{self.config_dn}",
                    search_filter="(&(objectClass=crossRef)(systemFlags:1.2.840.113556.1.4.803:=3))",
                    attributes=['nCName']
                )
                results.append({"Category": "Domain Count", "Value": str(len(domain_entries))})
            except:
                results.append({"Category": "Domain Count", "Value": "1"})

            # Site Count
            try:
                site_entries = self.search(
                    search_base=f"CN=Sites,{self.config_dn}",
                    search_filter='(objectCategory=site)',
                    attributes=['name']
                )
                results.append({"Category": "Site Count", "Value": str(len(site_entries))})
            except:
                results.append({"Category": "Site Count", "Value": "0"})

            # Global Catalog Count
            try:
                gc_entries = self.search(
                    search_base=f"CN=Sites,{self.config_dn}",
                    search_filter="(&(objectCategory=nTDSDSA)(options:1.2.840.113556.1.4.803:=1))",
                    attributes=['distinguishedName']
                )
                results.append({"Category": "Global Catalog Count", "Value": str(len(gc_entries))})
            except:
                results.append({"Category": "Global Catalog Count", "Value": "0"})

            # Domain (same as forest root)
            results.append({"Category": "Domain", "Value": forest_name})

            # Site - get first site
            try:
                site_entries = self.search(
                    search_base=f"CN=Sites,{self.config_dn}",
                    search_filter='(objectCategory=site)',
                    attributes=['name']
                )
                if site_entries:
                    results.append({"Category": "Site", "Value": get_attr(site_entries[0], 'name', '')})
            except:
                pass

            # GlobalCatalog - get first GC server
            try:
                dc_entries = self.search(
                    search_base=self.base_dn,
                    search_filter="(&(objectCategory=computer)(userAccountControl:1.2.840.113556.1.4.803:=8192))",
                    attributes=['dNSHostName']
                )
                if dc_entries:
                    results.append({"Category": "GlobalCatalog", "Value": get_attr(dc_entries[0], 'dNSHostName', '')})
            except:
                pass

            # Tombstone Lifetime
            try:
                dir_service_entries = self.search(
                    search_base=f"CN=Directory Service,CN=Windows NT,CN=Services,{self.config_dn}",
                    search_filter='(objectClass=*)',
                    attributes=['tombstoneLifetime'],
                    scope='BASE'
                )
                if dir_service_entries:
                    tombstone = get_attr(dir_service_entries[0], 'tombstoneLifetime', 180)
                    results.append({"Category": "Tombstone Lifetime", "Value": str(tombstone)})
                else:
                    results.append({"Category": "Tombstone Lifetime", "Value": "180"})
            except:
                results.append({"Category": "Tombstone Lifetime", "Value": "180"})

            # Recycle Bin (2008 R2 onwards) - check if msDS-EnabledFeature exists
            try:
                recycle_entries = self.search(
                    search_base=f"CN=Partitions,{self.config_dn}",
                    search_filter='(objectClass=crossRefContainer)',
                    attributes=['msDS-EnabledFeature']
                )
                recycle_enabled = False
                if recycle_entries:
                    enabled_features = get_attr_list(recycle_entries[0], 'msDS-EnabledFeature')
                    for feature in enabled_features:
                        if 'Recycle Bin Feature' in str(feature):
                            recycle_enabled = True
                            break
                results.append({"Category": "Recycle Bin (2008 R2 onwards)", "Value": "Enabled" if recycle_enabled else "Disabled"})
            except:
                results.append({"Category": "Recycle Bin (2008 R2 onwards)", "Value": "Disabled"})

            # Privileged Access Management (2016 onwards)
            try:
                pam_entries = self.search(
                    search_base=f"CN=Partitions,{self.config_dn}",
                    search_filter='(objectClass=crossRefContainer)',
                    attributes=['msDS-EnabledFeature']
                )
                pam_enabled = False
                if pam_entries:
                    enabled_features = get_attr_list(pam_entries[0], 'msDS-EnabledFeature')
                    for feature in enabled_features:
                        if 'Privileged Access Management' in str(feature) or 'PAM' in str(feature):
                            pam_enabled = True
                            break
                results.append({"Category": "Privileged Access Management (2016 onwards)", "Value": "Enabled" if pam_enabled else "Disabled"})
            except:
                results.append({"Category": "Privileged Access Management (2016 onwards)", "Value": "Disabled"})

            # Check for LAPS
            laps_entries = self.search(
                search_base=self.schema_dn,
                search_filter="(name=ms-Mcs-AdmPwd)",
                attributes=['whenCreated']
            )
            if laps_entries:
                results.append({"Category": "LAPS", "Value": "Enabled"})
                results.append({"Category": "LAPS Installed Date", "Value": format_datetime(self._parse_timestamp(get_attr(laps_entries[0], 'whenCreated')))})
            else:
                results.append({"Category": "LAPS", "Value": "Not Installed"})

        except Exception as e:
            logger.warning(f"Error collecting forest info: {e}")

        self.results['Forest'] = results
        logger.info(f"    Found {len(results)} forest properties")
        return results

    def collect_users(self) -> List[Dict]:
        """Collect user objects."""
        logger.info("[-] Collecting Users - May take some time...")
        results = []
        now = datetime.now(timezone.utc).replace(tzinfo=None)

        try:
            # Using (&(objectCategory=person)(objectClass=user)) instead of samAccountType
            filter_str = "(&(objectCategory=person)(objectClass=user))"
            if self.config.only_enabled:
                filter_str = "(&(objectCategory=person)(objectClass=user)(!(userAccountControl:1.2.840.113556.1.4.803:=2)))"

            entries = self.search(
                search_base=self.base_dn,
                search_filter=filter_str,
                attributes=['sAMAccountName', 'name', 'distinguishedName', 'canonicalName',
                           'userAccountControl', 'pwdLastSet', 'lastLogonTimestamp',
                           'accountExpires', 'adminCount', 'description', 'title',
                           'department', 'company', 'manager', 'mail', 'mobile',
                           'homeDirectory', 'profilePath', 'scriptPath', 'memberOf',
                           'primaryGroupID', 'objectSid', 'sIDHistory', 'servicePrincipalName',
                           'msDS-AllowedToDelegateTo', 'msDS-SupportedEncryptionTypes',
                           'givenName', 'sn', 'middleName', 'c', 'info', 'userWorkstations',
                           'whenCreated', 'whenChanged', 'ntSecurityDescriptor']
            )

            for entry in entries:
                
                uac = get_attr(entry, 'userAccountControl', 0)
                uac_parsed = parse_uac(uac)

                # Check "Cannot Change Password" via ACL (preferred) or UAC flag (fallback)
                cannot_change_pwd = check_cannot_change_password(entry)
                if not cannot_change_pwd:
                    # Fallback to UAC flag if ACL check didn't find it
                    cannot_change_pwd = uac_parsed.get('PasswordCantChange', False)

                # Password last set
                pwd_last_set = self._parse_timestamp(get_attr(entry, 'pwdLastSet'))
                pwd_last_set_dt = None
                pwd_age_days = None
                must_change_pwd = False
                pwd_not_changed_max = False
                
                if isinstance(pwd_last_set, datetime):
                    # Check if pwdLastSet is 0 (1601-01-01 epoch) which means "must change password at logon"
                    if pwd_last_set.year == 1601:
                        must_change_pwd = True
                        pwd_last_set_dt = None  # Don't show the 1601 date
                    else:
                        pwd_last_set_dt = pwd_last_set
                        # Calculate age in UTC
                        pwd_last_set_utc = pwd_last_set.replace(tzinfo=None) if pwd_last_set.tzinfo else pwd_last_set
                        pwd_age_days = (now - pwd_last_set_utc).days
                        if pwd_age_days > self.config.password_age_days:
                            pwd_not_changed_max = True
                else:
                    # If not a datetime, user must change password at next logon
                    must_change_pwd = True

                # Last logon
                last_logon = self._parse_timestamp(get_attr(entry, 'lastLogonTimestamp'))
                last_logon_dt = None
                logon_age_days = None
                never_logged_in = True
                dormant = False

                if isinstance(last_logon, datetime):
                    # Check if lastLogonTimestamp is 0 (1601-01-01 epoch) which means never logged in
                    if last_logon.year == 1601:
                        last_logon_dt = None
                        never_logged_in = True
                    else:
                        last_logon_dt = last_logon
                        never_logged_in = False
                        # Calculate age in UTC
                        last_logon_utc = last_logon.replace(tzinfo=None) if last_logon.tzinfo else last_logon
                        logon_age_days = (now - last_logon_utc).days
                        if logon_age_days > self.config.dormant_days:
                            dormant = True

                # Account expiration
                acc_expires = get_attr(entry, 'accountExpires')
                acc_expires_int = safe_int(acc_expires)
                acc_expires_dt = windows_timestamp_to_datetime(acc_expires_int) if acc_expires_int else None
                acc_expires_days = None
                if acc_expires_dt:
                    acc_expires_days = (acc_expires_dt - now).days

                # Kerberos encryption types
                enc_types = get_attr(entry, 'msDS-SupportedEncryptionTypes')
                kerb_enc = parse_kerb_enc_types(enc_types)

                # Delegation
                delegation_type = None
                delegation_protocol = None
                delegation_services = None

                if uac_parsed.get('TrustedForDelegation', False):
                    delegation_type = "Unconstrained"
                    delegation_services = "Any"

                allowed_to_delegate = get_attr_list(entry, 'msDS-AllowedToDelegateTo')
                if allowed_to_delegate:
                    delegation_type = "Constrained"
                    delegation_services = ", ".join([str(s) for s in allowed_to_delegate])

                if uac_parsed.get('TrustedToAuthForDelegation', False):
                    delegation_protocol = "Any"
                elif delegation_type:
                    delegation_protocol = "Kerberos"

                # SPNs
                spns = get_attr_list(entry, 'servicePrincipalName')
                has_spn = len(spns) > 0

                # SID History
                sid_history = get_attr_list(entry, 'sIDHistory')
                sid_history_str = ", ".join([sid_to_str(s) for s in sid_history]) if sid_history else ""

                results.append({
                    "UserName": get_attr(entry, 'sAMAccountName', ''),
                    "Name": get_attr(entry, 'name', ''),
                    "Enabled": uac_parsed.get('Enabled', ''),
                    "Must Change Password at Logon": must_change_pwd,
                    "Cannot Change Password": cannot_change_pwd,
                    "Password Never Expires": uac_parsed.get('PasswordNeverExpires', False),
                    "Reversible Password Encryption": uac_parsed.get('ReversibleEncryption', False),
                    "Smartcard Logon Required": uac_parsed.get('SmartcardRequired', False),
                    "Delegation Permitted": not uac_parsed.get('NotDelegated', False),
                    "Kerberos DES Only": uac_parsed.get('UseDESKeyOnly', False),
                    "Kerberos RC4": kerb_enc.get('RC4', 'Not Set'),
                    "Kerberos AES-128bit": kerb_enc.get('AES128', 'Not Set'),
                    "Kerberos AES-256bit": kerb_enc.get('AES256', 'Not Set'),
                    "Does Not Require Pre Auth": uac_parsed.get('DoesNotRequirePreAuth', False),
                    "Never Logged in": never_logged_in,
                    "Logon Age (days)": logon_age_days if logon_age_days is not None else "",
                    "Password Age (days)": pwd_age_days if pwd_age_days is not None else "",
                    f"Dormant (> {self.config.dormant_days} days)": dormant,
                    f"Password Age (> {self.config.password_age_days} days)": pwd_not_changed_max,
                    "Account Locked Out": uac_parsed.get('AccountLockedOut', False),
                    "Password Expired": uac_parsed.get('PasswordExpired', False),
                    "Password Not Required": uac_parsed.get('PasswordNotRequired', False),
                    "Delegation Type": delegation_type or "",
                    "Delegation Protocol": delegation_protocol or "",
                    "Delegation Services": delegation_services or "",
                    "Logon Workstations": get_attr(entry, 'userWorkstations', ''),
                    "AdminCount": get_attr(entry, 'adminCount', ''),
                    "Primary GroupID": get_attr(entry, 'primaryGroupID', ''),
                    "SID": sid_to_str(get_attr(entry, 'objectSid')),
                    "SIDHistory": sid_history_str,
                    "HasSPN": has_spn,
                    "Description": get_attr(entry, 'description', ''),
                    "Title": get_attr(entry, 'title', ''),
                    "Department": get_attr(entry, 'department', ''),
                    "Company": get_attr(entry, 'company', ''),
                    "Manager": get_attr(entry, 'manager', ''),
                    "Info": get_attr(entry, 'info', ''),
                    "Last Logon Date": format_datetime(last_logon_dt),
                    "Password LastSet": format_datetime(pwd_last_set_dt),
                    "Account Expiration Date": format_datetime(acc_expires_dt),
                    "Account Expiration (days)": acc_expires_days if acc_expires_days is not None else "",
                    "Mobile": get_attr(entry, 'mobile', ''),
                    "Email": get_attr(entry, 'mail', ''),
                    "HomeDirectory": get_attr(entry, 'homeDirectory', ''),
                    "ProfilePath": get_attr(entry, 'profilePath', ''),
                    "ScriptPath": get_attr(entry, 'scriptPath', ''),
                    "UserAccountControl": uac,
                    "First Name": get_attr(entry, 'givenName', ''),
                    "Middle Name": get_attr(entry, 'middleName', ''),
                    "Last Name": get_attr(entry, 'sn', ''),
                    "Country": get_attr(entry, 'c', ''),
                    "whenCreated": format_datetime(self._parse_timestamp(get_attr(entry, 'whenCreated'))),
                    "whenChanged": format_datetime(self._parse_timestamp(get_attr(entry, 'whenChanged'))),
                    "DistinguishedName": get_attr(entry, 'distinguishedName', ''),
                    "CanonicalName": get_attr(entry, 'canonicalName', ''),
                })

        except Exception as e:
            logger.warning(f"Error collecting users: {e}")
            import traceback
            traceback.print_exc()

        self.results['Users'] = results
        logger.info(f"    Found {len(results)} users")
        return results

    def collect_computers(self) -> List[Dict]:
        """Collect computer objects and service accounts (users ending in $)."""
        logger.info("[-] Collecting Computers - May take some time...")
        results = []
        now = datetime.now(timezone.utc).replace(tzinfo=None)

        # Build hostname-to-IP mapping from DNS records if available
        hostname_to_ip, hostname_to_ipv6 = self._build_hostname_to_ip_cache()

        try:
            # Collect both computer objects AND user accounts ending in $ (service accounts)
            # Filter 1: Computer objects
            filter_str = "(objectCategory=computer)"
            if self.config.only_enabled:
                filter_str = "(&(objectCategory=computer)(!(userAccountControl:1.2.840.113556.1.4.803:=2)))"

            entries = self.search(
                search_base=self.base_dn,
                search_filter=filter_str,
                attributes=['sAMAccountName', 'name', 'distinguishedName', 'dNSHostName',
                           'operatingSystem', 'operatingSystemVersion', 'operatingSystemServicePack',
                           'operatingSystemHotfix', 'userAccountControl', 'pwdLastSet',
                           'lastLogonTimestamp', 'description', 'primaryGroupID', 'objectSid',
                           'sIDHistory', 'servicePrincipalName', 'msDS-AllowedToDelegateTo',
                           'msDS-SupportedEncryptionTypes',
                           'whenCreated', 'whenChanged']
            )
            
            # Filter 2: User accounts ending in $ (service accounts that look like computers)
            service_filter = "(&(objectCategory=person)(objectClass=user)(sAMAccountName=*$))"
            if self.config.only_enabled:
                service_filter = "(&(objectCategory=person)(objectClass=user)(sAMAccountName=*$)(!(userAccountControl:1.2.840.113556.1.4.803:=2)))"
            
            service_entries = self.search(
                search_base=self.base_dn,
                search_filter=service_filter,
                attributes=['sAMAccountName', 'name', 'distinguishedName', 'dNSHostName',
                           'operatingSystem', 'operatingSystemVersion', 'operatingSystemServicePack',
                           'operatingSystemHotfix', 'userAccountControl', 'pwdLastSet',
                           'lastLogonTimestamp', 'description', 'primaryGroupID', 'objectSid',
                           'sIDHistory', 'servicePrincipalName', 'msDS-AllowedToDelegateTo',
                           'msDS-SupportedEncryptionTypes',
                           'whenCreated', 'whenChanged']
            )
            
            # Filter 3: Managed Service Accounts (MSAs)
            msa_filter = "(objectClass=msDS-ManagedServiceAccount)"
            if self.config.only_enabled:
                msa_filter = "(&(objectClass=msDS-ManagedServiceAccount)(!(userAccountControl:1.2.840.113556.1.4.803:=2)))"
            
            msa_entries = self.search(
                search_base=self.base_dn,
                search_filter=msa_filter,
                attributes=['sAMAccountName', 'name', 'distinguishedName', 'dNSHostName',
                           'operatingSystem', 'operatingSystemVersion', 'operatingSystemServicePack',
                           'operatingSystemHotfix', 'userAccountControl', 'pwdLastSet',
                           'lastLogonTimestamp', 'description', 'primaryGroupID', 'objectSid',
                           'sIDHistory', 'servicePrincipalName', 'msDS-AllowedToDelegateTo',
                           'msDS-SupportedEncryptionTypes',
                           'whenCreated', 'whenChanged']
            )
            
            # Combine all lists
            entries.extend(service_entries)
            entries.extend(msa_entries)

            for entry in entries:
                
                uac = get_attr(entry, 'userAccountControl', 0)
                uac_parsed = parse_uac(uac)

                # Password last set
                pwd_last_set = self._parse_timestamp(get_attr(entry, 'pwdLastSet'))
                pwd_last_set_dt = None
                if isinstance(pwd_last_set, datetime):
                    # Check if pwdLastSet is 0 (1601-01-01 epoch) which means password never set
                    if pwd_last_set.year == 1601:
                        pwd_last_set_dt = None
                    else:
                        pwd_last_set_dt = pwd_last_set
                pwd_age_days = None
                pwd_not_changed_max = False

                if pwd_last_set_dt:
                    # Calculate age in UTC
                    pwd_last_set_utc = pwd_last_set_dt.replace(tzinfo=None) if pwd_last_set_dt.tzinfo else pwd_last_set_dt
                    pwd_age_days = (now - pwd_last_set_utc).days
                    if pwd_age_days > self.config.password_age_days:
                        pwd_not_changed_max = True

                # Last logon
                last_logon = self._parse_timestamp(get_attr(entry, 'lastLogonTimestamp'))
                last_logon_dt = None
                if isinstance(last_logon, datetime):
                    # Check if lastLogonTimestamp is 0 (1601-01-01 epoch) which means never logged in
                    if last_logon.year == 1601:
                        last_logon_dt = None
                    else:
                        last_logon_dt = last_logon
                logon_age_days = None
                dormant = False

                if last_logon_dt:
                    # Calculate age in UTC
                    last_logon_utc = last_logon_dt.replace(tzinfo=None) if last_logon_dt.tzinfo else last_logon_dt
                    logon_age_days = (now - last_logon_utc).days
                    if logon_age_days > self.config.dormant_days:
                        dormant = True

                # Operating system
                os_name = safe_str(get_attr(entry, 'operatingSystem', ''))
                os_version = safe_str(get_attr(entry, 'operatingSystemVersion', ''))
                os_sp = safe_str(get_attr(entry, 'operatingSystemServicePack', ''))
                os_full = f"{os_name} {os_sp} {os_version}".strip()

                # Delegation
                delegation_type = None
                delegation_protocol = None
                delegation_services = None

                if uac_parsed.get('TrustedForDelegation', False):
                    delegation_type = "Unconstrained"
                    delegation_services = "Any"

                allowed_to_delegate = get_attr_list(entry, 'msDS-AllowedToDelegateTo')
                if allowed_to_delegate:
                    delegation_type = "Constrained"
                    delegation_services = ", ".join([str(s) for s in allowed_to_delegate])

                if uac_parsed.get('TrustedToAuthForDelegation', False):
                    delegation_protocol = "Any"
                elif delegation_type:
                    delegation_protocol = "Kerberos"

                # Kerberos encryption types
                enc_types = get_attr(entry, 'msDS-SupportedEncryptionTypes')
                kerb_enc = parse_kerb_enc_types(enc_types)

                # SID History
                sid_history = get_attr_list(entry, 'sIDHistory')
                sid_history_str = ", ".join([sid_to_str(s) for s in sid_history]) if sid_history else ""

                # Creator SID (for machine accounts created by users)
                # Note: msDS-CreatorSID is not supported by ADWS protocol
                # The LDAP version can retrieve this attribute, but ADWS returns a SOAP fault
                # when attempting to enumerate it. This is a known ADWS limitation.
                # Computer objects created by regular users will not show the creator's SID.
                creator_sid_str = ""

                # Get IP addresses from collected DNS records (offline lookup)
                ipv4_address = ""
                ipv6_address = ""
                dns_hostname = get_attr(entry, 'dNSHostName', '')
                if dns_hostname:
                    # Extract just the hostname part (before first dot) for lookup
                    hostname_part = dns_hostname.split('.')[0].lower()
                    ipv4_address = hostname_to_ip.get(hostname_part, '')
                    ipv6_address = hostname_to_ipv6.get(hostname_part, '')

                computer_dict = {
                    "UserName": get_attr(entry, 'sAMAccountName', ''),
                    "Name": get_attr(entry, 'name', ''),
                    "DNSHostName": dns_hostname,
                    "Enabled": uac_parsed.get('Enabled', ''),
                    "IPv4Address": ipv4_address,
                    "IPv6Address": ipv6_address,
                    "Operating System": os_full,
                    "Logon Age (days)": logon_age_days,
                    "Password Age (days)": pwd_age_days,
                    f"Dormant (> {self.config.dormant_days} days)": dormant,
                    f"Password Age (> {self.config.password_age_days} days)": pwd_not_changed_max,
                    "Delegation Type": delegation_type or "",
                    "Delegation Protocol": delegation_protocol or "",
                    "Delegation Services": delegation_services or "",
                    "Primary Group ID": get_attr(entry, 'primaryGroupID', ''),
                    "SID": sid_to_str(get_attr(entry, 'objectSid')),
                    "SIDHistory": sid_history_str,
                    "Description": get_attr(entry, 'description', ''),
                    "ms-ds-CreatorSid": creator_sid_str,
                    "Last Logon Date": format_datetime(last_logon_dt),
                    "Password LastSet": format_datetime(pwd_last_set_dt),
                    "UserAccountControl": uac,
                    "whenCreated": format_datetime(self._parse_timestamp(get_attr(entry, 'whenCreated'))),
                    "whenChanged": format_datetime(self._parse_timestamp(get_attr(entry, 'whenChanged'))),
                    "Distinguished Name": get_attr(entry, 'distinguishedName', ''),
                }
                results.append(computer_dict)

        except Exception as e:
            logger.warning(f"Error collecting computers: {e}")

        self.results['Computers'] = results
        logger.info(f"    Found {len(results)} computers")
        return results

    def collect_groups(self) -> List[Dict]:
        """Collect group objects."""
        logger.info("[-] Collecting Groups...")
        results = []

        try:
            entries = self.search(
                search_base=self.base_dn,
                search_filter='(objectCategory=group)',
                attributes=['sAMAccountName', 'name', 'distinguishedName', 'canonicalName',
                           'description', 'groupType', 'adminCount', 'managedBy',
                           'objectSid', 'sIDHistory', 'whenCreated', 'whenChanged']
            )

            for entry in entries:
                
                group_type = get_attr(entry, 'groupType')
                gt = safe_int(group_type)
                group_type_str = GROUP_TYPE.get(gt, "Unknown")
                if gt & 0x80000000:  # Security group
                    group_category = "Security"
                else:
                    group_category = "Distribution"

                if gt & 0x02:
                    group_scope = "Global"
                elif gt & 0x04:
                    group_scope = "DomainLocal"
                elif gt & 0x08:
                    group_scope = "Universal"
                else:
                    group_scope = "Unknown"

                # Managed by
                managed_by = get_attr(entry, 'managedBy', '')
                if managed_by:
                    match = re.search(r'CN=([^,]+)', str(managed_by))
                    if match:
                        managed_by = match.group(1)

                # SID History
                sid_history = get_attr_list(entry, 'sIDHistory')
                sid_history_str = ", ".join([sid_to_str(s) for s in sid_history]) if sid_history else ""

                results.append({
                    "Name": get_attr(entry, 'name', ''),
                    "AdminCount": get_attr(entry, 'adminCount', ''),
                    "GroupCategory": group_category,
                    "GroupScope": group_scope,
                    "ManagedBy": managed_by,
                    "SID": sid_to_str(get_attr(entry, 'objectSid')),
                    "SIDHistory": sid_history_str,
                    "Description": get_attr(entry, 'description', ''),
                    "whenCreated": format_datetime(self._parse_timestamp(get_attr(entry, 'whenCreated'))),
                    "whenChanged": format_datetime(self._parse_timestamp(get_attr(entry, 'whenChanged'))),
                    "DistinguishedName": get_attr(entry, 'distinguishedName', ''),
                    "CanonicalName": get_attr(entry, 'canonicalName', ''),
                })

        except Exception as e:
            logger.warning(f"Error collecting groups: {e}")

        self.results['Groups'] = results
        logger.info(f"    Found {len(results)} groups")
        return results

    def collect_group_members(self) -> List[Dict]:
        """Collect group memberships."""
        logger.info("[-] Collecting Group Members - May take some time...")
        results = []

        try:
            # Get all groups with members
            entries = self.search(
                search_base=self.base_dn,
                search_filter="(&(objectCategory=group)(member=*))",
                attributes=['sAMAccountName', 'name', 'member', 'objectSid']
            )

            for entry in entries:
                
                group_name = get_attr(entry, 'name', '')
                group_sam = get_attr(entry, 'sAMAccountName', '')
                members = get_attr_list(entry, 'member')

                for member_dn in members:
                    # Extract member name from DN
                    match = re.search(r'CN=([^,]+)', str(member_dn))
                    member_name = match.group(1) if match else str(member_dn)
                    
                    # Query member details
                    member_username = ""
                    member_sid = ""
                    account_type = ""
                    
                    try:
                        # Look up the member object to get its details
                        member_entries = self.search(
                            search_base=str(member_dn),
                            search_filter="(objectClass=*)",
                            attributes=['sAMAccountName', 'objectSid', 'objectClass'],
                            scope='BASE'
                        )
                        
                        if member_entries:
                            member_entry = member_entries[0]
                            member_username = get_attr(member_entry, 'sAMAccountName', '')
                            member_sid_raw = get_attr(member_entry, 'objectSid')
                            if member_sid_raw:
                                member_sid = sid_to_str(member_sid_raw)
                            
                            # Determine account type from objectClass
                            object_classes = get_attr_list(member_entry, 'objectClass')
                            object_classes_str = [str(oc).lower() for oc in object_classes]
                            
                            if 'user' in object_classes_str and 'computer' not in object_classes_str:
                                account_type = "user"
                            elif 'computer' in object_classes_str:
                                account_type = "computer"
                            elif 'group' in object_classes_str:
                                account_type = "group"
                            else:
                                account_type = "unknown"
                    except Exception:
                        # If lookup fails, try to guess from DN
                        member_dn_str = str(member_dn)
                        if ',CN=Users,' in member_dn_str or ',OU=Users,' in member_dn_str:
                            account_type = "user"
                        elif ',CN=Computers,' in member_dn_str or ',OU=Computers,' in member_dn_str:
                            account_type = "computer"
                        elif ',CN=Builtin,' in member_dn_str or 'CN=Groups' in member_dn_str:
                            account_type = "group"
                        else:
                            account_type = "unknown"

                    results.append({
                        "Group Name": group_name,
                        "Member UserName": member_username,
                        "Member Name": member_name,
                        "Member SID": member_sid,
                        "AccountType": account_type,
                    })

        except Exception as e:
            logger.warning(f"Error collecting group members: {e}")

        self.results['GroupMembers'] = results
        logger.info(f"    Found {len(results)} group memberships")
        return results

    def collect_ous(self) -> List[Dict]:
        """Collect organizational units."""
        logger.info("[-] Collecting Organizational Units...")
        results = []

        try:
            entries = self.search(
                search_base=self.base_dn,
                search_filter='(objectCategory=organizationalUnit)',
                attributes=['name', 'distinguishedName', 'description', 'gPLink',
                           'whenCreated', 'whenChanged']
            )

            for entry in entries:
                
                gp_link = get_attr(entry, 'gPLink', '')
                dn = get_attr(entry, 'distinguishedName', '')
                
                # Calculate depth based on DN
                depth = dn.count('OU=') if dn else 0

                results.append({
                    "Name": get_attr(entry, 'name', ''),
                    "Depth": depth,
                    "Description": get_attr(entry, 'description', ''),
                    "whenCreated": format_datetime(self._parse_timestamp(get_attr(entry, 'whenCreated'))),
                    "whenChanged": format_datetime(self._parse_timestamp(get_attr(entry, 'whenChanged'))),
                    "DistinguishedName": dn,
                })

        except Exception as e:
            logger.warning(f"Error collecting OUs: {e}")

        self.results['OUs'] = results
        logger.info(f"    Found {len(results)} organizational units")
        return results

    def collect_gpos(self) -> List[Dict]:
        """Collect Group Policy Objects."""
        logger.info("[-] Collecting GPOs...")
        results = []

        try:
            entries = self.search(
                search_base=self.base_dn,
                search_filter='(objectCategory=groupPolicyContainer)',
                attributes=['displayName', 'name', 'distinguishedName', 'gPCFileSysPath',
                           'whenCreated', 'whenChanged', 'flags']
            )

            for entry in entries:
                
                flags = get_attr(entry, 'flags', 0)
                flags = safe_int(flags)

                # GPO flags
                user_disabled = bool(flags & 1)
                computer_disabled = bool(flags & 2)

                results.append({
                    "DisplayName": get_attr(entry, 'displayName', ''),
                    "GUID": get_attr(entry, 'name', ''),
                    "whenCreated": format_datetime(self._parse_timestamp(get_attr(entry, 'whenCreated'))),
                    "whenChanged": format_datetime(self._parse_timestamp(get_attr(entry, 'whenChanged'))),
                    "DistinguishedName": get_attr(entry, 'distinguishedName', ''),
                    "FilePath": get_attr(entry, 'gPCFileSysPath', ''),
                })

        except Exception as e:
            logger.warning(f"Error collecting GPOs: {e}")

        self.results['GPOs'] = results
        logger.info(f"    Found {len(results)} GPOs")
        return results

    def collect_domain_controllers(self) -> List[Dict]:
        """Collect domain controller information."""
        logger.info("[-] Collecting Domain Controllers...")
        results = []

        try:
            # Get FSMO role holders
            fsmo_roles = {
                'pdc': None,
                'rid': None,
                'infra': None,
                'schema': None,
                'naming': None
            }
            
            # PDC Emulator - stored in domain object
            try:
                domain_entries = self.search(
                    search_base=self.base_dn,
                    search_filter='(objectCategory=domainDNS)',
                    attributes=['fSMORoleOwner']
                )
                if domain_entries:
                    pdc_owner = get_attr(domain_entries[0], 'fSMORoleOwner', '')
                    if pdc_owner:
                        fsmo_roles['pdc'] = str(pdc_owner).split(',')[1].replace('CN=', '').upper()
            except:
                pass
            
            # RID Master - stored in RID Manager object
            try:
                rid_entries = self.search(
                    search_base=f"CN=RID Manager$,CN=System,{self.base_dn}",
                    search_filter='(objectClass=*)',
                    attributes=['fSMORoleOwner'],
                    scope='BASE'
                )
                if rid_entries:
                    rid_owner = get_attr(rid_entries[0], 'fSMORoleOwner', '')
                    if rid_owner:
                        fsmo_roles['rid'] = str(rid_owner).split(',')[1].replace('CN=', '').upper()
            except:
                pass
            
            # Infrastructure Master - stored in Infrastructure object
            try:
                infra_entries = self.search(
                    search_base=f"CN=Infrastructure,{self.base_dn}",
                    search_filter='(objectClass=*)',
                    attributes=['fSMORoleOwner'],
                    scope='BASE'
                )
                if infra_entries:
                    infra_owner = get_attr(infra_entries[0], 'fSMORoleOwner', '')
                    if infra_owner:
                        fsmo_roles['infra'] = str(infra_owner).split(',')[1].replace('CN=', '').upper()
            except:
                pass
            
            # Schema Master - stored in Schema container
            try:
                schema_entries = self.search(
                    search_base=self.schema_dn,
                    search_filter='(objectClass=dMD)',
                    attributes=['fSMORoleOwner']
                )
                if schema_entries:
                    schema_owner = get_attr(schema_entries[0], 'fSMORoleOwner', '')
                    if schema_owner:
                        fsmo_roles['schema'] = str(schema_owner).split(',')[1].replace('CN=', '').upper()
            except:
                pass
            
            # Domain Naming Master - stored in Partitions container
            try:
                naming_entries = self.search(
                    search_base=f"CN=Partitions,{self.config_dn}",
                    search_filter='(objectClass=crossRefContainer)',
                    attributes=['fSMORoleOwner']
                )
                if naming_entries:
                    naming_owner = get_attr(naming_entries[0], 'fSMORoleOwner', '')
                    if naming_owner:
                        fsmo_roles['naming'] = str(naming_owner).split(',')[1].replace('CN=', '').upper()
            except:
                pass

            # Build hostname-to-IP mapping from DNS records if available
            hostname_to_ip, hostname_to_ipv6 = self._build_hostname_to_ip_cache()

            # Find DCs using userAccountControl
            entries = self.search(
                search_base=self.base_dn,
                search_filter="(&(objectCategory=computer)(userAccountControl:1.2.840.113556.1.4.803:=8192))",
                attributes=['name', 'dNSHostName', 'operatingSystem', 'serverReferenceBL']
            )

            for entry in entries:
                
                dc_name = safe_str(get_attr(entry, 'name', '')).upper()
                hostname = safe_str(get_attr(entry, 'dNSHostName', ''))
                
                # Get IPv4 address from offline DNS cache
                ipv4 = ""
                if hostname:
                    # Extract just the hostname part (before first dot) for lookup
                    hostname_part = hostname.split('.')[0].lower()
                    ipv4 = hostname_to_ip.get(hostname_part, '')
                    
                    # If not found and this is the DC we're connected to, use config IP
                    if not ipv4:
                        if hostname.lower() == self.config.domain_controller.lower() or \
                           dc_name.lower() == self.config.domain_controller.lower():
                            ipv4 = self.config.domain_controller
                
                # Get site information from serverReferenceBL
                site = ""
                server_ref = get_attr(entry, 'serverReferenceBL', '')
                if server_ref:
                    # serverReferenceBL looks like: CN=DC1,CN=Servers,CN=Default-First-Site-Name,CN=Sites,CN=Configuration,DC=...
                    parts = str(server_ref).split(',')
                    for i, part in enumerate(parts):
                        if part.startswith('CN=Sites'):
                            if i > 0:
                                site = parts[i-1].replace('CN=', '')
                            break
                
                # Check FSMO roles for this DC
                is_pdc = fsmo_roles['pdc'] == dc_name
                is_rid = fsmo_roles['rid'] == dc_name
                is_infra = fsmo_roles['infra'] == dc_name
                is_schema = fsmo_roles['schema'] == dc_name
                is_naming = fsmo_roles['naming'] == dc_name
                
                # SMB detection (simplified - would need actual SMB protocol testing)
                smb_port_open = "TRUE" if ipv4 else "FALSE"
                
                dc_info = {
                    "Domain": dn_to_fqdn(self.base_dn),
                    "Site": site,
                    "Name": dc_name.lower(),
                    "IPv4Address": ipv4,
                    "Operating System": safe_str(get_attr(entry, 'operatingSystem', '')),
                    "Hostname": hostname,
                    "Infra": "TRUE" if is_infra else "FALSE",
                    "Naming": "TRUE" if is_naming else "FALSE",
                    "Schema": "TRUE" if is_schema else "FALSE",
                    "RID": "TRUE" if is_rid else "FALSE",
                    "PDC": "TRUE" if is_pdc else "FALSE",
                    "SMB Port Open": smb_port_open,
                    "SMB1(NT LM 0.12)": "FALSE",  # Would require SMB protocol testing
                    "SMB2(0x0202)": "TRUE",  # Placeholder
                    "SMB2(0x0210)": "TRUE",  # Placeholder
                    "SMB3(0x0300)": "TRUE",  # Placeholder
                    "SMB3(0x0302)": "TRUE",  # Placeholder
                    "SMB3(0x0311)": "FALSE",  # Placeholder
                    "SMB Signing": "TRUE",  # Placeholder
                }

                results.append(dc_info)

        except Exception as e:
            logger.warning(f"Error collecting DCs: {e}")

        self.results['DomainControllers'] = results
        logger.info(f"    Found {len(results)} domain controllers")
        return results

    def collect_password_policy(self) -> List[Dict]:
        """Collect default domain password policy."""
        logger.info("[-] Collecting Default Password Policy...")
        results = []

        try:
            entries = self.search(
                search_base=self.base_dn,
                search_filter='(objectCategory=domainDNS)',
                attributes=['minPwdLength', 'pwdHistoryLength', 'pwdProperties',
                           'maxPwdAge', 'minPwdAge', 'lockoutThreshold',
                           'lockoutDuration', 'lockOutObservationWindow']
            )

            if entries:
                entry = entries[0]

                # Get password policy time intervals (these are raw integers, not timestamps)
                # Note: These are stored as negative 100-nanosecond intervals
                max_pwd_age = get_attr(entry, 'maxPwdAge')
                min_pwd_age = get_attr(entry, 'minPwdAge')
                lockout_duration = get_attr(entry, 'lockoutDuration')
                lockout_window = get_attr(entry, 'lockOutObservationWindow')

                def format_interval(interval, unit='days'):
                    """
                    Format AD time interval as days, hours, minutes, seconds (supports fractional values).
                    - For maxPwdAge, minPwdAge, lockoutDuration, lockoutObservationWindow: expects 100-nanosecond intervals (usually negative).
                    - For 0 or None, returns '0d 0h 0m 0s'.
                    """
                    if interval is None:
                        return '0d 0h 0m 0s'
                    try:
                        if isinstance(interval, timedelta):
                            total_seconds = abs(interval.total_seconds())
                        else:
                            interval_val = int(interval)
                            total_seconds = abs(interval_val) / 10000000
                        days = int(total_seconds // 86400)
                        hours = int((total_seconds % 86400) // 3600)
                        minutes = int((total_seconds % 3600) // 60)
                        seconds = int(total_seconds % 60)
                        if unit == 'days':
                            # Show as 'Xd Yh Zm Ws' for clarity
                            return f"{days}d {hours}h {minutes}m {seconds}s"
                        elif unit == 'minutes':
                            total_minutes = int(round(total_seconds / 60))
                            return f"{total_minutes}"
                        else:
                            return str(total_seconds)
                    except Exception as e:
                        logger.debug(f"Error formatting interval: {interval}, error: {e}")
                        return '0d 0h 0m 0s'

                pwd_props = get_attr(entry, 'pwdProperties', 0)
                pwd_props = safe_int(pwd_props)

                # Match ADRecon order and naming with security best practices
                results.append({
                    "Policy": "Enforce password history (passwords)",
                    "Current Value": str(get_attr(entry, 'pwdHistoryLength', '')),
                    "CIS Benchmark 2024-25": "24 or more passwords",
                    "PCI DSS v4.0.1": "4",
                    "PCI DSS Requirement": "Req. 8.3.7"
                })
                results.append({
                    "Policy": "Maximum password age (d h m s)",
                    "Current Value": format_interval(max_pwd_age, unit='days'),
                    "CIS Benchmark 2024-25": "1 to 365 days",
                    "PCI DSS v4.0.1": "90 days",
                    "PCI DSS Requirement": "Req. 8.3.9"
                })
                results.append({
                    "Policy": "Minimum password age (d h m s)",
                    "Current Value": format_interval(min_pwd_age, unit='days'),
                    "CIS Benchmark 2024-25": "1 or more days",
                    "PCI DSS v4.0.1": "N/A",
                    "PCI DSS Requirement": "-"
                })
                results.append({
                    "Policy": "Minimum password length (characters)",
                    "Current Value": str(get_attr(entry, 'minPwdLength', '')),
                    "CIS Benchmark 2024-25": "14 or more chars",
                    "PCI DSS v4.0.1": "12 chars",
                    "PCI DSS Requirement": "Req. 8.3.6"
                })
                results.append({
                    "Policy": "Password must meet complexity requirements",
                    "Current Value": "TRUE" if (pwd_props & 1) else "FALSE",
                    "CIS Benchmark 2024-25": "TRUE",
                    "PCI DSS v4.0.1": "TRUE",
                    "PCI DSS Requirement": "Req. 8.3.6"
                })
                results.append({
                    "Policy": "Store password using reversible encryption for all users in the domain",
                    "Current Value": "TRUE" if (pwd_props & 16) else "FALSE",
                    "CIS Benchmark 2024-25": "FALSE",
                    "PCI DSS v4.0.1": "N/A",
                    "PCI DSS Requirement": "-"
                })
                results.append({
                    "Policy": "Account lockout duration (d h m s)",
                    "Current Value": format_interval(lockout_duration, unit='days'),
                    "CIS Benchmark 2024-25": "15 mins or more",
                    "PCI DSS v4.0.1": "0 (manual unlock) or 30 mins",
                    "PCI DSS Requirement": "Req. 8.3.4"
                })
                results.append({
                    "Policy": "Account lockout threshold (attempts)",
                    "Current Value": str(get_attr(entry, 'lockoutThreshold', '')),
                    "CIS Benchmark 2024-25": "1 to 5 attempts",
                    "PCI DSS v4.0.1": "1 to 10 attempts",
                    "PCI DSS Requirement": "Req. 8.3.4"
                })
                results.append({
                    "Policy": "Reset account lockout counter after (mins)",
                    "Current Value": format_interval(lockout_window, unit='minutes'),
                    "CIS Benchmark 2024-25": "15 or more mins",
                    "PCI DSS v4.0.1": "N/A",
                    "PCI DSS Requirement": "-"
                })

        except Exception as e:
            logger.warning(f"Error collecting password policy: {e}")

        self.results['PasswordPolicy'] = results
        logger.info(f"    Found {len(results)} password policy settings")
        return results

    def collect_user_spns(self) -> List[Dict]:
        """Collect user SPNs (Service Principal Names)."""
        logger.info("[-] Collecting User SPNs...")
        results = []

        try:
            # Users with SPNs - using different filter structure
            filter_str = "(&(objectCategory=person)(objectClass=user)(servicePrincipalName=*))"
            if self.config.only_enabled:
                filter_str = "(&(objectCategory=person)(objectClass=user)(servicePrincipalName=*)(!(userAccountControl:1.2.840.113556.1.4.803:=2)))"

            entries = self.search(
                search_base=self.base_dn,
                search_filter=filter_str,
                attributes=['sAMAccountName', 'name', 'description', 'info', 'memberOf',
                           'servicePrincipalName', 'primaryGroupID', 'pwdLastSet', 'userAccountControl']
            )

            for entry in entries:
                
                spns = get_attr_list(entry, 'servicePrincipalName')
                uac = get_attr(entry, 'userAccountControl', 0)
                uac_parsed = parse_uac(uac)

                pwd_last_set = self._parse_timestamp(get_attr(entry, 'pwdLastSet'))
                pwd_last_set_dt = None
                if isinstance(pwd_last_set, datetime):
                    # Convert timezone-aware datetime to naive for comparison
                    pwd_last_set_dt = pwd_last_set.replace(tzinfo=None) if pwd_last_set.tzinfo else pwd_last_set

                # Get group memberships
                member_of = get_attr_list(entry, 'memberOf')
                groups = []
                for m in member_of:
                    match = re.search(r'CN=([^,]+)', str(m))
                    if match:
                        groups.append(match.group(1))

                for spn in spns:
                    spn_parts = str(spn).split('/')
                    service = spn_parts[0] if len(spn_parts) > 0 else ""
                    host = spn_parts[1] if len(spn_parts) > 1 else ""

                    results.append({
                        "UserName": get_attr(entry, 'sAMAccountName', ''),
                        "Name": get_attr(entry, 'name', ''),
                        "Enabled": uac_parsed.get('Enabled', ''),
                        "Service": service,
                        "Host": host,
                        "Password Last Set": format_datetime(pwd_last_set_dt),
                        "Description": get_attr(entry, 'description', ''),
                        "Info": get_attr(entry, 'info', ''),
                        "Primary GroupID": get_attr(entry, 'primaryGroupID', ''),
                        "Memberof": ", ".join(groups),
                    })

        except Exception as e:
            logger.warning(f"Error collecting user SPNs: {e}")

        self.results['UserSPNs'] = results
        logger.info(f"    Found {len(results)} user SPNs")
        return results

    def collect_computer_spns(self) -> List[Dict]:
        """Collect computer SPNs."""
        logger.info("[-] Collecting Computer SPNs...")
        results = []

        try:
            # Collect computer objects
            filter_str = "(&(objectCategory=computer)(servicePrincipalName=*))"
            if self.config.only_enabled:
                filter_str = "(&(objectCategory=computer)(servicePrincipalName=*)(!(userAccountControl:1.2.840.113556.1.4.803:=2)))"

            entries = self.search(
                search_base=self.base_dn,
                search_filter=filter_str,
                attributes=['name', 'sAMAccountName', 'servicePrincipalName']
            )

            # Also collect Managed Service Accounts (MSAs)
            msa_filter = "(objectClass=msDS-ManagedServiceAccount)"
            if self.config.only_enabled:
                msa_filter = "(&(objectClass=msDS-ManagedServiceAccount)(!(userAccountControl:1.2.840.113556.1.4.803:=2)))"
            
            msa_entries = self.search(
                search_base=self.base_dn,
                search_filter=msa_filter,
                attributes=['name', 'sAMAccountName', 'servicePrincipalName']
            )

            # Combine both result sets
            all_entries = list(entries) if entries else []
            if msa_entries:
                all_entries.extend(msa_entries)

            # Group SPNs by (UserName, Name, Service) and collect hosts
            grouped_spns = {}
            
            for entry in all_entries:
                
                spns = get_attr_list(entry, 'servicePrincipalName')
                sam_account = get_attr(entry, 'sAMAccountName', '')
                computer_name = get_attr(entry, 'name', '')

                for spn in spns:
                    # SPNs can be: Service/Host or Service/Host/Realm
                    # We want to extract Service and Host (ignoring Realm)
                    spn_parts = str(spn).split('/')
                    service = spn_parts[0] if len(spn_parts) > 0 else ""
                    host = spn_parts[1] if len(spn_parts) > 1 else ""
                    
                    # Remove port number if present (e.g., hostname:port)
                    if ':' in host:
                        host = host.split(':')[0]

                    # Group by (sam_account, computer_name, service)
                    key = (sam_account, computer_name, service)
                    if key not in grouped_spns:
                        grouped_spns[key] = []
                    if host and host not in grouped_spns[key]:
                        grouped_spns[key].append(host)

            # Convert grouped data to results list
            for (sam_account, computer_name, service), hosts in grouped_spns.items():
                results.append({
                    "UserName": sam_account,  # Already includes $ suffix
                    "Name": computer_name,
                    "Service": service,
                    "Host": ",".join(hosts),
                })

        except Exception as e:
            logger.warning(f"Error collecting computer SPNs: {e}")

        self.results['ComputerSPNs'] = results
        logger.info(f"    Found {len(results)} computer SPNs")
        return results

    def collect_laps(self) -> List[Dict]:
        """Collect LAPS (Local Administrator Password Solution) information."""
        logger.info("[-] Collecting LAPS - Needs Privileged Account...")
        results = []

        try:
            # Check if LAPS is installed
            laps_entries = self.search(
                search_base=self.schema_dn,
                search_filter="(name=ms-Mcs-AdmPwd)",
                attributes=['name']
            )

            if not laps_entries:
                logger.warning("[*] LAPS is not installed in this environment")
                return results

            # Get the schemaIDGUID for ms-Mcs-AdmPwd (forest-specific)
            logger.info("    Fetching ms-Mcs-AdmPwd schemaIDGUID...")
            laps_attr_guid = self._get_schema_guid("ms-Mcs-AdmPwd")
            if laps_attr_guid:
                logger.info(f"    Found LAPS attribute GUID: {laps_attr_guid}")
            else:
                logger.warning("    Could not retrieve LAPS attribute GUID - reader detection may be inaccurate")

            # Get LAPS passwords for computers - now including nTSecurityDescriptor to parse readers
            entries = self.search(
                search_base=self.base_dn,
                search_filter='(objectCategory=computer)',
                attributes=['name', 'dNSHostName', 'ms-Mcs-AdmPwd', 'ms-Mcs-AdmPwdExpirationTime',
                           'userAccountControl', 'nTSecurityDescriptor']
            )

            # Also get service accounts (user objects ending in $)
            service_entries = self.search(
                search_base=self.base_dn,
                search_filter="(&(objectCategory=person)(objectClass=user)(sAMAccountName=*$))",
                attributes=['name', 'sAMAccountName', 'dNSHostName', 'ms-Mcs-AdmPwd', 'ms-Mcs-AdmPwdExpirationTime',
                           'userAccountControl', 'nTSecurityDescriptor']
            )
            
            # Also get Managed Service Accounts (MSAs)
            msa_entries = self.search(
                search_base=self.base_dn,
                search_filter='(objectClass=msDS-ManagedServiceAccount)',
                attributes=['name', 'sAMAccountName', 'dNSHostName', 'ms-Mcs-AdmPwd', 'ms-Mcs-AdmPwdExpirationTime',
                           'userAccountControl', 'nTSecurityDescriptor']
            )

            # Combine all result sets
            all_entries = list(entries) if entries else []
            if service_entries:
                all_entries.extend(service_entries)
            if msa_entries:
                all_entries.extend(msa_entries)

            for entry in all_entries:
                
                uac = get_attr(entry, 'userAccountControl', 0)
                uac_parsed = parse_uac(uac)

                laps_pwd = safe_str(get_attr(entry, 'ms-Mcs-AdmPwd', ''))
                laps_exp = self._parse_timestamp(get_attr(entry, 'ms-Mcs-AdmPwdExpirationTime'))
                laps_exp_int = safe_int(get_attr(entry, 'ms-Mcs-AdmPwdExpirationTime'))
                laps_exp_dt = windows_timestamp_to_datetime(laps_exp_int) if laps_exp_int else None

                # Only include if we can read the password or it's set
                password_readable = bool(laps_pwd)
                password_stored = bool(laps_exp_int)  # If expiration time is set, password is stored

                # For service accounts, use sAMAccountName (without $), otherwise use dNSHostName
                hostname = get_attr(entry, 'dNSHostName', '')
                if not hostname:
                    # Service account - use sAMAccountName without the $ suffix
                    sam = get_attr(entry, 'sAMAccountName', '')
                    hostname = sam.rstrip('$') if sam else get_attr(entry, 'name', '')

                # Parse who can read the LAPS password with the attribute GUID
                laps_readers = self._parse_laps_readers(entry, laps_attr_guid)
                
                # Format reader information with details, filtering out well-known high-privilege principals
                if laps_readers:
                    # Format: "USERNAME (type) - why" - make usernames prominent
                    readers_formatted = []
                    for reader in laps_readers:
                        principal = reader.get('principal', 'Unknown')
                        
                        # Skip well-known high-privilege principals (Domain Admins, Enterprise Admins, etc.)
                        if self._is_well_known_high_privilege_principal(principal):
                            continue
                        
                        # Extract clean username from principal name
                        # e.g., "[User] lrvt" -> "lrvt" or "[Group] IT Admins" -> "IT Admins"
                        if principal.startswith('['):
                            # Format: [Type] Name
                            parts = principal.split('] ', 1)
                            if len(parts) == 2:
                                principal_type = parts[0].replace('[', '')
                                clean_name = parts[1]
                            else:
                                principal_type = 'Principal'
                                clean_name = principal
                        else:
                            principal_type = 'Principal'
                            clean_name = principal
                        
                        why = reader.get('why', '')
                        # Simplify the "why" message for readability
                        if 'Object-specific CONTROL_ACCESS for ms-Mcs-AdmPwd' in why:
                            simplified_why = 'Explicit LAPS read permission'
                        elif 'CONTROL_ACCESS on object' in why:
                            simplified_why = 'General read permission on computer'
                        elif 'High privilege on computer' in why:
                            simplified_why = 'High-level access to computer'
                        else:
                            simplified_why = why
                        
                        readers_formatted.append(f"{clean_name} ({principal_type}) - {simplified_why}")
                    
                    readers_str = ", ".join(readers_formatted) if readers_formatted else "Only default admins"
                else:
                    readers_str = "" if not laps_attr_guid else "None detected"

                results.append({
                    "Hostname": hostname,
                    "Enabled": uac_parsed.get('Enabled', ''),
                    "Stored": password_stored,
                    "Readable": password_readable,
                    "Password": laps_pwd if laps_pwd else "",
                    "Expiration": str(laps_exp_dt) if laps_exp_dt else "",
                    "Readers": readers_str,
                })

        except Exception as e:
            logger.warning(f"Error collecting LAPS: {e}")
            import traceback
            logger.debug(traceback.format_exc())

        self.results['LAPS'] = results
        logger.info(f"    Found {len(results)} LAPS entries")
        return results

    def collect_kerberoastable(self) -> List[Dict]:
        """Collect accounts vulnerable to Kerberoasting (user accounts with SPNs)."""
        logger.info("[-] Collecting Kerberoastable Accounts...")
        results = []
        now = datetime.now(timezone.utc).replace(tzinfo=None)

        try:
            # Find user accounts (not computers) with SPNs
            filter_str = "(&(objectCategory=person)(objectClass=user)(servicePrincipalName=*)(!(sAMAccountName=krbtgt)))"
            if self.config.only_enabled:
                filter_str = "(&(objectCategory=person)(objectClass=user)(servicePrincipalName=*)(!(sAMAccountName=krbtgt))(!(userAccountControl:1.2.840.113556.1.4.803:=2)))"

            entries = self.search(
                self.base_dn,
                filter_str,
                ['sAMAccountName', 'name', 'distinguishedName', 'userAccountControl',
                 'pwdLastSet', 'lastLogonTimestamp', 'adminCount', 'memberOf',
                 'servicePrincipalName', 'msDS-SupportedEncryptionTypes', 
                 'objectSid', 'description', 'info', 'whenCreated']
            )

            for entry in entries:
                uac = get_attr(entry, 'userAccountControl', 0)
                enabled = "True" if uac and not (int(uac) & 0x0002) else "False"
                
                # Password age
                pwd_last_set = get_attr(entry, 'pwdLastSet')
                pwd_age_days = None
                
                if isinstance(pwd_last_set, datetime):
                    if pwd_last_set.year != 1601:
                        pwd_last_set_utc = pwd_last_set.replace(tzinfo=None) if pwd_last_set.tzinfo else pwd_last_set
                        pwd_age_days = (now - pwd_last_set_utc).days
                elif pwd_last_set:
                    pwd_last_set_int = safe_int(pwd_last_set)
                    if pwd_last_set_int:
                        pwd_last_set_dt = windows_timestamp_to_datetime(pwd_last_set_int)
                        if pwd_last_set_dt and pwd_last_set_dt.year != 1601:
                            pwd_age_days = (now - pwd_last_set_dt).days
                
                # Last logon
                last_logon = get_attr(entry, 'lastLogonTimestamp')
                logon_age_days = None
                last_logon_dt = None
                
                if isinstance(last_logon, datetime):
                    if last_logon.year != 1601:
                        last_logon_dt = last_logon
                        last_logon_utc = last_logon.replace(tzinfo=None) if last_logon.tzinfo else last_logon
                        logon_age_days = (now - last_logon_utc).days
                elif last_logon:
                    last_logon_int = safe_int(last_logon)
                    if last_logon_int:
                        last_logon_dt = windows_timestamp_to_datetime(last_logon_int)
                        if last_logon_dt and last_logon_dt.year != 1601:
                            logon_age_days = (now - last_logon_dt).days
                
                # SPNs
                spns = get_attr_list(entry, 'servicePrincipalName')
                spn_count = len(spns)
                spns_str = "; ".join(spns[:3])  # Show first 3
                if spn_count > 3:
                    spns_str += f" ... ({spn_count} total)"
                
                # Kerberos encryption types
                enc_types = get_attr(entry, 'msDS-SupportedEncryptionTypes')
                kerb_enc = parse_kerb_enc_types(enc_types)
                
                # Check if privileged
                admin_count = get_attr(entry, 'adminCount', '')
                is_privileged = str(admin_count) == '1'
                
                # Risk assessment
                risk_level = "Medium"
                risk_factors = ["Kerberoastable (user account with SPN)"]
                
                if is_privileged:
                    risk_level = "CRITICAL"
                    risk_factors.append("Privileged account (AdminCount=1)")
                
                # Check encryption type - RC4 is easier to crack
                rc4_val = kerb_enc.get('RC4', 'Not Set')
                aes256_val = kerb_enc.get('AES256', 'Not Set')
                if rc4_val in ['Supported', 'Default'] and aes256_val not in ['Supported', 'Default']:
                    risk_level = "HIGH" if risk_level == "Medium" else risk_level
                    risk_factors.append("Uses weak RC4 encryption (easier to crack)")
                
                if pwd_age_days and pwd_age_days > 365:
                    risk_factors.append(f"Old password ({pwd_age_days} days)")
                elif pwd_age_days and pwd_age_days < 90:
                    risk_factors.append("Password recently changed")
                
                # Get group memberships
                member_of = get_attr_list(entry, 'memberOf')
                groups = []
                for group_dn in member_of:
                    match = re.search(r'CN=([^,]+)', str(group_dn))
                    if match:
                        groups.append(match.group(1))
                
                results.append({
                    "SAM Account Name": get_attr(entry, 'sAMAccountName', ''),
                    "Name": get_attr(entry, 'name', ''),
                    "Enabled": enabled,
                    "Risk Level": risk_level,
                    "Risk Factors": "; ".join(risk_factors),
                    "SPN Count": spn_count,
                    "Service Principal Names": spns_str,
                    "AdminCount": admin_count,
                    "Kerberos RC4": kerb_enc.get('RC4', 'Not Set'),
                    "Kerberos AES-128": kerb_enc.get('AES128', 'Not Set'),
                    "Kerberos AES-256": kerb_enc.get('AES256', 'Not Set'),
                    "Password Age (days)": pwd_age_days if pwd_age_days is not None else "Unknown",
                    "Last Logon": format_datetime(last_logon_dt) if last_logon_dt else "Never",
                    "Logon Age (days)": logon_age_days if logon_age_days is not None else "Unknown",
                    "Member Of": "; ".join(groups[:5]) if groups else "None",
                    "SID": sid_to_str(get_attr(entry, 'objectSid')),
                    "Description": get_attr(entry, 'description', ''),
                    "Info": get_attr(entry, 'info', ''),
                    "Created": format_datetime(self._parse_timestamp(get_attr(entry, 'whenCreated'))),
                    "Distinguished Name": get_attr(entry, 'distinguishedName', ''),
                })

        except Exception as e:
            logger.warning(f"Error collecting Kerberoastable accounts: {e}")
            logger.debug(f"Exception details: ", exc_info=True)

        self.results['Kerberoastable'] = results
        logger.info(f"    Found {len(results)} Kerberoastable accounts")
        return results

    def collect_asrep_roastable(self) -> List[Dict]:
        """Collect accounts vulnerable to AS-REP Roasting (no Kerberos pre-authentication required)."""
        logger.info("[-] Collecting AS-REP Roastable Accounts...")
        results = []
        now = datetime.now(timezone.utc).replace(tzinfo=None)

        try:
            # Find accounts with DONT_REQ_PREAUTH flag set (0x400000)
            filter_str = "(&(objectCategory=person)(objectClass=user)(userAccountControl:1.2.840.113556.1.4.803:=4194304))"
            if self.config.only_enabled:
                filter_str = "(&(objectCategory=person)(objectClass=user)(userAccountControl:1.2.840.113556.1.4.803:=4194304)(!(userAccountControl:1.2.840.113556.1.4.803:=2)))"

            entries = self.search(
                self.base_dn,
                filter_str,
                ['sAMAccountName', 'name', 'distinguishedName', 'userAccountControl',
                 'pwdLastSet', 'lastLogonTimestamp', 'adminCount', 'memberOf',
                 'servicePrincipalName', 'objectSid', 'description', 'info', 'whenCreated']
            )

            for entry in entries:
                uac = get_attr(entry, 'userAccountControl', 0)
                enabled = "True" if uac and not (int(uac) & 0x0002) else "False"
                
                # Password age
                pwd_last_set = get_attr(entry, 'pwdLastSet')
                pwd_age_days = None
                
                if isinstance(pwd_last_set, datetime):
                    if pwd_last_set.year != 1601:
                        pwd_last_set_utc = pwd_last_set.replace(tzinfo=None) if pwd_last_set.tzinfo else pwd_last_set
                        pwd_age_days = (now - pwd_last_set_utc).days
                elif pwd_last_set:
                    pwd_last_set_int = safe_int(pwd_last_set)
                    if pwd_last_set_int:
                        pwd_last_set_dt = windows_timestamp_to_datetime(pwd_last_set_int)
                        if pwd_last_set_dt and pwd_last_set_dt.year != 1601:
                            pwd_age_days = (now - pwd_last_set_dt).days
                
                # Last logon
                last_logon = get_attr(entry, 'lastLogonTimestamp')
                logon_age_days = None
                last_logon_dt = None
                
                if isinstance(last_logon, datetime):
                    if last_logon.year != 1601:
                        last_logon_dt = last_logon
                        last_logon_utc = last_logon.replace(tzinfo=None) if last_logon.tzinfo else last_logon
                        logon_age_days = (now - last_logon_utc).days
                elif last_logon:
                    last_logon_int = safe_int(last_logon)
                    if last_logon_int:
                        last_logon_dt = windows_timestamp_to_datetime(last_logon_int)
                        if last_logon_dt and last_logon_dt.year != 1601:
                            logon_age_days = (now - last_logon_dt).days
                
                # Check if privileged
                admin_count = get_attr(entry, 'adminCount', '')
                is_privileged = str(admin_count) == '1'
                
                # Check for SPNs
                spns = get_attr_list(entry, 'servicePrincipalName')
                has_spn = len(spns) > 0
                
                # Risk assessment
                risk_level = "Medium"
                risk_factors = ["AS-REP Roasting vulnerable (no pre-authentication required)"]
                
                if is_privileged:
                    risk_level = "CRITICAL"
                    risk_factors.append("Privileged account (AdminCount=1)")
                elif has_spn:
                    risk_level = "HIGH"
                    risk_factors.append("Has SPN (can also be Kerberoasted)")
                
                if pwd_age_days and pwd_age_days < 90:
                    risk_factors.append("Password recently changed")
                
                # Get group memberships
                member_of = get_attr_list(entry, 'memberOf')
                groups = []
                for group_dn in member_of:
                    match = re.search(r'CN=([^,]+)', str(group_dn))
                    if match:
                        groups.append(match.group(1))
                
                results.append({
                    "SAM Account Name": get_attr(entry, 'sAMAccountName', ''),
                    "Name": get_attr(entry, 'name', ''),
                    "Enabled": enabled,
                    "Risk Level": risk_level,
                    "Risk Factors": "; ".join(risk_factors),
                    "AdminCount": admin_count,
                    "Has SPN": "Yes" if has_spn else "No",
                    "Password Age (days)": pwd_age_days if pwd_age_days is not None else "Unknown",
                    "Last Logon":format_datetime(last_logon_dt) if last_logon_dt else "Never",
                    "Logon Age (days)": logon_age_days if logon_age_days is not None else "Unknown",
                    "Member Of": "; ".join(groups[:5]) if groups else "None",
                    "SID": sid_to_str(get_attr(entry, 'objectSid')),
                    "Description": get_attr(entry, 'description', ''),
                    "Info": get_attr(entry, 'info', ''),
                    "Created": format_datetime(self._parse_timestamp(get_attr(entry, 'whenCreated'))),
                    "Distinguished Name": get_attr(entry, 'distinguishedName', ''),
                })

        except Exception as e:
            logger.warning(f"Error collecting AS-REP roastable accounts: {e}")
            logger.debug(f"Exception details: ", exc_info=True)

        self.results['ASREPRoastable'] = results
        logger.info(f"    Found {len(results)} AS-REP roastable accounts")
        return results

    # Helper methods
    def _parse_timestamp(self, value):
        """Parse various timestamp formats to datetime."""
        if value is None or value == 0:
            return None
        
        # Handle Windows FILETIME (large integers)
        if isinstance(value, int) or (isinstance(value, str) and value.isdigit()):
            timestamp = int(value) if isinstance(value, str) else value
            if timestamp == 0 or timestamp == 9223372036854775807:
                return None
            return windows_timestamp_to_datetime(timestamp)
        
        # Handle generalized time strings
        if isinstance(value, str):
            return generalized_time_to_datetime(value)
        
        # Already a datetime
        if isinstance(value, datetime):
            return value
        
        return None

    def _build_hostname_to_ip_cache(self) -> Tuple[Dict[str, str], Dict[str, str]]:
        """
        Build hostname-to-IP mapping from already-collected DNS records.
        This avoids the need for live DNS lookups and works offline.
        
        First tries to use self.results['DNSRecords'], then falls back to reading
        the DNSRecords.csv file if it exists in the output directory.
        
        Returns:
            Tuple of (hostname_to_ipv4, hostname_to_ipv6) dictionaries
        """
        hostname_to_ip = {}
        hostname_to_ipv6 = {}
        
        # Try to use already-collected DNS records first
        if 'DNSRecords' in self.results and self.results['DNSRecords']:
            logger.debug(f"Building offline DNS cache from {len(self.results['DNSRecords'])} DNS records")
            for record in self.results['DNSRecords']:
                if record.get('RecordType') == 'A':
                    hostname = record.get('Name', '').rstrip('.')
                    ip = record.get('Data', '')
                    if hostname and ip:
                        hostname_to_ip[hostname.lower()] = ip
                elif record.get('RecordType') == 'AAAA':
                    hostname = record.get('Name', '').rstrip('.')
                    ipv6 = record.get('Data', '')
                    if hostname and ipv6:
                        hostname_to_ipv6[hostname.lower()] = ipv6
            
            logger.debug(f"Built offline DNS cache: {len(hostname_to_ip)} IPv4, {len(hostname_to_ipv6)} IPv6 records")
        else:
            # Try to read from CSV file if it exists (for offline analysis)
            csv_path = os.path.join(self.config.output_dir, 'CSV-Files', 'DNSRecords.csv')
            if os.path.exists(csv_path):
                try:
                    logger.debug(f"Reading DNS records from {csv_path} for offline lookup")
                    with open(csv_path, 'r', encoding='utf-8') as f:
                        reader = csv.DictReader(f)
                        for record in reader:
                            if record.get('RecordType') == 'A':
                                hostname = record.get('Name', '').rstrip('.')
                                ip = record.get('Data', '')
                                if hostname and ip:
                                    hostname_to_ip[hostname.lower()] = ip
                            elif record.get('RecordType') == 'AAAA':
                                hostname = record.get('Name', '').rstrip('.')
                                ipv6 = record.get('Data', '')
                                if hostname and ipv6:
                                    hostname_to_ipv6[hostname.lower()] = ipv6
                    
                    logger.debug(f"Built offline DNS cache from CSV: {len(hostname_to_ip)} IPv4, {len(hostname_to_ipv6)} IPv6 records")
                except Exception as e:
                    logger.debug(f"Could not read DNS records from CSV: {e}")
            else:
                logger.debug("No DNSRecords available for offline lookup (not yet collected and no CSV found)")
        
        return hostname_to_ip, hostname_to_ipv6

    def _convert_time_to_days(self, value) -> float:
        """Convert Windows time interval to days."""
        if value is None:
            return 0.0
        try:
            val = int(value) if not isinstance(value, int) else value
            # Windows time is in 100-nanosecond intervals, negative for durations
            return abs(val) * 1e-7 / 86400
        except (ValueError, TypeError):
            return 0.0

    def _convert_time_to_minutes(self, value) -> float:
        """Convert Windows time interval to minutes."""
        if value is None:
            return 0.0
        try:
            val = int(value) if not isinstance(value, int) else value
            return abs(val) * 1e-7 / 60
        except (ValueError, TypeError):
            return 0.0

    def _get_functional_level(self, level: int) -> str:
        """Get functional level string from integer."""
        return DOMAIN_FUNCTIONAL_LEVELS.get(level, f"Unknown ({level})")

    # Stub methods for modules not yet fully implemented
    def collect_trusts(self) -> List[Dict]:
        """Collect trust relationships."""
        logger.info("[-] Collecting Trust Relationships...")
        results = []

        try:
            # Using objectCategory=trustedDomain for slightly different query
            entries = self.search(
                search_base=self.base_dn,
                search_filter='(objectCategory=trustedDomain)',
                attributes=['distinguishedName', 'trustPartner', 'trustDirection', 'trustType',
                           'trustAttributes', 'whenCreated', 'whenChanged', 'securityIdentifier']
            )

            for entry in entries:
                
                # Get the trusted domain's SID and build mapping
                trusted_domain_sid_bytes = get_attr(entry, 'securityIdentifier')
                trusted_domain_name = get_attr(entry, 'trustPartner', '')
                
                if trusted_domain_sid_bytes and trusted_domain_name:
                    try:
                        # Convert SID to string format (handles both bytes and base64 strings)
                        trusted_domain_sid = sid_to_str(trusted_domain_sid_bytes)
                        # Store mapping for SID resolution
                        self.domain_sid_to_name[trusted_domain_sid] = trusted_domain_name
                        logger.debug(f"Mapped trusted domain SID {trusted_domain_sid} to {trusted_domain_name}")
                    except Exception as e:
                        logger.debug(f"Failed to map trusted domain SID: {e}")
                
                trust_dir = get_attr(entry, 'trustDirection', 0)
                trust_type = get_attr(entry, 'trustType', 0)
                trust_attrs = get_attr(entry, 'trustAttributes', 0)

                # Parse trust attributes
                attrs_list = []
                if trust_attrs:
                    trust_attrs = safe_int(trust_attrs)
                    if trust_attrs & 0x01: attrs_list.append("Non Transitive")
                    if trust_attrs & 0x02: attrs_list.append("UpLevel")
                    if trust_attrs & 0x04: attrs_list.append("Quarantined")
                    if trust_attrs & 0x08: attrs_list.append("Forest Transitive")
                    if trust_attrs & 0x10: attrs_list.append("Cross Organization")
                    if trust_attrs & 0x20: attrs_list.append("Within Forest")
                    if trust_attrs & 0x40: attrs_list.append("Treat as External")
                    if trust_attrs & 0x80: attrs_list.append("Uses RC4 Encryption")
                    if trust_attrs & 0x200: attrs_list.append("No TGT Delegation")

                results.append({
                    "Source Domain": dn_to_fqdn(str(get_attr(entry, 'distinguishedName', ''))),
                    "Target Domain": get_attr(entry, 'trustPartner', ''),
                    "Trust Direction": TRUST_DIRECTION.get(safe_int(trust_dir), "Unknown"),
                    "Trust Type": TRUST_TYPE.get(safe_int(trust_type), "Unknown"),
                    "Attributes": ", ".join(attrs_list),
                    "whenCreated": format_datetime(self._parse_timestamp(get_attr(entry, 'whenCreated'))),
                    "whenChanged": format_datetime(self._parse_timestamp(get_attr(entry, 'whenChanged'))),
                })

        except Exception as e:
            logger.warning(f"Error collecting trusts: {e}")

        self.results['Trusts'] = results
        logger.info(f"    Found {len(results)} trust relationships")
        return results

    def collect_sites(self) -> List[Dict]:
        """Collect AD sites."""
        logger.info("[-] Collecting Sites...")
        results = []

        try:
            entries = self.search(
                search_base=f"CN=Sites,{self.config_dn}",
                search_filter='(objectCategory=site)',
                attributes=['name', 'description', 'whenCreated', 'whenChanged']
            )

            for entry in entries:
                
                results.append({
                    "Name": get_attr(entry, 'name', ''),
                    "Description": get_attr(entry, 'description', ''),
                    "whenCreated": format_datetime(self._parse_timestamp(get_attr(entry, 'whenCreated'))),
                    "whenChanged": format_datetime(self._parse_timestamp(get_attr(entry, 'whenChanged'))),
                })

        except Exception as e:
            logger.warning(f"Error collecting sites: {e}")

        self.results['Sites'] = results
        logger.info(f"    Found {len(results)} sites")
        return results

    def collect_subnets(self) -> List[Dict]:
        """Collect AD subnets."""
        logger.info("[-] Collecting Subnets...")
        results = []

        try:
            entries = self.search(
                search_base=f"CN=Subnets,CN=Sites,{self.config_dn}",
                search_filter='(objectCategory=subnet)',
                attributes=['name', 'description', 'siteObject', 'whenCreated', 'whenChanged']
            )

            for entry in entries:
                
                site_obj = get_attr(entry, 'siteObject', '')
                site_name = ""
                if site_obj:
                    # Extract site name from DN
                    match = re.search(r'CN=([^,]+)', str(site_obj))
                    if match:
                        site_name = match.group(1)

                results.append({
                    "Site": site_name,
                    "Name": get_attr(entry, 'name', ''),
                    "Description": get_attr(entry, 'description', ''),
                    "whenCreated": format_datetime(self._parse_timestamp(get_attr(entry, 'whenCreated'))),
                    "whenChanged": format_datetime(self._parse_timestamp(get_attr(entry, 'whenChanged'))),
                })

        except Exception as e:
            logger.warning(f"Error collecting subnets: {e}")

        self.results['Subnets'] = results
        logger.info(f"    Found {len(results)} subnets")
        return results

    def collect_schema_history(self) -> List[Dict]:
        """Collect schema history/updates."""
        logger.info("[-] Collecting Schema History - May take some time...")
        results = []

        try:
            entries = self.search(
                search_base=self.schema_dn,
                search_filter='(objectCategory=*)',
                attributes=['name', 'objectClass', 'whenCreated', 'whenChanged', 'distinguishedName']
            )

            for entry in entries:
                
                obj_class = get_attr(entry, 'objectClass')
                if isinstance(obj_class, list):
                    obj_class = ", ".join([str(c) for c in obj_class])

                results.append({
                    "ObjectClass": obj_class,
                    "Name": get_attr(entry, 'name', ''),
                    "whenCreated": format_datetime(self._parse_timestamp(get_attr(entry, 'whenCreated'))),
                    "whenChanged": format_datetime(self._parse_timestamp(get_attr(entry, 'whenChanged'))),
                    "DistinguishedName": get_attr(entry, 'distinguishedName', ''),
                })

        except Exception as e:
            logger.warning(f"Error collecting schema history: {e}")

        self.results['SchemaHistory'] = results
        logger.info(f"    Found {len(results)} schema objects")
        return results

    def collect_fine_grained_password_policies(self) -> List[Dict]:
        """Collect fine-grained password policies."""
        logger.info("[-] Collecting Fine-Grained Password Policies...")
        results = []

        try:
            entries = self.search(
                search_base=f"CN=Password Settings Container,CN=System,{self.base_dn}",
                search_filter='(objectCategory=msDS-PasswordSettings)',
                attributes=['name', 'msDS-PasswordSettingsPrecedence', 'msDS-MinimumPasswordLength',
                           'msDS-PasswordHistoryLength', 'msDS-PasswordComplexityEnabled',
                           'msDS-PasswordReversibleEncryptionEnabled', 'msDS-MaximumPasswordAge',
                           'msDS-MinimumPasswordAge', 'msDS-LockoutThreshold', 'msDS-LockoutDuration',
                           'msDS-LockoutObservationWindow', 'msDS-PSOAppliesTo']
            )

            for entry in entries:
                
                applies_to = get_attr_list(entry, 'msDS-PSOAppliesTo')
                
                # Get boolean values and format consistently with LDAP tool
                complexity_enabled = get_attr(entry, 'msDS-PasswordComplexityEnabled', '')
                reversible_encryption = get_attr(entry, 'msDS-PasswordReversibleEncryptionEnabled', '')

                results.append({
                    "Name": get_attr(entry, 'name', ''),
                    "Precedence": str(get_attr(entry, 'msDS-PasswordSettingsPrecedence', '')),
                    "Min Password Length": str(get_attr(entry, 'msDS-MinimumPasswordLength', '')),
                    "Password History": str(get_attr(entry, 'msDS-PasswordHistoryLength', '')),
                    "Complexity Enabled": str(complexity_enabled).capitalize() if str(complexity_enabled).upper() in ['TRUE', 'FALSE'] else str(complexity_enabled),
                    "Reversible Encryption": str(reversible_encryption).capitalize() if str(reversible_encryption).upper() in ['TRUE', 'FALSE'] else str(reversible_encryption),
                    "Lockout Threshold": str(get_attr(entry, 'msDS-LockoutThreshold', '')),
                    "Applies To": ", ".join([str(a) for a in applies_to]) if applies_to else "",
                })

        except Exception as e:
            logger.warning(f"Error collecting FGPP: {e}")

        self.results['FineGrainedPasswordPolicy'] = results
        logger.info(f"    Found {len(results)} fine-grained password policies")
        return results

    def collect_dns_zones(self) -> List[Dict]:
        """Collect DNS zones."""
        logger.info("[-] Collecting DNS Zones...")
        results = []

        try:
            # Check multiple locations for DNS zones
            search_bases = [
                self.base_dn,
                f"DC=DomainDnsZones,{self.base_dn}",
                f"DC=ForestDnsZones,{self.forest_root_dn}",
            ]

            for search_base in search_bases:
                try:
                    entries = self.search(
                        search_base=search_base,
                        search_filter='(objectCategory=dnsZone)',
                        attributes=['name', 'distinguishedName', 'whenCreated', 'whenChanged']
                    )

                    for entry in entries:
                        
                        results.append({
                            "Name": get_attr(entry, 'name', ''),
                            "whenCreated": format_datetime(self._parse_timestamp(get_attr(entry, 'whenCreated'))),
                            "whenChanged": format_datetime(self._parse_timestamp(get_attr(entry, 'whenChanged'))),
                            "DistinguishedName": get_attr(entry, 'distinguishedName', ''),
                        })
                except Exception:
                    pass

        except Exception as e:
            logger.warning(f"Error collecting DNS zones: {e}")

        self.results['DNSZones'] = results
        logger.info(f"    Found {len(results)} DNS zones")
        return results

    def collect_dns_records(self) -> List[Dict]:
        """Collect DNS records with full parsing."""
        logger.info("[-] Collecting DNS Records - May take some time...")
        results = []

        def parse_dns_record(record_bytes):
            """Parse DNS record from binary format."""
            if not isinstance(record_bytes, bytes) or len(record_bytes) < 24:
                return None
            
            try:
                # DNS record header structure
                data_length = struct.unpack('<H', record_bytes[0:2])[0]
                record_type = struct.unpack('<H', record_bytes[2:4])[0]
                version = struct.unpack('B', record_bytes[4:5])[0]
                rank = struct.unpack('B', record_bytes[5:6])[0]
                flags = struct.unpack('<H', record_bytes[6:8])[0]
                serial = struct.unpack('<I', record_bytes[8:12])[0]
                ttl_seconds = struct.unpack('<I', record_bytes[12:16])[0]
                reserved = struct.unpack('<I', record_bytes[16:20])[0]
                timestamp = struct.unpack('<I', record_bytes[20:24])[0]
                
                # Calculate age from timestamp (hours since 1/1/1601)
                age = timestamp
                
                # Format timestamp
                if timestamp == 0:
                    timestamp_str = "[static]"
                else:
                    try:
                        # Windows timestamp: hours since 1/1/1601
                        base = datetime(1601, 1, 1)
                        ts_datetime = base + timedelta(hours=timestamp)
                        timestamp_str = ts_datetime.strftime('%m/%d/%Y %I:%M:%S %p')
                    except:
                        timestamp_str = "[static]"
                
                record_data = record_bytes[24:]
                data_str = ""
                type_name = ""
                
                # Parse different record types
                if record_type == 1:  # A record
                    type_name = "A"
                    if len(record_data) >= 4:
                        data_str = '.'.join(str(b) for b in record_data[0:4])
                
                elif record_type == 2:  # NS record
                    type_name = "NS"
                    data_str, _ = parse_dns_name(record_data, 2)
                
                elif record_type == 5:  # CNAME record
                    type_name = "CNAME"
                    data_str, _ = parse_dns_name(record_data, 2)
                
                elif record_type == 6:  # SOA record
                    type_name = "SOA"
                    if len(record_data) >= 20:
                        soa_serial = struct.unpack('>I', record_data[0:4])[0]
                        refresh = struct.unpack('>I', record_data[4:8])[0]
                        retry = struct.unpack('>I', record_data[8:12])[0]
                        expire = struct.unpack('>I', record_data[12:16])[0]
                        minimum = struct.unpack('>I', record_data[16:20])[0]
                        primary_ns, offset = parse_dns_name(record_data, 22)
                        admin_email, offset = parse_dns_name(record_data, offset + 2)
                        data_str = f"[{soa_serial}][{primary_ns}][{admin_email}][{refresh}][{retry}][{expire}][{minimum}]"
                
                elif record_type == 12:  # PTR record
                    type_name = "PTR"
                    data_str, _ = parse_dns_name(record_data, 2)
                
                elif record_type == 15:  # MX record
                    type_name = "MX"
                    if len(record_data) >= 2:
                        preference = struct.unpack('>H', record_data[0:2])[0]
                        exchange, _ = parse_dns_name(record_data, 2)
                        data_str = f"[{preference}][{exchange}]"
                
                elif record_type == 16:  # TXT record
                    type_name = "TXT"
                    if len(record_data) > 0:
                        txt_len = record_data[0]
                        if len(record_data) >= txt_len + 1:
                            data_str = record_data[1:txt_len+1].decode('utf-8', errors='ignore')
                
                elif record_type == 28:  # AAAA record (IPv6)
                    type_name = "AAAA"
                    if len(record_data) >= 16:
                        ipv6_parts = []
                        for i in range(0, 16, 2):
                            part = struct.unpack('>H', record_data[i:i+2])[0]
                            ipv6_parts.append(f"{part:04x}")
                        data_str = ':'.join(ipv6_parts)
                
                elif record_type == 33:  # SRV record
                    type_name = "SRV"
                    if len(record_data) >= 8:
                        priority = struct.unpack('>H', record_data[0:2])[0]
                        weight = struct.unpack('>H', record_data[2:4])[0]
                        port = struct.unpack('>H', record_data[4:6])[0]
                        target, _ = parse_dns_name(record_data, 8)
                        data_str = f"[{priority}][{weight}][{port}][{target}]"
                
                else:
                    type_name = f"TYPE{record_type}"
                    data_str = record_data.hex() if record_data else ""
                
                return {
                    'RecordType': type_name,
                    'Data': data_str,
                    'TTL': ttl_seconds,
                    'Age': age,
                    'TimeStamp': timestamp_str,
                    'UpdatedAtSerial': serial
                }
                
            except Exception as e:
                logger.debug(f"Error parsing DNS record: {e}")
                return None
        
        def parse_dns_name(data, offset=0):
            """Parse DNS name from record data in label format."""
            if not data or offset >= len(data):
                return "", offset
            
            try:
                name_parts = []
                jumped = False
                saved_offset = 0
                
                while offset < len(data):
                    if offset >= len(data):
                        break
                    
                    length = data[offset]
                    
                    if length == 0:
                        offset += 1
                        break
                    
                    if length >= 0xC0:
                        if not jumped:
                            saved_offset = offset + 2
                        if offset + 1 < len(data):
                            pointer = struct.unpack('>H', data[offset:offset+2])[0] & 0x3FFF
                            pointed_name, _ = parse_dns_name(data, pointer)
                            if pointed_name and pointed_name != '.':
                                name_parts.append(pointed_name.rstrip('.'))
                            jumped = True
                            offset = saved_offset
                        break
                    
                    offset += 1
                    if offset + length > len(data):
                        break
                    
                    label_bytes = data[offset:offset+length]
                    try:
                        label = label_bytes.decode('ascii', errors='replace')
                        label = ''.join(c for c in label if c.isprintable() or c in '.-_')
                        if label:
                            name_parts.append(label)
                    except:
                        pass
                    
                    offset += length
                
                if jumped and saved_offset > 0:
                    offset = saved_offset
                
                result = '.'.join(name_parts)
                if result and not result.endswith('.'):
                    result += '.'
                return result, offset
                
            except Exception as e:
                logger.debug(f"Error parsing DNS name: {e}")
                return "", offset
        
        def sanitize_value(value):
            """Sanitize value for Excel - remove illegal characters."""
            if isinstance(value, str):
                sanitized = ''.join(char for char in value if ord(char) >= 32 or char in '\t\n\r')
                sanitized = sanitized.replace('\x00', '').replace('\x0b', '').replace('\x0c', '')
                return sanitized
            return value

        try:
            # Get DNS records from each zone
            if 'DNSZones' not in self.results:
                self.collect_dns_zones()

            for zone in self.results.get('DNSZones', []):
                zone_name = zone.get('Name', '')
                zone_dn = zone.get('DistinguishedName', '')
                if not zone_dn:
                    continue

                try:
                    entries = self.search(
                        search_base=zone_dn,
                        search_filter='(objectCategory=dnsNode)',
                        attributes=['name', 'dnsRecord', 'dNSTombstoned', 'whenCreated', 'whenChanged',
                                   'showInAdvancedViewOnly', 'distinguishedName']
                    )

                    for entry in entries:
                        
                        name = get_attr(entry, 'name', '')
                        dns_records = get_attr_list(entry, 'dnsRecord')
                        tombstoned = get_attr(entry, 'dNSTombstoned', False)
                        when_created = format_datetime(self._parse_timestamp(get_attr(entry, 'whenCreated')))
                        when_changed = format_datetime(self._parse_timestamp(get_attr(entry, 'whenChanged')))
                        show_advanced = str(get_attr(entry, 'showInAdvancedViewOnly', 'TRUE')).upper()
                        dn = get_attr(entry, 'distinguishedName', '')
                        
                        if dns_records:
                            for record_bytes in dns_records:
                                # Convert string to bytes if needed (ADWS returns base64-encoded strings)
                                if isinstance(record_bytes, str):
                                    try:
                                        import base64
                                        record_bytes = base64.b64decode(record_bytes)
                                    except:
                                        logger.debug(f"Could not decode DNS record for {name}")
                                        continue
                                
                                parsed = parse_dns_record(record_bytes)
                                if parsed:
                                    results.append({
                                        "ZoneName": sanitize_value(zone_name),
                                        "Name": sanitize_value(name),
                                        "RecordType": sanitize_value(parsed['RecordType']),
                                        "Data": sanitize_value(parsed['Data']),
                                        "TTL": parsed['TTL'],
                                        "Age": parsed['Age'],
                                        "TimeStamp": sanitize_value(parsed['TimeStamp']),
                                        "UpdatedAtSerial": parsed['UpdatedAtSerial'],
                                        "whenCreated": sanitize_value(when_created),
                                        "whenChanged": sanitize_value(when_changed),
                                        "showInAdvancedViewOnly": sanitize_value(show_advanced),
                                        "DistinguishedName": sanitize_value(dn)
                                    })
                        else:
                            results.append({
                                "ZoneName": sanitize_value(zone_name),
                                "Name": sanitize_value(name),
                                "RecordType": "",
                                "Data": "",
                                "TTL": "",
                                "Age": "",
                                "TimeStamp": "",
                                "UpdatedAtSerial": "",
                                "whenCreated": sanitize_value(when_created),
                                "whenChanged": sanitize_value(when_changed),
                                "showInAdvancedViewOnly": sanitize_value(show_advanced),
                                "DistinguishedName": sanitize_value(dn)
                            })
                except Exception as e:
                    logger.debug(f"Error processing zone {zone_name}: {e}")

        except Exception as e:
            logger.warning(f"Error collecting DNS records: {e}")

        self.results['DNSRecords'] = results
        logger.info(f"    Found {len(results)} DNS records")
        return results

    def collect_printers(self) -> List[Dict]:
        """Collect printer objects."""
        logger.info("[-] Collecting Printers...")
        results = []

        try:
            entries = self.search(
                self.base_dn,
                "(objectCategory=printQueue)",
                ['printerName', 'name', 'serverName', 'portName', 'driverName',
                 'driverVersion', 'printShareName', 'uNCName', 'url',
                 'whenCreated', 'whenChanged']
            )

            for entry in entries:
                results.append({
                    "Name": get_attr(entry, 'name', ''),
                    "Server Name": get_attr(entry, 'serverName', ''),
                    "Share Name": get_attr(entry, 'printShareName', ''),
                    "Port Name": get_attr(entry, 'portName', ''),
                    "Driver Name": get_attr(entry, 'driverName', ''),
                    "Driver Version": get_attr(entry, 'driverVersion', ''),
                    "UNC Name": get_attr(entry, 'uNCName', ''),
                    "URL": get_attr(entry, 'url', ''),
                    "whenCreated": format_datetime(self._parse_timestamp(get_attr(entry, 'whenCreated'))),
                    "whenChanged": format_datetime(self._parse_timestamp(get_attr(entry, 'whenChanged'))),
                })

        except Exception as e:
            logger.warning(f"Error collecting printers: {e}")

        self.results['Printers'] = results
        logger.info(f"    Found {len(results)} printers")
        return results

    def collect_bitlocker(self) -> List[Dict]:
        """Collect BitLocker recovery keys."""
        logger.info("[-] Collecting BitLocker Recovery Keys...")
        results = []

        try:
            entries = self.search(
                search_base=self.base_dn,
                search_filter='(objectCategory=msFVE-RecoveryInformation)',
                attributes=['name', 'msFVE-RecoveryPassword', 'msFVE-RecoveryGuid', 
                           'whenCreated', 'distinguishedName']
            )

            for entry in entries:
                recovery_pwd = get_attr(entry, 'msFVE-RecoveryPassword', '')
                recovery_guid = get_attr(entry, 'msFVE-RecoveryGuid', '')
                when_created = format_datetime(self._parse_timestamp(get_attr(entry, 'whenCreated')))
                dn = get_attr(entry, 'distinguishedName', '')
                
                # Extract computer name from DN
                computer_name = ""
                if dn:
                    parts = dn.split(',')
                    for part in parts:
                        if part.startswith('CN=') and not part.startswith('CN={'):
                            computer_name = part[3:]
                            break

                results.append({
                    "Computer": computer_name,
                    "RecoveryPassword": recovery_pwd,
                    "RecoveryGuid": recovery_guid,
                    "whenCreated": when_created,
                    "DistinguishedName": dn
                })

        except Exception as e:
            logger.warning(f"Error collecting BitLocker: {e}")

        self.results['Bitlocker'] = results
        logger.info(f"    Found {len(results)} BitLocker recovery keys")
        return results

    def collect_dmsa(self) -> List[Dict]:
        """Collect Delegated Managed Service Accounts (dMSA) - Windows Server 2025+."""
        logger.info("[-] Collecting Delegated Managed Service Accounts (dMSA)...")
        results = []
        now = datetime.now(timezone.utc).replace(tzinfo=None)

        try:
            # First check if the schema supports dMSA (Windows Server 2025+ feature)
            try:
                schema_check = self.search(
                    self.schema_dn,
                    "(lDAPDisplayName=msDS-DelegatedManagedServiceAccount)",
                    ['lDAPDisplayName']
                )
                
                if not schema_check:
                    logger.warning("[*] dMSA not supported - requires Windows Server 2025+ AD schema")
                    self.results['dMSA'] = results
                    return results
            except Exception as e:
                logger.warning(f"[*] dMSA not supported in this AD environment: {e}")
                self.results['dMSA'] = results
                return results
            
            # Collect dMSA (Delegated Managed Service Account - Windows Server 2025+)
            dmsa_entries = self.search(
                self.base_dn,
                "(objectClass=msDS-DelegatedManagedServiceAccount)",
                ['sAMAccountName', 'name', 'distinguishedName', 'description',
                 'userAccountControl', 'pwdLastSet', 'lastLogonTimestamp',
                 'servicePrincipalName', 'objectSid', 'whenCreated', 'whenChanged',
                 'dNSHostName', 'msDS-AllowedToDelegateTo', 'msDS-ManagedAccountPrecededByLink',
                 'msDS-DelegatedMSAState']
            )

            # Process dMSA entries
            if dmsa_entries:
                for entry in dmsa_entries:
                    uac = get_attr(entry, 'userAccountControl', 0)
                    enabled = "TRUE" if uac and not (int(uac) & 0x0002) else "FALSE"

                    # Password last set
                    pwd_last_set = get_attr(entry, 'pwdLastSet')
                    pwd_last_set_dt = None
                    if isinstance(pwd_last_set, datetime):
                        pwd_last_set_dt = pwd_last_set.replace(tzinfo=None) if pwd_last_set.tzinfo else pwd_last_set
                    elif pwd_last_set:
                        # Convert Windows FileTime integer to datetime
                        pwd_last_set_int = safe_int(pwd_last_set)
                        if pwd_last_set_int:
                            pwd_last_set_dt = windows_timestamp_to_datetime(pwd_last_set_int)
                    
                    pwd_age_days = None
                    if pwd_last_set_dt:
                        pwd_age = now - pwd_last_set_dt
                        pwd_age_days = pwd_age.days

                    # Last logon
                    last_logon = get_attr(entry, 'lastLogonTimestamp')
                    last_logon_dt = None
                    if isinstance(last_logon, datetime):
                        last_logon_dt = last_logon.replace(tzinfo=None) if last_logon.tzinfo else last_logon
                    elif last_logon:
                        # Convert Windows FileTime integer to datetime
                        last_logon_int = safe_int(last_logon)
                        if last_logon_int:
                            last_logon_dt = windows_timestamp_to_datetime(last_logon_int)
                    
                    logon_age_days = None
                    if last_logon_dt:
                        logon_age = now - last_logon_dt
                        logon_age_days = logon_age.days

                    # Check for badSuccessor vulnerability (dMSA specific)
                    superseded_account = get_attr(entry, 'msDS-ManagedAccountPrecededByLink', '')
                    bad_successor_vuln = bool(superseded_account)
                    dmsa_state = get_attr(entry, 'msDS-DelegatedMSAState', '')

                    # SPNs
                    spns = get_attr_list(entry, 'servicePrincipalName')
                    spns_str = ", ".join(spns) if spns else ""

                    # Delegation
                    allowed_to_delegate = get_attr_list(entry, 'msDS-AllowedToDelegateTo')
                    delegation_str = ", ".join(allowed_to_delegate) if allowed_to_delegate else ""

                    results.append({
                        "SAM Account Name": get_attr(entry, 'sAMAccountName', ''),
                        "Name": get_attr(entry, 'name', ''),
                        "Enabled": enabled,
                        "Description": get_attr(entry, 'description', ''),
                        "DNS Hostname": get_attr(entry, 'dNSHostName', ''),
                        "SPNs": spns_str,
                        "Password Age (days)": pwd_age_days,
                        "Logon Age (days)": logon_age_days,
                        "Delegation Services": delegation_str,
                        "Superseded Account (BadSuccessor)": superseded_account,
                        "dMSA State": dmsa_state,
                        "badSuccessor Vulnerable": bad_successor_vuln,
                        "Risk Level": "CRITICAL" if bad_successor_vuln else "Low",
                        "SID": sid_to_str(get_attr(entry, 'objectSid')),
                        "Password LastSet": format_datetime(pwd_last_set_dt),
                        "Last Logon Date": format_datetime(last_logon_dt),
                        "whenCreated": format_datetime(get_attr(entry, 'whenCreated')),
                        "whenChanged": format_datetime(get_attr(entry, 'whenChanged')),
                        "Distinguished Name": get_attr(entry, 'distinguishedName', ''),
                    })

        except Exception as e:
            logger.warning(f"Error collecting dMSA accounts: {e}")
            logger.debug(f"Exception details: ", exc_info=True)

        self.results['dMSA'] = results
        logger.info(f"    Found {len(results)} dMSA accounts")
        return results

    def collect_gmsa(self) -> List[Dict]:
        """Collect Group Managed Service Accounts (gMSA)."""
        logger.info("[-] Collecting Group Managed Service Accounts (gMSA)...")
        results = []
        now = datetime.now(timezone.utc).replace(tzinfo=None)

        try:
            # Optional schema check (informational only)
            try:
                schema_check = self.search(
                    self.schema_dn,
                    "(lDAPDisplayName=msDS-GroupManagedServiceAccount)",
                    ['lDAPDisplayName']
                )
                
                if not schema_check:
                    logger.debug("[*] gMSA schema class not found in schema, but will still attempt to search for gMSA objects")
            except Exception as e:
                logger.debug(f"[*] Could not check schema for gMSA support: {e}")
            
            # Collect gMSA accounts
            gmsa_entries = self.search(
                self.base_dn,
                "(objectClass=msDS-GroupManagedServiceAccount)",
                ['sAMAccountName', 'name', 'distinguishedName', 'description',
                 'userAccountControl', 'pwdLastSet', 'lastLogonTimestamp',
                 'servicePrincipalName', 'objectSid', 'whenCreated', 'whenChanged',
                 'dNSHostName', 'msDS-AllowedToDelegateTo',
                 'msDS-ManagedPasswordInterval', 'msDS-GroupMSAMembership',
                 'nTSecurityDescriptor']
            )

            # Process gMSA entries
            if gmsa_entries:
                for entry in gmsa_entries:
                    uac = get_attr(entry, 'userAccountControl', 0)
                    enabled = "True" if uac and not (int(uac) & 0x0002) else "False"

                    # Password last set
                    pwd_last_set = get_attr(entry, 'pwdLastSet')
                    pwd_last_set_dt = None
                    if isinstance(pwd_last_set, datetime):
                        pwd_last_set_dt = pwd_last_set.replace(tzinfo=None) if pwd_last_set.tzinfo else pwd_last_set
                    elif pwd_last_set:
                        # Convert Windows FileTime integer to datetime
                        pwd_last_set_int = safe_int(pwd_last_set)
                        if pwd_last_set_int:
                            pwd_last_set_dt = windows_timestamp_to_datetime(pwd_last_set_int)
                    
                    pwd_age_days = None
                    if pwd_last_set_dt:
                        pwd_age_delta = now - pwd_last_set_dt
                        pwd_age_days = pwd_age_delta.days

                    # Last logon
                    last_logon = get_attr(entry, 'lastLogonTimestamp')
                    last_logon_dt = None
                    if isinstance(last_logon, datetime):
                        last_logon_dt = last_logon.replace(tzinfo=None) if last_logon.tzinfo else last_logon
                    elif last_logon:
                        # Convert Windows FileTime integer to datetime
                        last_logon_int = safe_int(last_logon)
                        if last_logon_int:
                            last_logon_dt = windows_timestamp_to_datetime(last_logon_int)
                    
                    logon_age_days = None
                    if last_logon_dt:
                        logon_age_delta = now - last_logon_dt
                        logon_age_days = logon_age_delta.days

                    # Password rotation interval (in days)
                    pwd_interval = get_attr(entry, 'msDS-ManagedPasswordInterval', 30)
                    pwd_interval_days = int(pwd_interval) if pwd_interval else 30

                    # SPNs
                    spns = get_attr_list(entry, 'servicePrincipalName')
                    spns_str = ", ".join(spns) if spns else ""

                    # Delegation
                    allowed_to_delegate = get_attr_list(entry, 'msDS-AllowedToDelegateTo')
                    delegation_str = ", ".join(allowed_to_delegate) if allowed_to_delegate else ""

                    # Parse security descriptor for owner
                    owner_principal = None
                    sd_data = get_attr(entry, 'nTSecurityDescriptor')
                    if sd_data and isinstance(sd_data, bytes) and IMPACKET_AVAILABLE:
                        try:
                            sd = ldaptypes.SR_SECURITY_DESCRIPTOR(data=sd_data)
                            if sd['OwnerSid']:
                                owner_sid = sd['OwnerSid'].formatCanonical()
                                owner_principal = self._resolve_sid_to_name(owner_sid)
                        except Exception as e:
                            logger.debug(f"Error parsing gMSA owner: {e}")
                    
                    # Parse who can retrieve the managed password from msDS-GroupMSAMembership
                    allowed_principals = []
                    try:
                        membership_sd = get_attr(entry, 'msDS-GroupMSAMembership')
                        logger.debug(f"gMSA {get_attr(entry, 'sAMAccountName', 'Unknown')}: msDS-GroupMSAMembership type={type(membership_sd)}, value={membership_sd[:50] if membership_sd else None}")
                        
                        if membership_sd:
                            if isinstance(membership_sd, bytes) and IMPACKET_AVAILABLE:
                                try:
                                    sd = ldaptypes.SR_SECURITY_DESCRIPTOR(data=membership_sd)
                                    logger.debug(f"  Parsed SD, has DACL: {sd['Dacl'] is not None}")
                                    if sd['Dacl']:
                                        logger.debug(f"  DACL has {len(sd['Dacl'].aces)} ACEs")
                                        for ace in sd['Dacl'].aces:
                                            try:
                                                # ACCESS_ALLOWED_ACE_TYPE = 0
                                                if ace['AceType'] == 0:
                                                    sid = ace['Ace']['Sid'].formatCanonical()
                                                    principal_name = self._resolve_sid_to_name(sid)
                                                    logger.debug(f"    Found principal: {principal_name}")
                                                    allowed_principals.append(principal_name)
                                            except Exception as e2:
                                                logger.debug(f"    Error parsing ACE: {e2}")
                                                continue
                                except Exception as e:
                                    logger.debug(f"  Error parsing gMSA membership SD: {e}")
                            else:
                                logger.debug(f"  msDS-GroupMSAMembership not bytes or impacket unavailable (bytes={isinstance(membership_sd, bytes)}, impacket={IMPACKET_AVAILABLE})")
                        else:
                            logger.debug(f"  msDS-GroupMSAMembership is None/empty")
                    except Exception as e:
                        logger.debug(f"Error getting gMSA membership: {e}")
                    
                    allowed_principals_str = ", ".join(allowed_principals) if allowed_principals else "Not readable"

                    # Parse write permissions (who can modify the account)
                    write_principals = self._parse_write_permissions(entry)
                    write_perms_str = ", ".join(write_principals) if write_principals else "None"
                    
                    owner_str = owner_principal if owner_principal else "Unknown"

                    results.append({
                        "SAM Account Name": get_attr(entry, 'sAMAccountName', ''),
                        "Name": get_attr(entry, 'name', ''),
                        "Enabled": enabled,
                        "Owner": owner_str,
                        "Description": get_attr(entry, 'description', ''),
                        "DNS Hostname": get_attr(entry, 'dNSHostName', ''),
                        "SPNs": spns_str,
                        "Password Age (days)": pwd_age_days,
                        "Password Rotation Interval (days)": pwd_interval_days,
                        "Logon Age (days)": logon_age_days,
                        "Delegation Services": delegation_str,
                        "Allowed To Retrieve Password": allowed_principals_str,
                        "Write Permissions": write_perms_str,
                        "SID": sid_to_str(get_attr(entry, 'objectSid')),
                        "Password LastSet": format_datetime(pwd_last_set_dt),
                        "Last Logon Date": format_datetime(last_logon_dt),
                        "whenCreated": format_datetime(self._parse_timestamp(get_attr(entry, 'whenCreated'))),
                        "whenChanged": format_datetime(self._parse_timestamp(get_attr(entry, 'whenChanged'))),
                        "Distinguished Name": get_attr(entry, 'distinguishedName', ''),
                    })

        except Exception as e:
            logger.warning(f"Error collecting gMSA accounts: {e}")
            logger.debug(f"Exception details: ", exc_info=True)

        self.results['gMSA'] = results
        logger.info(f"    Found {len(results)} gMSA accounts")
        return results

    def collect_certificate_templates(self) -> List[Dict]:
        """Collect AD Certificate Services certificate templates."""
        logger.info("[-] Collecting ADCS Certificate Templates...")
        
        # Warn if querying from child domain - ACLs may be incomplete
        if hasattr(self, 'forest_root_dn') and self.forest_root_dn != self.base_dn:
            logger.warning("[!] Connected to child domain - certificate template ACLs may be incomplete!")
            logger.warning("    For complete ACL data, connect to forest root DC instead.")
            logger.warning(f"    Forest root: {dn_to_fqdn(self.forest_root_dn)}")
        
        results = []

        try:
            # Certificate templates are stored in CN=Certificate Templates,CN=Public Key Services,CN=Services,CN=Configuration
            pki_dn = f"CN=Certificate Templates,CN=Public Key Services,CN=Services,{self.config_dn}"
            
            # Search for certificate templates
            self.search(
                search_base=pki_dn,
                search_filter='(objectClass=pKICertificateTemplate)',
                attributes=['cn', 'displayName', 'distinguishedName', 'whenCreated', 'whenChanged',
                           'pKIDefaultKeySpec', 'pKIMaxIssuingDepth', 'pKIExpirationPeriod',
                           'pKIOverlapPeriod', 'pKICriticalExtensions', 'pKIExtendedKeyUsage',
                           'msPKI-Certificate-Name-Flag', 'msPKI-Enrollment-Flag',
                           'msPKI-Private-Key-Flag', 'msPKI-Certificate-Application-Policy',
                           'msPKI-RA-Signature', 'pKIDefaultCSPs', 'flags',
                           'nTSecurityDescriptor', 'msPKI-Template-Schema-Version',
                           'msPKI-Template-Minor-Revision', 'msPKI-Cert-Template-OID']
            )
            entries = self.conn.entries if hasattr(self.conn, 'entries') else []

            for entry in entries:
                # Parse enrollment permissions and owner from security descriptor
                enrollment_principals, auto_enrollment_principals, owner_principal = self._parse_enrollment_rights(entry)
                
                # Parse write permissions for ESC4
                write_principals = self._parse_write_permissions(entry)
                
                # Parse flags
                enrollment_flag = safe_int(get_attr(entry, 'msPKI-Enrollment-Flag', 0))
                cert_name_flag = safe_int(get_attr(entry, 'msPKI-Certificate-Name-Flag', 0))
                private_key_flag = safe_int(get_attr(entry, 'msPKI-Private-Key-Flag', 0))
                
                # Check for vulnerable configurations
                enrollee_supplies_subject = bool(cert_name_flag & 0x00000001)  # CT_FLAG_ENROLLEE_SUPPLIES_SUBJECT
                allows_san = bool(cert_name_flag & 0x00010000)  # CT_FLAG_SUBJECT_ALT_REQUIRE_UPN/DNS/EMAIL
                
                # Check authorization flags
                no_security_extension = bool(enrollment_flag & 0x80000000)  # ESC9
                auto_enroll = bool(enrollment_flag & 0x00000020)
                requires_manager_approval = bool(enrollment_flag & 0x00000002)
                
                # Check if authorized signatures are required (ESC3 check)
                ra_signature_count = safe_int(get_attr(entry, 'msPKI-RA-Signature', 0))
                requires_authorized_signatures = ra_signature_count > 0
                
                # Check if exportable
                exportable_key = bool(private_key_flag & 0x00000010)  # CT_FLAG_EXPORTABLE_KEY
                
                # Get extended key usage
                eku = get_attr(entry, 'pKIExtendedKeyUsage', [])
                if isinstance(eku, str):
                    eku = [eku]
                eku_str = ', '.join(eku) if eku else ''
                
                # ESC vulnerability checks
                # ESC1: Enrollee supplies subject + Client Authentication
                has_client_auth = '1.3.6.1.5.5.7.3.2' in eku_str or not eku_str
                
                # ESC2: Any Purpose EKU (allows any usage including client auth)
                has_any_purpose_eku = '2.5.29.37.0' in eku_str
                
                # ESC3: Certificate Request Agent / Enrollment Agent
                has_enrollment_agent = '1.3.6.1.4.1.311.20.2.1' in eku_str
                
                # ESC9: No security extension (StrongCertificateBindingEnforcement not set)
                vulnerable_to_esc9 = no_security_extension
                
                # Determine risk level
                risk_level = 'Low'
                risk_factors = []
                
                # Check if low-privileged users can enroll using SID-based detection
                low_priv_can_enroll = False
                only_admins_can_enroll = False
                
                if enrollment_principals:
                    for principal in enrollment_principals:
                        # Use SID-based low-privilege detection
                        if self._is_low_privilege_principal(principal):
                            low_priv_can_enroll = True
                            break
                    
                    # Check if only administrative principals have enrollment rights
                    if not low_priv_can_enroll:
                        admin_keywords = [
                            # Domain-level admin groups
                            'domain admins', 'enterprise admins', 'schema admins', 'administrator',
                            # Built-in privileged groups
                            'builtin\\administrators', 'account operators', 'server operators', 
                            'backup operators', 'print operators', 'group policy creator owners',
                            # Service and infrastructure groups
                            'ras and ias servers', 'cert publishers', 'cryptographic operators',
                            'network configuration operators', 'dnsadmins', 'dns admins',
                            'domain controllers', 'enterprise read-only domain controllers',
                            'key admins', 'enterprise key admins', 'protected users',
                            # Well-known privileged SIDs
                            's-1-5-9',      # Enterprise Domain Controllers
                            's-1-5-18',     # Local System
                            's-1-5-19',     # Local Service
                            's-1-5-20',     # Network Service
                            's-1-5-32-544', # Administrators
                            's-1-5-32-548', # Account Operators
                            's-1-5-32-549', # Server Operators
                            's-1-5-32-550', # Print Operators
                            's-1-5-32-551'  # Backup Operators
                        ]
                        only_admins_can_enroll = all(
                            any(admin_keyword in principal.lower() for admin_keyword in admin_keywords)
                            for principal in enrollment_principals
                        )
                
                # Check for non-admin write permissions (ESC4) using SID-based detection
                low_priv_can_write = False
                for principal in write_principals:
                    if self._is_low_privilege_principal(principal):
                        low_priv_can_write = True
                        break
                
                # Risk assessment with ESC1-4 and ESC9
                esc_vulns = []
                
                # ESC1: Domain escalation via Subject Alternative Name
                if enrollee_supplies_subject and has_client_auth and enrollment_principals and not only_admins_can_enroll:
                    esc_vulns.append('ESC1')
                    if requires_manager_approval:
                        # Manager approval mitigates but doesn't eliminate the risk
                        if risk_level == 'Low':
                            risk_level = 'MEDIUM'
                        risk_factors.append('ESC1: Enrollee supplies subject + Client Auth (mitigated by manager approval)')
                    elif low_priv_can_enroll:
                        risk_level = 'CRITICAL'
                        risk_factors.append('ESC1: Enrollee supplies subject + Client Auth + Low-priv enrollment')
                    else:
                        risk_level = 'HIGH'
                        risk_factors.append('ESC1: Enrollee supplies subject + Client Auth')
                
                # ESC2: Any Purpose EKU
                if has_any_purpose_eku and enrollment_principals and not only_admins_can_enroll:
                    esc_vulns.append('ESC2')
                    if requires_manager_approval:
                        # Manager approval mitigates but doesn't eliminate the risk
                        if risk_level == 'Low':
                            risk_level = 'Low'
                        risk_factors.append('ESC2: Any Purpose EKU (mitigated by manager approval requirement)')
                    elif low_priv_can_enroll:
                        if risk_level != 'CRITICAL':
                            risk_level = 'HIGH'
                        risk_factors.append('ESC2: Any Purpose EKU allows arbitrary certificate usage')
                    else:
                        if risk_level not in ['CRITICAL', 'HIGH']:
                            risk_level = 'MEDIUM'
                        risk_factors.append('ESC2: Any Purpose EKU allows arbitrary certificate usage')
                
                # ESC3: Enrollment Agent template
                # ESC3 requires: Certificate Request Agent EKU + No authorized signatures + No manager approval
                if has_enrollment_agent and enrollment_principals and not only_admins_can_enroll:
                    esc_vulns.append('ESC3')
                    # Check if both manager approval AND authorized signatures are missing (both mitigations absent)
                    if not requires_manager_approval and not requires_authorized_signatures:
                        # No mitigations in place - this is exploitable
                        if low_priv_can_enroll:
                            if risk_level not in ['CRITICAL', 'HIGH']:
                                risk_level = 'HIGH'
                            risk_factors.append('ESC3: Certificate Request Agent EKU with no approval/signature requirements')
                        else:
                            if risk_level not in ['CRITICAL', 'HIGH']:
                                risk_level = 'MEDIUM'
                            risk_factors.append('ESC3: Certificate Request Agent EKU with no approval/signature requirements')
                    elif requires_manager_approval and requires_authorized_signatures:
                        # Both mitigations present
                        if risk_level == 'Low':
                            risk_level = 'Low'
                        risk_factors.append('ESC3: Certificate Request Agent EKU (mitigated by manager approval + authorized signatures)')
                    elif requires_manager_approval or requires_authorized_signatures:
                        # One mitigation present
                        if risk_level == 'Low':
                            risk_level = 'Low'
                        mitigation = 'manager approval' if requires_manager_approval else 'authorized signatures'
                        risk_factors.append(f'ESC3: Certificate Request Agent EKU (partially mitigated by {mitigation})')
                    else:
                        # Edge case - shouldn't reach here
                        if risk_level not in ['CRITICAL', 'HIGH']:
                            risk_level = 'MEDIUM'
                        risk_factors.append('ESC3: Certificate Request Agent EKU')
                
                # ESC4: Vulnerable template ACLs
                if low_priv_can_write:
                    # ESC4 is not mitigated by manager approval - attackers can modify the template to disable it
                    if risk_level not in ['CRITICAL', 'HIGH']:
                        risk_level = 'HIGH'
                    esc_vulns.append('ESC4')
                    risk_factors.append('ESC4: Non-admin principals have write access to template')
                
                # ESC9: No security extension + Client Auth + No approvals + Low-priv enrollment
                # Requires ALL of: CT_FLAG_NO_SECURITY_EXTENSION, Client Authentication EKU,
                # No Manager Approval, No Authorized Signatures, Low-privileged enrollment
                if (vulnerable_to_esc9 and has_client_auth and 
                    not requires_manager_approval and not requires_authorized_signatures and 
                    low_priv_can_enroll):
                    esc_vulns.append('ESC9')
                    if risk_level not in ['CRITICAL', 'HIGH']:
                        risk_level = 'HIGH'
                    risk_factors.append('ESC9: No security extension + Client Auth + No approvals (allows certificate theft & impersonation)')
                
                # Additional risk factors
                if only_admins_can_enroll and (enrollee_supplies_subject and has_client_auth):
                    risk_factors.append('ESC1 config present but only admins can enroll')
                
                # Only flag as Medium if non-admins can actually enroll
                if enrollee_supplies_subject and not has_client_auth and risk_level == 'Low' and not only_admins_can_enroll:
                    risk_level = 'Medium'
                    risk_factors.append('Enrollee can supply subject name')
                elif allows_san and risk_level == 'Low' and not only_admins_can_enroll:
                    risk_level = 'Medium'
                    risk_factors.append('Allows Subject Alternative Name')
                    
                if exportable_key and risk_level != 'Low':
                    risk_factors.append('Private key exportable')
                
                if low_priv_can_enroll and enrollment_principals and risk_level == 'Low':
                    risk_factors.append('Low-privileged users can enroll')
                
                # Check if owner is low-privileged (potential ESC4 variant)
                if owner_principal and self._is_low_privilege_principal(owner_principal):
                    # ESC4 is not mitigated by manager approval - owner can modify the template
                    if risk_level not in ['CRITICAL', 'HIGH']:
                        risk_level = 'HIGH'
                    if 'ESC4' not in esc_vulns:
                        esc_vulns.append('ESC4')
                    risk_factors.append('ESC4: Low-privileged owner can modify template')
                
                # If no ESC vulnerabilities and no risk factors, set risk level to None
                if not esc_vulns and not risk_factors:
                    risk_level = 'None'
                    
                # Format enrollment and write permissions
                enroll_perms_str = '; '.join(enrollment_principals) if enrollment_principals else 'None'
                auto_enroll_perms_str = '; '.join(auto_enrollment_principals) if auto_enrollment_principals else 'None'
                write_perms_str = '; '.join(write_principals) if write_principals else 'None'
                esc_list = ', '.join(esc_vulns) if esc_vulns else 'None'
                owner_str = owner_principal if owner_principal else 'Unknown'
                    
                results.append({
                    "Template Name": get_attr(entry, 'cn', ''),
                    "Display Name": get_attr(entry, 'displayName', ''),
                    "Distinguished Name": get_attr(entry, 'distinguishedName', ''),
                    "Owner": owner_str,
                    "Schema Version": safe_int(get_attr(entry, 'msPKI-Template-Schema-Version', 0)),
                    "Created": format_datetime(self._parse_timestamp(get_attr(entry, 'whenCreated'))),
                    "Modified": format_datetime(self._parse_timestamp(get_attr(entry, 'whenChanged'))),
                    "Extended Key Usage": eku_str,
                    "Enrollee Supplies Subject": enrollee_supplies_subject,
                    "Allows SAN": allows_san,
                    "Client Authentication": has_client_auth,
                    "Any Purpose EKU": has_any_purpose_eku,
                    "Enrollment Agent": has_enrollment_agent,
                    "Exportable Key": exportable_key,
                    "Auto-Enrollment": auto_enroll,
                    "Requires Manager Approval": bool(enrollment_flag & 0x00000002),
                    "Authorized Signatures Required": ra_signature_count,
                    "Enrollment Flag": hex(enrollment_flag),
                    "Certificate Name Flag": hex(cert_name_flag),
                    "Private Key Flag": hex(private_key_flag),
                    "Enrollment Rights": enroll_perms_str,
                    "Auto-Enrollment Rights": auto_enroll_perms_str,
                    "Write Permissions": write_perms_str,
                    "ESC Vulnerabilities": esc_list,
                    "Risk Level": risk_level,
                    "Risk Factors": '; '.join(risk_factors) if risk_factors else 'None',
                })

        except Exception as e:
            logger.warning(f"Error collecting certificate templates: {e}")
            logger.debug(f"Exception details: ", exc_info=True)

        logger.info(f"    Found {len(results)} certificate templates")
        return results

    def collect_certificate_authorities(self) -> List[Dict]:
        """Collect AD Certificate Services certificate authorities."""
        logger.info("[-] Collecting ADCS Certificate Authorities...")
        results = []

        try:
            # CAs are stored in CN=Enrollment Services,CN=Public Key Services,CN=Services,CN=Configuration
            pki_dn = f"CN=Enrollment Services,CN=Public Key Services,CN=Services,{self.config_dn}"
            
            self.search(
                search_base=pki_dn,
                search_filter='(objectClass=pKIEnrollmentService)',
                attributes=['cn', 'displayName', 'distinguishedName', 'dNSHostName',
                           'whenCreated', 'whenChanged', 'cACertificate', 'certificateTemplates',
                           'msPKI-Enrollment-Servers', 'flags']
            )
            entries = self.conn.entries if hasattr(self.conn, 'entries') else []

            for entry in entries:
                # Get certificate templates this CA issues
                templates = get_attr(entry, 'certificateTemplates', [])
                if isinstance(templates, str):
                    templates = [templates]
                templates_str = ', '.join(templates) if templates else 'None'
                
                results.append({
                    "CA Name": get_attr(entry, 'cn', ''),
                    "Display Name": get_attr(entry, 'displayName', ''),
                    "DNS Hostname": get_attr(entry, 'dNSHostName', ''),
                    "Distinguished Name": get_attr(entry, 'distinguishedName', ''),
                    "Created": format_datetime(self._parse_timestamp(get_attr(entry, 'whenCreated'))),
                    "Modified": format_datetime(self._parse_timestamp(get_attr(entry, 'whenChanged'))),
                    "Certificate Templates": templates_str,
                    "Template Count": len(templates),
                })

        except Exception as e:
            logger.warning(f"Error collecting certificate authorities: {e}")
            logger.debug(f"Exception details: ", exc_info=True)

        logger.info(f"    Found {len(results)} certificate authorities")
        return results

    def collect_protected_groups(self) -> List[Dict]:
        """Collect protected groups and AdminSDHolder information."""
        logger.info("[-] Collecting Protected Groups & AdminSDHolder...")
        results = []

        try:
            # Well-known protected groups (these inherit from AdminSDHolder)
            protected_groups = [
                'Account Operators',
                'Administrators',
                'Backup Operators',
                'Domain Admins',
                'Domain Controllers',
                'Enterprise Admins',
                'Print Operators',
                'Read-only Domain Controllers',
                'Replicator',
                'Schema Admins',
                'Server Operators',
                'Key Admins',
                'Enterprise Key Admins',
                'Cert Publishers',
                'DnsAdmins',
                'Group Policy Creator Owners',
                'Incoming Forest Trust Builders',
                'Network Configuration Operators',
                'Windows Authorization Access Group',
                'Terminal Server License Servers',
                'RAS and IAS Servers',
                'Allowed RODC Password Replication Group',
                'Denied RODC Password Replication Group',
                'Cloneable Domain Controllers',
                'Protected Users',
                'Cryptographic Operators'
            ]
            
            # Well-known built-in accounts that have adminCount=1 by design
            builtin_accounts = [
                'krbtgt',
                'Administrator'
            ]
            
            # Service accounts that should NOT be in Protected Users (incompatible)
            service_accounts_exclusions = [
                'krbtgt'
            ]
            
            # Highly privileged groups that should be in Protected Users group
            highly_privileged_groups = [
                'Domain Admins',
                'Enterprise Admins',
                'Schema Admins',
                'Administrators'
            ]
            
            # First, get all users/groups with adminCount=1
            entries = self.search(
                self.base_dn,
                "(&(adminCount=1)(|(objectCategory=person)(objectCategory=group)))",
                ['sAMAccountName', 'name', 'objectCategory', 'distinguishedName', 
                 'memberOf', 'objectSid', 'whenChanged', 'lastLogonTimestamp']
            )
            
            for entry in entries:
                obj_category = str(get_attr(entry, 'objectCategory', '')).lower()
                is_user = 'person' in obj_category
                obj_type = "User" if is_user else "Group"
                
                sam_account = get_attr(entry, 'sAMAccountName', '')
                
                # Check if this is the built-in Administrator account (RID 500)
                is_builtin_administrator = False
                obj_sid = get_attr(entry, 'objectSid')
                if obj_sid:
                    sid_str = sid_to_str(obj_sid)
                    # Built-in Administrator always has RID 500 (ends with -500)
                    if sid_str and sid_str.endswith('-500'):
                        is_builtin_administrator = True
                name = get_attr(entry, 'name', '')
                member_of = get_attr_list(entry, 'memberOf')
                
                # Check if this GROUP object IS one of the protected groups
                is_protected_group_object = not is_user and name in protected_groups
                
                # Check if this is a built-in account (like krbtgt)
                is_builtin_account = is_user and sam_account.lower() in [a.lower() for a in builtin_accounts]
                is_service_account_exclusion = is_user and sam_account.lower() in [a.lower() for a in service_accounts_exclusions]
                
                # Check if this account is in a protected group (for users)
                in_protected_group = False
                protected_group_membership = []
                in_highly_privileged_group = False
                in_protected_users_group = False
                
                if is_user:
                    # For users, check their group memberships
                    for group_dn in member_of:
                        # Extract CN from DN
                        match = re.search(r'CN=([^,]+)', str(group_dn))
                        if match:
                            group_name = match.group(1)
                            if group_name in protected_groups:
                                in_protected_group = True
                                protected_group_membership.append(group_name)
                            if group_name in highly_privileged_groups:
                                in_highly_privileged_group = True
                            if group_name == 'Protected Users':
                                in_protected_users_group = True
                
                # Determine status
                if is_protected_group_object:
                    # This IS a protected group object itself
                    status = "Protected Group (Built-in)"
                elif is_builtin_account:
                    # Built-in account like krbtgt
                    status = "Built-in Account"
                elif is_user and in_protected_group:
                    # User is member of protected groups
                    status = "Active (Member of Protected Groups)"
                elif is_user and not in_protected_group:
                    # User has adminCount=1 but not in protected groups (orphaned)
                    status = "Orphaned (AdminCount=1 but not in protected group)"
                else:
                    # Group that is not a protected group but has adminCount=1
                    status = "Non-Standard Group with AdminCount=1"
                
                # Check if highly privileged but NOT in Protected Users group (security risk)
                risk_level = "Low"
                recommendations = []
                
                if is_protected_group_object:
                    # Built-in protected groups are normal, no risk
                    risk_level = "Low"
                    recommendations.append("Built-in AD object (normal)")
                elif is_service_account_exclusion:
                    # Service accounts like krbtgt that can't be in Protected Users
                    risk_level = "Low"
                    recommendations.append("Service account - Protected Users group not applicable")
                elif is_builtin_administrator:
                    # Built-in Administrator account (RID 500) - special case for emergency access
                    if not in_protected_users_group:
                        risk_level = "Medium"
                        recommendations.append("Built-in Administrator account - Consider Protected Users group only if emergency access doesn't require NTLM/RC4")
                    else:
                        risk_level = "Low"
                        recommendations.append("Built-in Administrator account properly protected")
                elif is_builtin_account:
                    # Other built-in accounts that aren't Administrator or service accounts
                    if not in_protected_users_group:
                        risk_level = "Medium"
                        recommendations.append("Built-in account - Consider Protected Users group")
                    else:
                        risk_level = "Low"
                        recommendations.append("Built-in account properly protected")
                elif is_user and (in_highly_privileged_group or in_protected_group):
                    # User is in privileged groups
                    if not in_protected_users_group:
                        risk_level = "HIGH"
                        recommendations.append("Highly privileged user should be in 'Protected Users' group for additional security")
                elif is_user and not in_protected_group:
                    # User has adminCount=1 but not in protected groups (orphaned)
                    risk_level = "Medium"
                    recommendations.append("Orphaned AdminSDHolder - may indicate past privilege or misconfiguration")
                elif not is_user and not is_protected_group_object:
                    # Non-standard group with adminCount=1
                    risk_level = "Medium"
                    recommendations.append("Non-standard group with AdminCount=1 - review membership and purpose")
                
                # Get last logon (only for users, not groups)
                last_logon_str = "N/A"
                if is_user:
                    last_logon = get_attr(entry, 'lastLogonTimestamp')
                    last_logon_dt = None
                    
                    if isinstance(last_logon, datetime):
                        if last_logon.year != 1601:
                            last_logon_dt = last_logon
                    elif last_logon:
                        # Convert Windows FileTime integer to datetime
                        last_logon_int = safe_int(last_logon)
                        if last_logon_int:
                            last_logon_dt = windows_timestamp_to_datetime(last_logon_int)
                            if last_logon_dt and last_logon_dt.year == 1601:
                                last_logon_dt = None
                    
                    last_logon_str = format_datetime(last_logon_dt) if last_logon_dt else "Never"
                
                results.append({
                    "Name": name,
                    "SAM Account Name": sam_account,
                    "Type": obj_type,
                    "Status": status,
                    "Last Logon": last_logon_str,
                    "In Protected Users Group": "Yes" if in_protected_users_group else "No",
                    "Risk Level": risk_level,
                    "Recommendations": "; ".join(recommendations) if recommendations else "None",
                    "Protected Group Membership": "; ".join(protected_group_membership) if protected_group_membership else "None",
                    "SID": sid_to_str(get_attr(entry, 'objectSid')),
                    "Last Modified": format_datetime(self._parse_timestamp(get_attr(entry, 'whenChanged'))),
                    "Distinguished Name": get_attr(entry, 'distinguishedName', ''),
                })

        except Exception as e:
            logger.warning(f"Error collecting protected groups: {e}")
            logger.debug(f"Exception details: ", exc_info=True)

        self.results['ProtectedGroups'] = results
        logger.info(f"    Found {len(results)} accounts with AdminCount=1")
        return results

    def collect_krbtgt(self) -> List[Dict]:
        """Collect krbtgt account information."""
        logger.info("[-] Collecting krbtgt Account Information...")
        results = []
        now = datetime.now(timezone.utc).replace(tzinfo=None)

        try:
            # Find krbtgt account
            entries = self.search(
                search_base=self.base_dn,
                search_filter="(&(objectCategory=person)(objectClass=user)(sAMAccountName=krbtgt))",
                attributes=['sAMAccountName', 'name', 'distinguishedName', 'pwdLastSet',
                           'msDS-SupportedEncryptionTypes', 'userAccountControl', 'whenCreated',
                           'whenChanged', 'objectSid', 'description']
            )

            if entries:
                entry = entries[0]
                
                # Password age
                pwd_last_set = self._parse_timestamp(get_attr(entry, 'pwdLastSet'))
                pwd_age_days = None
                pwd_last_set_dt = None
                
                if isinstance(pwd_last_set, datetime):
                    if pwd_last_set.year != 1601:
                        pwd_last_set_dt = pwd_last_set
                        pwd_last_set_utc = pwd_last_set.replace(tzinfo=None) if pwd_last_set.tzinfo else pwd_last_set
                        pwd_age_days = (now - pwd_last_set_utc).days
                
                # Kerberos encryption types
                enc_types = get_attr(entry, 'msDS-SupportedEncryptionTypes')
                kerb_enc = parse_kerb_enc_types(enc_types)
                
                # Determine risk level based on password age
                risk_level = "Low"
                risk_factors = []
                
                if pwd_age_days:
                    if pwd_age_days > 365:
                        risk_level = "CRITICAL"
                        risk_factors.append(f"krbtgt password not changed in {pwd_age_days} days (Golden Ticket risk)")
                    elif pwd_age_days > 180:
                        risk_level = "HIGH"
                        risk_factors.append(f"krbtgt password should be changed (last changed {pwd_age_days} days ago)")
                    else:
                        risk_factors.append(f"Password changed {pwd_age_days} days ago (acceptable)")
                
                # Check encryption types
                rc4_val = kerb_enc.get('RC4', 'Not Set')
                aes256_val = kerb_enc.get('AES256', 'Not Set')
                if rc4_val in ['Supported', 'Default'] and aes256_val not in ['Supported', 'Default']:
                    risk_factors.append("Using weak RC4 encryption (AES256 recommended)")
                    if risk_level == "Low":
                        risk_level = "Medium"
                
                results.append({
                    "Account Name": get_attr(entry, 'sAMAccountName', ''),
                    "Distinguished Name": get_attr(entry, 'distinguishedName', ''),
                    "Password Last Set": format_datetime(pwd_last_set_dt) if pwd_last_set_dt else "Never",
                    "Password Age (days)": pwd_age_days if pwd_age_days is not None else "Unknown",
                    "Risk Level": risk_level,
                    "Risk Factors": "; ".join(risk_factors) if risk_factors else "None",
                    "Kerberos RC4": kerb_enc.get('RC4', 'Not Set'),
                    "Kerberos AES-128": kerb_enc.get('AES128', 'Not Set'),
                    "Kerberos AES-256": kerb_enc.get('AES256', 'Not Set'),
                    "SID": sid_to_str(get_attr(entry, 'objectSid')),
                    "Created": format_datetime(self._parse_timestamp(get_attr(entry, 'whenCreated'))),
                    "Last Modified": format_datetime(self._parse_timestamp(get_attr(entry, 'whenChanged'))),
                    "Description": get_attr(entry, 'description', '')
                })

        except Exception as e:
            logger.warning(f"Error collecting krbtgt: {e}")

        self.results['krbtgt'] = results
        logger.info(f"    Found {len(results)} krbtgt account(s)")
        return results

    def collect_gplinks(self) -> List[Dict]:
        """Collect GPO links (Scope of Management)."""
        logger.info("[-] Collecting gPLinks...")
        results = []

        try:
            # Get GPO dictionary for name lookup
            gpo_dict = {}
            gpo_entries = self.search(
                search_base=self.base_dn,
                search_filter='(objectCategory=groupPolicyContainer)',
                attributes=['name', 'displayName']
            )
            for gpo in gpo_entries:
                gpo_name = safe_str(get_attr(gpo, 'name', '')).lower()
                gpo_display = safe_str(get_attr(gpo, 'displayName', ''))
                if gpo_name:
                    gpo_dict[gpo_name] = gpo_display

            # Get domain and OUs with gPLink
            entries = self.search(
                search_base=self.base_dn,
                search_filter='(|(objectCategory=domainDNS)(objectCategory=organizationalUnit))',
                attributes=['name', 'distinguishedName', 'gPLink', 'gPOptions']
            )

            # Also get sites
            site_entries = self.search(
                search_base=f"CN=Sites,{self.config_dn}",
                search_filter='(objectCategory=site)',
                attributes=['name', 'distinguishedName', 'gPLink', 'gPOptions']
            )
            entries.extend(site_entries)

            for entry in entries:
                
                gp_link = get_attr(entry, 'gPLink', '')
                gp_options = get_attr(entry, 'gPOptions', 0)
                block_inheritance = bool(safe_int(gp_options) & 1)
                
                dn = get_attr(entry, 'distinguishedName', '')
                depth = dn.count('OU=') if dn else 0
                name = get_attr(entry, 'name', '')

                # If no gPLink, add one entry showing the OU/Site exists but has no links
                if not gp_link:
                    results.append({
                        "Name": name,
                        "Depth": depth,
                        "DistinguishedName": dn,
                        "Link Order": "",
                        "GPO": "",
                        "Enforced": "",
                        "Link Enabled": "",
                        "BlockInheritance": block_inheritance,
                        "gPLink": "",
                        "gPOptions": gp_options,
                    })
                    continue

                # Parse gPLink - format: [LDAP://cn={GUID},cn=policies,...;options][...]
                links = re.findall(r'\[LDAP://([^;]+);(\d+)\]', str(gp_link), re.IGNORECASE)

                # Link order is reversed - first link in list has highest order number
                total_links = len(links)
                for idx, (link_dn, link_options) in enumerate(links):
                    link_options = int(link_options)
                    enforced = bool(link_options & 2)
                    disabled = bool(link_options & 1)
                    
                    # Calculate link order (reverse: first = highest number)
                    link_order = total_links - idx

                    # Extract GPO GUID
                    guid_match = re.search(r'cn=(\{[^}]+\})', link_dn, re.IGNORECASE)
                    gpo_guid = guid_match.group(1) if guid_match else ""
                    gpo_display_name = gpo_dict.get(gpo_guid.lower(), gpo_guid)

                    results.append({
                        "Name": name,
                        "Depth": depth,
                        "DistinguishedName": dn,
                        "Link Order": link_order,
                        "GPO": gpo_display_name,
                        "Enforced": enforced,
                        "Link Enabled": not disabled,
                        "BlockInheritance": block_inheritance,
                        "gPLink": gp_link,
                        "gPOptions": gp_options,
                    })

        except Exception as e:
            logger.warning(f"Error collecting gPLinks: {e}")

        self.results['gPLinks'] = results
        logger.info(f"    Found {len(results)} GPO links")
        return results

    def collect_about(self) -> List[Dict]:
        """Collect metadata about the PyADRecon run."""
        logger.info("[-] Collecting About PyADRecon...")
        results: List[Dict] = []

        try:
            # Calculate execution time
            end_time = datetime.now()
            duration = end_time - self.start_time
            duration_secs = duration.total_seconds()

            # Get computer name where script is running
            import os
            import platform
            import socket

            local_computer = (
                socket.gethostname().strip()
                or os.environ.get("COMPUTERNAME", "").strip()
                or platform.node().strip()
            )

            computer_type = "Non-Domain System"

            # Try to determine computer type by searching AD
            try:
                candidates = []
                if local_computer:
                    candidates.append(local_computer)

                    # If it's not an FQDN, try appending the target domain
                    if "." not in local_computer and self.base_dn:
                        target_fqdn = dn_to_fqdn(self.base_dn)
                        if target_fqdn:
                            candidates.append(f"{local_computer}.{target_fqdn}")

                comp_entries = None
                for cand in candidates:
                    # Prefer dNSHostName matches
                    comp_entries = self.search(
                        self.base_dn,
                        f"(&(objectCategory=computer)(|(dNSHostName={cand})(dNSHostName={cand}*)))",
                        ['dNSHostName', 'operatingSystem']
                    )
                    if comp_entries:
                        break

                    # Fallback to CN/name for NetBIOS-style hostnames
                    comp_entries = self.search(
                        self.base_dn,
                        f"(&(objectCategory=computer)(|(cn={cand})(name={cand})))",
                        ['dNSHostName', 'operatingSystem']
                    )
                    if comp_entries:
                        break

                if comp_entries:
                    os_name = get_attr(comp_entries[0], 'operatingSystem', '') or ''
                    if 'Server' in os_name:
                        computer_type = "Domain Member Server"
                    else:
                        computer_type = "Domain Member Workstation"

                    # Prefer displaying the AD dNSHostName if present
                    ad_hostname = get_attr(comp_entries[0], 'dNSHostName', '') or ''
                    if ad_hostname:
                        local_computer = ad_hostname

            except Exception:
                pass

            if not local_computer:
                local_computer = "Unknown Host"

            results.append({"Category": "PyADRecon-ADWS Version", "Value": VERSION})
            results.append({"Category": "Date", "Value": self.start_time.strftime("%m.%d.%Y %H:%M")})
            results.append({"Category": "GitHub Repository", "Value": "github.com/l4rm4nd/PyADRecon-ADWS"})
            results.append({"Category": "Executed By", "Value": self.config.username if self.config.username else "Current User"})
            results.append({"Category": "Executed From", "Value": f"{local_computer} ({computer_type})"})
            results.append({"Category": "Execution Time", "Value": f"{duration_secs:.2f} seconds"})
            results.append({"Category": "Target Domain", "Value": dn_to_fqdn(self.base_dn)})
            results.append({"Category": "Protocol", "Value": "ADWS (Port 9389)"})

        except Exception as e:
            logger.warning(f"Error collecting about info: {e}")

        logger.info(f"    Found {len(results)} metadata items")
        return results

    def run(self) -> bool:
        """Run the reconnaissance collection."""
        try:
            # Connect
            if not self.connect():
                return False

            # Initialize SID mappings for cross-domain resolution
            # This ensures SID-to-domain-name mapping works even with --collect flags
            self._initialize_sid_mappings()

            # Collect about info first
            self.results['AboutPyADRecon'] = self.collect_about()

            # Collect data based on config
            collection_methods = [
                ('Domain', self.config.collect_domain, self.collect_domain_info),
                ('Forest', self.config.collect_forest, self.collect_forest_info),
                ('Trusts', self.config.collect_trusts, self.collect_trusts),
                ('Sites', self.config.collect_sites, self.collect_sites),
                ('Subnets', self.config.collect_subnets, self.collect_subnets),
                ('SchemaHistory', self.config.collect_schema, self.collect_schema_history),
                ('DNSZones', self.config.collect_dns_zones, self.collect_dns_zones),
                ('DNSRecords', self.config.collect_dns_records, self.collect_dns_records),
                ('DomainControllers', self.config.collect_dcs, self.collect_domain_controllers),
                ('PasswordPolicy', self.config.collect_password_policy, self.collect_password_policy),
                ('FineGrainedPasswordPolicy', self.config.collect_fgpp, self.collect_fine_grained_password_policies),
                ('Users', self.config.collect_users, self.collect_users),
                ('UserSPNs', self.config.collect_user_spns, self.collect_user_spns),
                ('Groups', self.config.collect_groups, self.collect_groups),
                ('GroupMembers', self.config.collect_group_members, self.collect_group_members),
                ('Computers', self.config.collect_computers, self.collect_computers),
                ('ComputerSPNs', self.config.collect_computer_spns, self.collect_computer_spns),
                ('OUs', self.config.collect_ous, self.collect_ous),
                ('GPOs', self.config.collect_gpos, self.collect_gpos),
                ('gPLinks', self.config.collect_gplinks, self.collect_gplinks),
                ('LAPS', self.config.collect_laps, self.collect_laps),
                ('Bitlocker', self.config.collect_bitlocker, self.collect_bitlocker),
                ('Printers', self.config.collect_printers, self.collect_printers),
                ('gMSA', self.config.collect_gmsa, self.collect_gmsa),
                ('dMSA', self.config.collect_dmsa, self.collect_dmsa),
                ('ProtectedGroups', self.config.collect_protected_groups, self.collect_protected_groups),
                ('krbtgt', self.config.collect_krbtgt, self.collect_krbtgt),
                ('Kerberoastable', self.config.collect_kerberoastable, self.collect_kerberoastable),
                ('ASREPRoastable', self.config.collect_asrep_roastable, self.collect_asrep_roastable),
                ('CertificateTemplates', self.config.collect_adcs, self.collect_certificate_templates),
                ('CertificateAuthorities', self.config.collect_adcs, self.collect_certificate_authorities),
            ]

            for name, enabled, method in collection_methods:
                if enabled:
                    try:
                        self.results[name] = method()
                    except Exception as e:
                        logger.error(f"Error collecting {name}: {e}")
                        self.results[name] = []

            return True

        except Exception as e:
            logger.error(f"Run failed: {e}")
            traceback.print_exc()
            return False

    def _sanitize_value(self, value):
        """Sanitize value for CSV output."""
        if value is None:
            return ""
        if isinstance(value, bytes):
            return value.hex()
        if isinstance(value, datetime):
            return format_datetime(value)
        if isinstance(value, list):
            return '; '.join([str(v) for v in value])
        return str(value)

    def _sanitize_results(self):
        """Sanitize all results for export."""
        for module_name, entries in self.results.items():
            for entry in entries:
                for key, value in entry.items():
                    entry[key] = self._sanitize_value(value)

    def export_csv(self, output_dir: str) -> str:
        """Export results to CSV files."""
        csv_dir = os.path.join(output_dir, "CSV-Files")
        os.makedirs(csv_dir, exist_ok=True)

        self._sanitize_results()

        for module_name, entries in self.results.items():
            if not entries:
                continue

            csv_file = os.path.join(csv_dir, f"{module_name}.csv")
            
            try:
                with open(csv_file, 'w', newline='', encoding='utf-8') as f:
                    if entries:
                        fieldnames = list(entries[0].keys())
                        writer = csv.DictWriter(f, fieldnames=fieldnames)
                        writer.writeheader()
                        writer.writerows(entries)
                
                logger.info(f"[+] Exported {module_name} to {csv_file}")
            
            except Exception as e:
                logger.error(f"Error exporting {module_name}: {e}")

        return csv_dir

    def apply_security_formatting(self, wb):
        """Apply conditional formatting to highlight security issues in Users and Computers sheets.
        
        Color scheme:
        - Red: Critical security issues (e.g., password never expires, unconstrained delegation, passwords in description/info)
        - Orange: Medium security issues (e.g., SPNs, password age)
        - Yellow: Informational/warnings (e.g., dormant accounts, never logged in, disabled status)
        - Gray: Disabled accounts (row background, unless overridden by security colors)
        """
        from openpyxl.styles import PatternFill
        import re
        
        # Define color schemes for security issues
        red_fill = PatternFill(start_color="FFB3BA", end_color="FFB3BA", fill_type="solid")  # Light red
        orange_fill = PatternFill(start_color="FFD9B3", end_color="FFD9B3", fill_type="solid")  # Light orange
        yellow_fill = PatternFill(start_color="FFFFB3", end_color="FFFFB3", fill_type="solid")  # Light yellow
        
        # Regex pattern to detect passwords in description/info fields
        # Matches common password-related keywords in multiple languages
        password_pattern = re.compile(
            r'\b(pw|password|passwort|kennwort|initial|pwd|pass|secret|cred|credential)\b',
            re.IGNORECASE
        )
        
        logger.info("    Applying security-based conditional formatting...")
        
        # Format Users sheet
        if "Users" in wb.sheetnames:
            ws = wb["Users"]
            if ws.max_row > 1:  # Has data beyond header
                # Find column indices for security-relevant fields
                headers = {cell.value: cell.column for cell in ws[1]}
                
                # Critical security issues (RED)
                red_columns = {
                    "Password Never Expires": (True, "TRUE", "True"),
                    "Reversible Password Encryption": (True, "TRUE", "True"),
                    "Does Not Require Pre Auth": (True, "TRUE", "True"),
                    "Password Not Required": (True, "TRUE", "True"),
                    "Kerberos DES Only": (True, "TRUE", "True"),
                    "AdminCount": (1, "1"),
                }
                
                # Medium security issues (ORANGE)
                orange_columns = {
                    "Delegation Type": ("Unconstrained", "Unconstrained"),
                    "Must Change Password at Logon": (True, "TRUE", "True"),
                    "Cannot Change Password": (True, "TRUE", "True"),
                    "Account Locked Out": (True, "TRUE", "True"),
                    "HasSPN": (True, "TRUE", "True"),
                }
                
                # Informational/warning (YELLOW)
                yellow_columns = [
                    "Never Logged in",
                    f"Dormant (> {self.config.dormant_days} days)",
                    f"Password Age (> {self.config.password_age_days} days)",
                ]
                
                # Apply formatting to data rows (skip header)
                for row_idx in range(2, ws.max_row + 1):
                    # Check Description and Info fields for password-related keywords (CRITICAL - RED)
                    for field_name in ["Description", "Info"]:
                        if field_name in headers:
                            cell = ws.cell(row=row_idx, column=headers[field_name])
                            if cell.value and isinstance(cell.value, str):
                                if password_pattern.search(cell.value):
                                    cell.fill = red_fill
                    
                    # First, check Enabled status and color the Enabled cell itself
                    is_disabled = False
                    if "Enabled" in headers:
                        enabled_cell = ws.cell(row=row_idx, column=headers["Enabled"])
                        if enabled_cell.value in (False, "FALSE", "False"):
                            is_disabled = True
                            # Color Enabled cell yellow for disabled accounts (informational/warning)
                            # Yellow indicates account status without implying "bad" (red) or "good" (green)
                            enabled_cell.fill = yellow_fill
                        # Enabled accounts remain uncolored (neutral state)
                    
                    # RED: Critical security issues
                    for col_name, trigger_values in red_columns.items():
                        if col_name in headers:
                            cell = ws.cell(row=row_idx, column=headers[col_name])
                            if cell.value in trigger_values:
                                cell.fill = red_fill
                    
                    # ORANGE: Medium security issues
                    for col_name, trigger_values in orange_columns.items():
                        if col_name in headers:
                            cell = ws.cell(row=row_idx, column=headers[col_name])
                            if cell.value in trigger_values:
                                cell.fill = orange_fill
                    
                    # YELLOW: Informational/warnings
                    for col_name in yellow_columns:
                        if col_name in headers:
                            cell = ws.cell(row=row_idx, column=headers[col_name])
                            if cell.value in (True, "TRUE", "True"):
                                cell.fill = yellow_fill
                    
                    # Special: Gray out entire row for disabled users (after other formatting)
                    if is_disabled:
                        gray_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                        for col in range(1, ws.max_column + 1):
                            # Only gray cells that haven't been colored by security issues
                            cell = ws.cell(row=row_idx, column=col)
                            # Skip the Enabled column itself (keep it yellow)
                            if col != headers.get("Enabled"):
                                # Keep security highlighting visible
                                if cell.fill.start_color.rgb not in ["00FFB3BA", "00FFD9B3", "00FFFFB3"]:
                                    cell.fill = gray_fill
        
        # Format Computers sheet
        if "Computers" in wb.sheetnames:
            ws = wb["Computers"]
            if ws.max_row > 1:  # Has data beyond header
                # Find column indices for security-relevant fields
                headers = {cell.value: cell.column for cell in ws[1]}
                
                # Critical security issues (RED)
                red_columns = {
                    "Delegation Type": ("Unconstrained", "Unconstrained"),
                }
                
                # Medium security issues (ORANGE)
                orange_columns = {
                    f"Password Age (> {self.config.password_age_days} days)": (True, "TRUE", "True"),
                }
                
                # Informational/warning (YELLOW)
                yellow_columns = [
                    f"Dormant (> {self.config.dormant_days} days)",
                ]
                
                # Apply formatting to data rows (skip header)
                for row_idx in range(2, ws.max_row + 1):
                    # First, check Enabled status and color the Enabled cell itself
                    is_disabled = False
                    if "Enabled" in headers:
                        enabled_cell = ws.cell(row=row_idx, column=headers["Enabled"])
                        if enabled_cell.value in (False, "FALSE", "False"):
                            is_disabled = True
                            # Color Enabled cell yellow for disabled computers (informational/warning)
                            # Yellow indicates computer status without implying "bad" (red) or "good" (green)
                            enabled_cell.fill = yellow_fill
                        # Enabled computers remain uncolored (neutral state)
                    
                    # RED: Critical security issues
                    for col_name, trigger_values in red_columns.items():
                        if col_name in headers:
                            cell = ws.cell(row=row_idx, column=headers[col_name])
                            if cell.value in trigger_values:
                                cell.fill = red_fill
                    
                    # ORANGE: Medium security issues
                    for col_name, trigger_values in orange_columns.items():
                        if col_name in headers:
                            cell = ws.cell(row=row_idx, column=headers[col_name])
                            if cell.value in trigger_values:
                                cell.fill = orange_fill
                    
                    # YELLOW: Informational/warnings
                    for col_name in yellow_columns:
                        if col_name in headers:
                            cell = ws.cell(row=row_idx, column=headers[col_name])
                            if cell.value in (True, "TRUE", "True"):
                                cell.fill = yellow_fill
                    
                    # Special: Gray out entire row for disabled computers (after other formatting)
                    if is_disabled:
                        gray_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                        for col in range(1, ws.max_column + 1):
                            # Only gray cells that haven't been colored by security issues
                            cell = ws.cell(row=row_idx, column=col)
                            # Skip the Enabled column itself (keep it yellow)
                            if col != headers.get("Enabled"):
                                # Keep security highlighting visible
                                if cell.fill.start_color.rgb not in ["00FFB3BA", "00FFD9B3", "00FFFFB3"]:
                                    cell.fill = gray_fill
        
        # Format PasswordPolicy sheet
        if "PasswordPolicy" in wb.sheetnames:
            ws = wb["PasswordPolicy"]
            if ws.max_row > 1:  # Has data beyond header
                # Find column indices
                headers = {cell.value: cell.column for cell in ws[1]}
                import re
                def parse_interval_to_days(interval_str):
                    # Parse 'Xd Yh Zm Ws' to total days (float)
                    if not interval_str or not isinstance(interval_str, str):
                        return None
                    pattern = r"(?:(\d+(?:\.\d+)?)d)?\s*(?:(\d+(?:\.\d+)?)h)?\s*(?:(\d+(?:\.\d+)?)m)?\s*(?:(\d+(?:\.\d+)?)s)?"
                    match = re.match(pattern, interval_str.strip())
                    if not match:
                        return None
                    days, hours, mins, secs = match.groups(default="0")
                    try:
                        total_days = float(days) + float(hours)/24 + float(mins)/1440 + float(secs)/86400
                        return total_days
                    except Exception:
                        return None
                def parse_interval_to_minutes(interval_str):
                    # Parse 'Xd Yh Zm Ws' to total minutes (float)
                    if not interval_str or not isinstance(interval_str, str):
                        return None
                    pattern = r"(?:(\d+(?:\.\d+)?)d)?\s*(?:(\d+(?:\.\d+)?)h)?\s*(?:(\d+(?:\.\d+)?)m)?\s*(?:(\d+(?:\.\d+)?)s)?"
                    match = re.match(pattern, interval_str.strip())
                    if not match:
                        return None
                    days, hours, mins, secs = match.groups(default="0")
                    try:
                        total_minutes = float(days)*1440 + float(hours)*60 + float(mins) + float(secs)/60
                        return total_minutes
                    except Exception:
                        return None
                if "Current Value" in headers and "CIS Benchmark 2024-25" in headers:
                    current_col = headers["Current Value"]
                    cis_col = headers["CIS Benchmark 2024-25"]
                    policy_col = headers["Policy"]
                    # Apply formatting to data rows (skip header)
                    for row_idx in range(2, ws.max_row + 1):
                        policy = ws.cell(row=row_idx, column=policy_col).value
                        current_val = ws.cell(row=row_idx, column=current_col).value
                        cis_val = ws.cell(row=row_idx, column=cis_col).value
                        try:
                            current_val_str = str(current_val).strip()
                            cis_val_str = str(cis_val).strip()
                            non_compliant = False
                            if "Enforce password history" in str(policy):
                                # CIS: 24 or more
                                if current_val_str.isdigit() and int(current_val_str) < 24:
                                    non_compliant = True
                            elif "Maximum password age" in str(policy):
                                # CIS: 1 to 365 days
                                days = parse_interval_to_days(current_val_str)
                                if days is None or days < 1 or days > 365:
                                    non_compliant = True
                            elif "Minimum password age" in str(policy):
                                # CIS: 1 or more days
                                days = parse_interval_to_days(current_val_str)
                                if days is None or days < 1:
                                    non_compliant = True
                            elif "Minimum password length" in str(policy):
                                if current_val_str.isdigit() and int(current_val_str) < 14:
                                    non_compliant = True
                            elif "complexity requirements" in str(policy):
                                if current_val_str.upper() != "TRUE":
                                    non_compliant = True
                            elif "reversible encryption" in str(policy):
                                if current_val_str.upper() != "FALSE":
                                    non_compliant = True
                            elif "Account lockout duration" in str(policy):
                                # CIS: 15 or more days (0 is also acceptable for manual unlock)
                                days = parse_interval_to_days(current_val_str)
                                if days is None or (days != 0 and days < 15):
                                    non_compliant = True
                            elif "Account lockout threshold" in str(policy):
                                if current_val_str.isdigit():
                                    val = int(current_val_str)
                                    if val < 1 or val > 5:
                                        non_compliant = True
                                elif current_val_str == "Not Set" or current_val_str == "0":
                                    non_compliant = True
                            elif "Reset account lockout counter" in str(policy):
                                # CIS: 15 or more minutes
                                mins = parse_interval_to_minutes(current_val_str)
                                if mins is None or mins < 15:
                                    non_compliant = True
                            # Highlight the Current Value cell - orange for non-compliant, green for compliant
                            if non_compliant:
                                ws.cell(row=row_idx, column=current_col).fill = orange_fill
                            else:
                                green_fill = PatternFill(start_color="B3FFB3", end_color="B3FFB3", fill_type="solid")
                                ws.cell(row=row_idx, column=current_col).fill = green_fill
                        except Exception as e:
                            logger.debug(f"Error comparing password policy values: {e}")
        
        # Format LAPS sheet
        if "LAPS" in wb.sheetnames:
            ws = wb["LAPS"]
            if ws.max_row > 1:  # Has data beyond header
                # Find column indices
                headers = {cell.value: cell.column for cell in ws[1]}
                
                if "Password" in headers:
                    password_col = headers["Password"]
                    
                    # Apply formatting to data rows (skip header)
                    for row_idx in range(2, ws.max_row + 1):
                        password_cell = ws.cell(row=row_idx, column=password_col)
                        # Highlight password field YELLOW if it contains a value
                        if password_cell.value and str(password_cell.value).strip():
                            password_cell.fill = yellow_fill
        
        # Format gMSA sheet
        if "gMSA" in wb.sheetnames:
            ws = wb["gMSA"]
            if ws.max_row > 1:  # Has data beyond header
                # Find column indices for security-relevant fields
                headers = {cell.value: cell.column for cell in ws[1]}
                
                # Apply formatting to data rows (skip header)
                for row_idx in range(2, ws.max_row + 1):
                    # Check for badSuccessor vulnerability (CRITICAL - RED)
                    if "badSuccessor Vulnerable" in headers:
                        vuln_cell = ws.cell(row=row_idx, column=headers["badSuccessor Vulnerable"])
                        if vuln_cell.value in (True, "TRUE", "True"):
                            vuln_cell.fill = red_fill
                    
                    # Highlight Risk Level column
                    if "Risk Level" in headers:
                        risk_cell = ws.cell(row=row_idx, column=headers["Risk Level"])
                        risk_value = str(risk_cell.value).upper() if risk_cell.value else ""
                        
                        if risk_value == "CRITICAL":
                            risk_cell.fill = red_fill
                        elif risk_value == "HIGH":
                            risk_cell.fill = red_fill
                        elif risk_value == "MEDIUM":
                            risk_cell.fill = orange_fill
                    
                    # Highlight disabled accounts (YELLOW - informational)
                    if "Enabled" in headers:
                        enabled_cell = ws.cell(row=row_idx, column=headers["Enabled"])
                        if enabled_cell.value in (False, "FALSE", "False"):
                            enabled_cell.fill = yellow_fill
        
        # Format dMSA sheet

        # Format dMSA sheet
        if "dMSA" in wb.sheetnames:
            ws = wb["dMSA"]
            if ws.max_row > 1:  # Has data beyond header
                # Find column indices for security-relevant fields
                headers = {cell.value: cell.column for cell in ws[1]}
                
                # Apply formatting to data rows (skip header)
                for row_idx in range(2, ws.max_row + 1):
                    # Check for badSuccessor vulnerability (CRITICAL - RED)
                    if "badSuccessor Vulnerable" in headers:
                        vuln_cell = ws.cell(row=row_idx, column=headers["badSuccessor Vulnerable"])
                        if vuln_cell.value in (True, "TRUE", "True"):
                            vuln_cell.fill = red_fill
                    
                    # Highlight Risk Level column
                    if "Risk Level" in headers:
                        risk_cell = ws.cell(row=row_idx, column=headers["Risk Level"])
                        risk_value = str(risk_cell.value).upper() if risk_cell.value else ""
                        
                        if risk_value == "CRITICAL":
                            risk_cell.fill = red_fill
                        elif risk_value == "HIGH":
                            risk_cell.fill = red_fill
                        elif risk_value == "MEDIUM":
                            risk_cell.fill = orange_fill
                    
                    # Highlight disabled accounts (YELLOW - informational)
                    if "Enabled" in headers:
                        enabled_cell = ws.cell(row=row_idx, column=headers["Enabled"])
                        if enabled_cell.value in (False, "FALSE", "False"):
                            enabled_cell.fill = yellow_fill
        

        # Format CertificateTemplates sheet
        if "CertificateTemplates" in wb.sheetnames:
            ws = wb["CertificateTemplates"]
            if ws.max_row > 1:  # Has data beyond header
                # Find column indices for security-relevant fields
                headers = {cell.value: cell.column for cell in ws[1]}
                
                # Apply formatting to data rows (skip header)
                for row_idx in range(2, ws.max_row + 1):
                    # Highlight Risk Level column (RED for CRITICAL/HIGH, ORANGE for MEDIUM)
                    if "Risk Level" in headers:
                        risk_cell = ws.cell(row=row_idx, column=headers["Risk Level"])
                        risk_value = str(risk_cell.value).upper() if risk_cell.value else ""
                        
                        if risk_value in ["CRITICAL", "HIGH"]:
                            risk_cell.fill = red_fill
                        elif risk_value == "MEDIUM":
                            risk_cell.fill = orange_fill
                    
                    # Highlight ESC Vulnerabilities column (RED if any ESC vulns present)
                    if "ESC Vulnerabilities" in headers:
                        esc_cell = ws.cell(row=row_idx, column=headers["ESC Vulnerabilities"])
                        esc_value = str(esc_cell.value) if esc_cell.value else ""
                        
                        # Check if any ESC vulnerability is present (not "None")
                        if esc_value and esc_value.strip().upper() != "NONE":
                            # Highlight based on specific ESC types
                            if "ESC1" in esc_value or "ESC4" in esc_value:
                                esc_cell.fill = red_fill
                            elif "ESC2" in esc_value or "ESC3" in esc_value:
                                esc_cell.fill = orange_fill
                            elif "ESC9" in esc_value:
                                esc_cell.fill = yellow_fill
                    
                    # Highlight critical individual flags
                    if "Enrollee Supplies Subject" in headers:
                        enrollee_cell = ws.cell(row=row_idx, column=headers["Enrollee Supplies Subject"])
                        if enrollee_cell.value in (True, "TRUE", "True"):
                            # Only highlight red if this is actually part of an ESC1 vulnerability
                            # Check if this row has ESC1 flagged
                            is_esc1 = False
                            if "ESC Vulnerabilities" in headers:
                                esc_cell = ws.cell(row=row_idx, column=headers["ESC Vulnerabilities"])
                                esc_value = str(esc_cell.value) if esc_cell.value else ""
                                is_esc1 = "ESC1" in esc_value
                            
                            if is_esc1:
                                enrollee_cell.fill = red_fill
                            # Don't highlight if not part of ESC1 (benign use case)
                    
                    # Highlight Client Authentication (ORANGE - critical for exploitation)
                    if "Client Authentication" in headers:
                        client_auth_cell = ws.cell(row=row_idx, column=headers["Client Authentication"])
                        if client_auth_cell.value in (True, "TRUE", "True"):
                            client_auth_cell.fill = orange_fill
                    
                    # Highlight Any Purpose EKU (ORANGE - allows arbitrary usage)
                    if "Any Purpose EKU" in headers:
                        any_purpose_cell = ws.cell(row=row_idx, column=headers["Any Purpose EKU"])
                        if any_purpose_cell.value in (True, "TRUE", "True"):
                            any_purpose_cell.fill = orange_fill
                    
                    # Highlight Enrollment Agent (ORANGE - can request certs for others)
                    if "Enrollment Agent" in headers:
                        enrollment_agent_cell = ws.cell(row=row_idx, column=headers["Enrollment Agent"])
                        if enrollment_agent_cell.value in (True, "TRUE", "True"):
                            enrollment_agent_cell.fill = orange_fill
                    
                    # Highlight Allows SAN (ORANGE - similar to Enrollee Supplies Subject)
                    if "Allows SAN" in headers:
                        allows_san_cell = ws.cell(row=row_idx, column=headers["Allows SAN"])
                        if allows_san_cell.value in (True, "TRUE", "True"):
                            allows_san_cell.fill = orange_fill
                    
                    # Highlight Exportable Key (YELLOW - private key can be exported)
                    if "Exportable Key" in headers:
                        exportable_key_cell = ws.cell(row=row_idx, column=headers["Exportable Key"])
                        if exportable_key_cell.value in (True, "TRUE", "True"):
                            exportable_key_cell.fill = yellow_fill
                    
                    # Highlight Auto-Enrollment (YELLOW - informational, auto-enrollment enabled)
                    if "Auto-Enrollment" in headers:
                        auto_enrollment_cell = ws.cell(row=row_idx, column=headers["Auto-Enrollment"])
                        if auto_enrollment_cell.value in (True, "TRUE", "True"):
                            auto_enrollment_cell.fill = yellow_fill
                    
                    # Highlight Requires Manager Approval (GREEN when TRUE, ORANGE when FALSE)
                    if "Requires Manager Approval" in headers:
                        approval_cell = ws.cell(row=row_idx, column=headers["Requires Manager Approval"])
                        if approval_cell.value in (True, "TRUE", "True"):
                            # Green for approval required (good security)
                            green_fill = PatternFill(start_color="B3FFB3", end_color="B3FFB3", fill_type="solid")
                            approval_cell.fill = green_fill
                        elif approval_cell.value in (False, "FALSE", "False"):
                            # Orange for no approval required (risky)
                            approval_cell.fill = orange_fill
        
        # Format Groups sheet
        if "Groups" in wb.sheetnames:
            ws = wb["Groups"]
            if ws.max_row > 1:
                headers = {cell.value: cell.column for cell in ws[1]}
                
                for row_idx in range(2, ws.max_row + 1):
                    # Highlight AdminCount (ORANGE - privileged group)
                    if "AdminCount" in headers:
                        admin_cell = ws.cell(row=row_idx, column=headers["AdminCount"])
                        if admin_cell.value in [1, "1", True, "TRUE"]:
                            admin_cell.fill = orange_fill
                    
                    # Highlight SIDHistory (RED - potential security risk, indicates domain migration or SID injection)
                    if "SIDHistory" in headers:
                        sid_history_cell = ws.cell(row=row_idx, column=headers["SIDHistory"])
                        if sid_history_cell.value and str(sid_history_cell.value).strip() and str(sid_history_cell.value) != "":
                            sid_history_cell.fill = red_fill
        
        # Format DomainControllers sheet
        if "DomainControllers" in wb.sheetnames:
            ws = wb["DomainControllers"]
            if ws.max_row > 1:
                headers = {cell.value: cell.column for cell in ws[1]}
                
                for row_idx in range(2, ws.max_row + 1):
                    # Highlight LDAP Signing (RED if not required)
                    if "LDAP Signing" in headers:
                        signing_cell = ws.cell(row=row_idx, column=headers["LDAP Signing"])
                        if signing_cell.value in ["Not Required", "NOT REQUIRED", False, "FALSE"]:
                            signing_cell.fill = red_fill
                        elif signing_cell.value in ["Required", "REQUIRED", True, "TRUE"]:
                            green_fill = PatternFill(start_color="B3FFB3", end_color="B3FFB3", fill_type="solid")
                            signing_cell.fill = green_fill
                    
                    # Highlight LDAP Channel Binding (RED if not required)
                    if "LDAP Channel Binding" in headers:
                        binding_cell = ws.cell(row=row_idx, column=headers["LDAP Channel Binding"])
                        if binding_cell.value in ["Not Required", "NOT REQUIRED", False, "FALSE"]:
                            binding_cell.fill = red_fill
                        elif binding_cell.value in ["Required", "REQUIRED", True, "TRUE"]:
                            green_fill = PatternFill(start_color="B3FFB3", end_color="B3FFB3", fill_type="solid")
                            binding_cell.fill = green_fill
                    
                    # Highlight SMB1 (RED if enabled - deprecated and insecure)
                    if "SMB1(NT LM 0.12)" in headers:
                        smb1_cell = ws.cell(row=row_idx, column=headers["SMB1(NT LM 0.12)"])
                        if smb1_cell.value in (True, "TRUE", "True"):
                            smb1_cell.fill = red_fill
                    
                    # Highlight SMB Signing (RED if not enabled)
                    if "SMB Signing" in headers:
                        signing_cell = ws.cell(row=row_idx, column=headers["SMB Signing"])
                        if signing_cell.value in (False, "FALSE", "False"):
                            signing_cell.fill = red_fill
                        elif signing_cell.value in (True, "TRUE", "True"):
                            green_fill = PatternFill(start_color="B3FFB3", end_color="B3FFB3", fill_type="solid")
                            signing_cell.fill = green_fill
        
        # Format FineGrainedPasswordPolicy sheet
        if "FineGrainedPasswordPolicy" in wb.sheetnames:
            ws = wb["FineGrainedPasswordPolicy"]
            if ws.max_row > 1:
                headers = {cell.value: cell.column for cell in ws[1]}
                
                for row_idx in range(2, ws.max_row + 1):
                    # Highlight weak password policies (similar to PasswordPolicy sheet)
                    if "Minimum Password Length" in headers:
                        length_cell = ws.cell(row=row_idx, column=headers["Minimum Password Length"])
                        try:
                            length_val = int(length_cell.value) if length_cell.value else 0
                            if length_val < 14:
                                length_cell.fill = orange_fill
                        except (ValueError, TypeError):
                            pass
                    
                    # Highlight Reversible Encryption (RED if enabled)
                    if "Reversible Encryption Enabled" in headers:
                        rev_enc_cell = ws.cell(row=row_idx, column=headers["Reversible Encryption Enabled"])
                        if rev_enc_cell.value in (True, "TRUE", "True"):
                            rev_enc_cell.fill = red_fill
                    
                    # Highlight low lockout threshold
                    if "Lockout Threshold" in headers:
                        lockout_cell = ws.cell(row=row_idx, column=headers["Lockout Threshold"])
                        try:
                            lockout_val = int(lockout_cell.value) if lockout_cell.value else 0
                            if lockout_val == 0 or lockout_val > 5:
                                lockout_cell.fill = orange_fill
                        except (ValueError, TypeError):
                            pass
        
        # Format UserSPNs sheet
        if "UserSPNs" in wb.sheetnames:
            ws = wb["UserSPNs"]
            if ws.max_row > 1:
                headers = {cell.value: cell.column for cell in ws[1]}
                
                for row_idx in range(2, ws.max_row + 1):
                    # Highlight disabled accounts (YELLOW - informational)
                    if "Enabled" in headers:
                        enabled_cell = ws.cell(row=row_idx, column=headers["Enabled"])
                        if enabled_cell.value in (False, "FALSE", "False"):
                            enabled_cell.fill = yellow_fill
        
        # Format ComputerSPNs sheet
        if "ComputerSPNs" in wb.sheetnames:
            ws = wb["ComputerSPNs"]
            if ws.max_row > 1:
                headers = {cell.value: cell.column for cell in ws[1]}
                
                for row_idx in range(2, ws.max_row + 1):
                    # Highlight disabled accounts (YELLOW - informational)
                    if "Enabled" in headers:
                        enabled_cell = ws.cell(row=row_idx, column=headers["Enabled"])
                        if enabled_cell.value in (False, "FALSE", "False"):
                            enabled_cell.fill = yellow_fill
        
        # Format Domain sheet
        if "Domain" in wb.sheetnames:
            ws = wb["Domain"]
            if ws.max_row > 1:
                headers = {cell.value: cell.column for cell in ws[1]}
                
                for row_idx in range(2, ws.max_row + 1):
                    # Highlight Machine Account Quota (ORANGE if > 0, GREEN if 0)
                    if "Machine Account Quota" in headers:
                        quota_cell = ws.cell(row=row_idx, column=headers["Machine Account Quota"])
                        try:
                            quota_val = int(quota_cell.value) if quota_cell.value else -1
                            if quota_val > 0:
                                quota_cell.fill = orange_fill
                            elif quota_val == 0:
                                green_fill = PatternFill(start_color="B3FFB3", end_color="B3FFB3", fill_type="solid")
                                quota_cell.fill = green_fill
                        except (ValueError, TypeError):
                            pass
        
        # Format ProtectedGroups sheet
        if "ProtectedGroups" in wb.sheetnames:
            ws = wb["ProtectedGroups"]
            if ws.max_row > 1:  # Has data beyond header
                headers = {cell.value: cell.column for cell in ws[1]}
                
                # Apply formatting to data rows (skip header)
                for row_idx in range(2, ws.max_row + 1):
                    # Highlight Risk Level column
                    if "Risk Level" in headers:
                        risk_cell = ws.cell(row=row_idx, column=headers["Risk Level"])
                        risk_value = str(risk_cell.value).upper() if risk_cell.value else ""
                        
                        if risk_value in ["CRITICAL", "HIGH"]:
                            risk_cell.fill = red_fill
                        elif risk_value == "MEDIUM":
                            risk_cell.fill = orange_fill
                        # Low risk remains uncolored (not a problem, just normal/expected)
                    
                    # Highlight In Protected Users Group column (GREEN when Yes, RED when No for high-risk accounts)
                    if "In Protected Users Group" in headers and "Risk Level" in headers:
                        protected_cell = ws.cell(row=row_idx, column=headers["In Protected Users Group"])
                        risk_cell = ws.cell(row=row_idx, column=headers["Risk Level"])
                        risk_value = str(risk_cell.value).upper() if risk_cell.value else ""
                        
                        if protected_cell.value in ["Yes", "YES", True, "TRUE"]:
                            green_fill = PatternFill(start_color="B3FFB3", end_color="B3FFB3", fill_type="solid")
                            protected_cell.fill = green_fill
                        elif protected_cell.value in ["No", "NO", False, "FALSE"]:
                            # Only highlight red for HIGH risk accounts not in Protected Users
                            if risk_value in ["CRITICAL", "HIGH"]:
                                protected_cell.fill = red_fill
                            elif risk_value == "MEDIUM":
                                protected_cell.fill = orange_fill
        
        # Format krbtgt sheet
        if "krbtgt" in wb.sheetnames:
            ws = wb["krbtgt"]
            if ws.max_row > 1:  # Has data beyond header
                headers = {cell.value: cell.column for cell in ws[1]}
                
                # Apply formatting to data rows (skip header)
                for row_idx in range(2, ws.max_row + 1):
                    # Highlight Risk Level column
                    if "Risk Level" in headers:
                        risk_cell = ws.cell(row=row_idx, column=headers["Risk Level"])
                        risk_value = str(risk_cell.value).upper() if risk_cell.value else ""
                        
                        if risk_value in ["CRITICAL", "HIGH"]:
                            risk_cell.fill = red_fill
                        elif risk_value == "MEDIUM":
                            risk_cell.fill = orange_fill
                        # Low risk remains uncolored (not a problem, just normal/expected)
                    
                    # Highlight Password Age Days if > 365 (CRITICAL)
                    if "Password Age (Days)" in headers:
                        age_cell = ws.cell(row=row_idx, column=headers["Password Age (Days)"])
                        try:
                            age_val = float(age_cell.value) if age_cell.value else 0
                            if age_val > 365:
                                age_cell.fill = red_fill
                            elif age_val > 180:
                                age_cell.fill = orange_fill
                        except (ValueError, TypeError):
                            pass
                    
                    # Highlight Supports RC4 (ORANGE - weak encryption)
                    if "Supports RC4" in headers:
                        rc4_cell = ws.cell(row=row_idx, column=headers["Supports RC4"])
                        if rc4_cell.value in [True, "TRUE", "Yes", "YES"]:
                            rc4_cell.fill = orange_fill
        
        # Format ASREPRoastable sheet
        if "ASREPRoastable" in wb.sheetnames:
            ws = wb["ASREPRoastable"]
            if ws.max_row > 1:  # Has data beyond header
                headers = {cell.value: cell.column for cell in ws[1]}
                
                # Apply formatting to data rows (skip header)
                for row_idx in range(2, ws.max_row + 1):
                    # Highlight Risk Level column
                    if "Risk Level" in headers:
                        risk_cell = ws.cell(row=row_idx, column=headers["Risk Level"])
                        risk_value = str(risk_cell.value).upper() if risk_cell.value else ""
                        
                        if risk_value in ["CRITICAL", "HIGH"]:
                            risk_cell.fill = red_fill
                        elif risk_value == "MEDIUM":
                            risk_cell.fill = orange_fill
                        # Low risk remains uncolored (not a problem, just normal/expected)
                    
                    # Highlight Does Not Require Pre Auth (always RED - critical vulnerability)
                    if "Does Not Require Pre Auth" in headers:
                        preauth_cell = ws.cell(row=row_idx, column=headers["Does Not Require Pre Auth"])
                        if preauth_cell.value in [True, "TRUE", "Yes", "YES"]:
                            preauth_cell.fill = red_fill
                    
                    # Highlight AdminCount (ORANGE - privileged account)
                    if "AdminCount" in headers:
                        admin_cell = ws.cell(row=row_idx, column=headers["AdminCount"])
                        if admin_cell.value in [1, "1", True, "TRUE"]:
                            admin_cell.fill = orange_fill
        
        # Format Kerberoastable sheet
        if "Kerberoastable" in wb.sheetnames:
            ws = wb["Kerberoastable"]
            if ws.max_row > 1:  # Has data beyond header
                headers = {cell.value: cell.column for cell in ws[1]}
                
                # Apply formatting to data rows (skip header)
                for row_idx in range(2, ws.max_row + 1):
                    # Highlight Risk Level column
                    if "Risk Level" in headers:
                        risk_cell = ws.cell(row=row_idx, column=headers["Risk Level"])
                        risk_value = str(risk_cell.value).upper() if risk_cell.value else ""
                        
                        if risk_value in ["CRITICAL", "HIGH"]:
                            risk_cell.fill = red_fill
                        elif risk_value == "MEDIUM":
                            risk_cell.fill = orange_fill
                        # Low risk remains uncolored (not a problem, just normal/expected)
                    
                    # Highlight Supports RC4 (ORANGE - weak encryption, easier to crack)
                    if "Supports RC4" in headers:
                        rc4_cell = ws.cell(row=row_idx, column=headers["Supports RC4"])
                        if rc4_cell.value in [True, "TRUE", "Yes", "YES"]:
                            rc4_cell.fill = orange_fill
                    
                    # Highlight AdminCount (ORANGE - privileged account)
                    if "AdminCount" in headers:
                        admin_cell = ws.cell(row=row_idx, column=headers["AdminCount"])
                        if admin_cell.value in [1, "1", True, "TRUE"]:
                            admin_cell.fill = orange_fill
                    
                    # Highlight Password Never Expires (RED - critical, allows unlimited cracking time)
                    if "Password Never Expires" in headers:
                        pwd_cell = ws.cell(row=row_idx, column=headers["Password Never Expires"])
                        if pwd_cell.value in [True, "TRUE", "Yes", "YES"]:
                            pwd_cell.fill = red_fill

    def apply_striped_formatting(self, wb):
        """Apply striped row formatting (alternating light gray) to all data sheets.
        This is applied after security formatting to preserve security highlights.
        """
        from openpyxl.styles import PatternFill
        
        stripe_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # RGB values of colors to preserve (with 00 prefix that openpyxl adds)
        preserve_colors = {
            "00FFB3BA",  # Red (security)
            "FFB3BA",    # Red (without prefix)
            "00FFD9B3",  # Orange (security)
            "FFD9B3",    # Orange (without prefix)
            "00FFFFB3",  # Yellow (security)
            "FFFFB3",    # Yellow (without prefix)
            "00E0E0E0",  # Gray (disabled accounts)
            "E0E0E0",    # Gray (without prefix)
            "000066CC",  # Blue (headers)
            "0066CC",    # Blue (without prefix)
            "00B3FFB3",  # Green (positive indicators)
            "B3FFB3",    # Green (without prefix)
        }
        
        logger.info("    Applying striped row formatting...")
        
        skip_sheets = set()
        
        for sheet_name in wb.sheetnames:
            if sheet_name in skip_sheets:
                continue
                
            ws = wb[sheet_name]
            if ws.max_row <= 1:
                continue
            
            for row_idx in range(2, ws.max_row + 1):
                if row_idx % 2 == 0:
                    for col_idx in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        # Only apply striping if cell has no fill or has default fill
                        # Skip cells with colored fills (headers, security highlights, etc.)
                        cell_rgb = cell.fill.start_color.rgb if cell.fill and hasattr(cell.fill.start_color, 'rgb') else None
                        if cell_rgb is None or cell_rgb == '00000000' or cell_rgb not in preserve_colors:
                            cell.fill = stripe_fill

    def export_xlsx(self, output_dir: str, domain_name: str = ""):
        """Export results to Excel file (optimized for large datasets)."""
        if not OPENPYXL_AVAILABLE:
            logger.warning("[*] openpyxl not available - Excel export disabled")
            return None

        logger.info("[*] Generating Excel Report (optimized mode)...")
        start_time = datetime.now()

        try:
            from openpyxl import Workbook
            from openpyxl.cell import WriteOnlyCell

            wb = Workbook(write_only=True)

            # Define styles for headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
            left_alignment = Alignment(horizontal='left', vertical='top')

            column_widths = {}

            # Define sheet order to match ADRecon
            SHEET_ORDER = [
                'AboutPyADRecon', 'Users', 'UserSPNs',
                'krbtgt', 'ProtectedGroups', 'ASREPRoastable', 'Kerberoastable',
                'gMSA', 'dMSA', 'GroupMembers', 'Groups', 'OUs', 'Computers',
                'ComputerSPNs', 'LAPS', 'ManagedServiceAccounts', 'Printers', 'DNSZones', 'DNSRecords', 'gPLinks', 'GPOs',
                'DomainControllers', 'CertificateTemplates', 'CertificateAuthorities',
                'PasswordPolicy', 'FineGrainedPasswordPolicy',
                'SchemaHistory', 'Sites', 'Domain', 'Forest'
            ]
            
            SHEET_NAME_MAPPING = {
                'AboutPyADRecon': 'About PyADRecon',
                'ManagedServiceAccounts': 'Managed Service Accounts'
            }

            # Create About PyADRecon sheet first
            if 'AboutPyADRecon' in self.results and self.results['AboutPyADRecon']:
                data = self.results['AboutPyADRecon']
                logger.info(f"    [1/1] Writing AboutPyADRecon ({len(data):,} records)...")
                display_name = SHEET_NAME_MAPPING.get('AboutPyADRecon', 'AboutPyADRecon')
                ws = wb.create_sheet(display_name[:31])
                
                if data and len(data) > 0:
                    headers = list(data[0].keys())
                    header_row = []
                    for header in headers:
                        cell = WriteOnlyCell(ws, value=header)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = left_alignment
                        header_row.append(cell)
                    ws.append(header_row)
                    
                    for item in data:
                        ws.append([item.get(h, '') for h in headers])
            
            # Build TOC data
            toc_data = [["Sheet Name", "Record Count"]]
            
            stats_sheets = []
            if 'Users' in self.results and self.results['Users']:
                stats_sheets.append(['User Stats', 'Statistics'])
            if 'Computers' in self.results and self.results['Computers']:
                stats_sheets.append(['Computer Stats', 'Statistics'])
            
            for name in SHEET_ORDER:
                if name == 'AboutPyADRecon':
                    continue
                if name in self.results and self.results[name]:
                    friendly_name = SHEET_NAME_MAPPING.get(name, name)
                    toc_data.append([friendly_name, len(self.results[name])])

            # Create Table of Contents sheet
            toc_ws = wb.create_sheet("Table of Contents")
            
            header_row = []
            for header in toc_data[0]:
                cell = WriteOnlyCell(toc_ws, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = left_alignment
                header_row.append(cell)
            toc_ws.append(header_row)
            
            for sheet_info in stats_sheets:
                sheet_name = sheet_info[0]
                record_count = sheet_info[1]
                
                name_cell = WriteOnlyCell(toc_ws, value=sheet_name)
                name_cell.hyperlink = f"#'{sheet_name}'!A1"
                name_cell.font = Font(color="0563C1", underline="single")
                count_cell = WriteOnlyCell(toc_ws, value=record_count)
                
                toc_ws.append([name_cell, count_cell])
            
            for row_data in toc_data[1:]:
                sheet_name = row_data[0]
                record_count = row_data[1]
                
                name_cell = WriteOnlyCell(toc_ws, value=sheet_name)
                name_cell.hyperlink = f"#'{sheet_name}'!A1"
                name_cell.font = Font(color="0563C1", underline="single")
                
                count_cell = WriteOnlyCell(toc_ws, value=record_count)
                
                toc_ws.append([name_cell, count_cell])
            
            # Create User Stats tab
            if 'Users' in self.results and self.results['Users']:
                logger.info("    Creating User Stats tab...")
                user_stats = calculate_user_stats(self.results['Users'], self.config.password_age_days, self.config.dormant_days)
                user_stats_ws = wb.create_sheet("User Stats")
                
                heading_cell = WriteOnlyCell(user_stats_ws, value="Status of User Accounts")
                heading_cell.font = Font(bold=True, size=14)
                user_stats_ws.append([heading_cell])
                user_stats_ws.append([])
                
                summary_header_row = []
                for header in ["Category", "Count"]:
                    cell = WriteOnlyCell(user_stats_ws, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = left_alignment
                    summary_header_row.append(cell)
                user_stats_ws.append(summary_header_row)
                
                user_stats_ws.append(["Enabled Users", user_stats['enabled_count']])
                user_stats_ws.append(["Disabled Users", user_stats['disabled_count']])
                user_stats_ws.append(["Total Users", user_stats['total_count']])
                user_stats_ws.append([])
                
                stats_headers = ["Category", "Enabled Count", "Enabled Percentage", 
                               "Disabled Count", "Disabled Percentage", "Total Count", "Total Percentage"]
                stats_header_row = []
                for header in stats_headers:
                    cell = WriteOnlyCell(user_stats_ws, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = left_alignment
                    stats_header_row.append(cell)
                user_stats_ws.append(stats_header_row)
                
                total_enabled = user_stats['enabled_count']
                total_disabled = user_stats['disabled_count']
                total_count = user_stats['total_count']
                
                for category, counts in user_stats['categories'].items():
                    enabled_count = counts['enabled_count']
                    disabled_count = counts['disabled_count']
                    cat_total = counts['total_count']
                    
                    enabled_pct = f"{(enabled_count / total_enabled * 100):.1f}%" if total_enabled > 0 else "0.0%"
                    disabled_pct = f"{(disabled_count / total_disabled * 100):.1f}%" if total_disabled > 0 else "0.0%"
                    total_pct = f"{(cat_total / total_count * 100):.1f}%" if total_count > 0 else "0.0%"
                    
                    user_stats_ws.append([
                        category,
                        enabled_count,
                        enabled_pct,
                        disabled_count,
                        disabled_pct,
                        cat_total,
                        total_pct
                    ])
            
            # Create Computer Stats tab
            if 'Computers' in self.results and self.results['Computers']:
                logger.info("    Creating Computer Stats tab...")
                laps_data = self.results.get('LAPS', [])
                computer_stats = calculate_computer_stats(self.results['Computers'], laps_data, 30, self.config.dormant_days)
                computer_stats_ws = wb.create_sheet("Computer Stats")
                
                heading_cell = WriteOnlyCell(computer_stats_ws, value="Status of Computer Accounts")
                heading_cell.font = Font(bold=True, size=14)
                computer_stats_ws.append([heading_cell])
                computer_stats_ws.append([])
                
                summary_header_row = []
                for header in ["Category", "Count"]:
                    cell = WriteOnlyCell(computer_stats_ws, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = left_alignment
                    summary_header_row.append(cell)
                computer_stats_ws.append(summary_header_row)
                
                computer_stats_ws.append(["Enabled Computers", computer_stats['enabled_count']])
                computer_stats_ws.append(["Disabled Computers", computer_stats['disabled_count']])
                computer_stats_ws.append(["Total Computers", computer_stats['total_count']])
                computer_stats_ws.append([])
                
                stats_headers = ["Category", "Enabled Count", "Enabled Percentage", 
                               "Disabled Count", "Disabled Percentage", "Total Count", "Total Percentage"]
                stats_header_row = []
                for header in stats_headers:
                    cell = WriteOnlyCell(computer_stats_ws, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = left_alignment
                    stats_header_row.append(cell)
                computer_stats_ws.append(stats_header_row)
                
                total_enabled = computer_stats['enabled_count']
                total_disabled = computer_stats['disabled_count']
                total_count = computer_stats['total_count']
                
                for category, counts in computer_stats['categories'].items():
                    enabled_count = counts['enabled_count']
                    disabled_count = counts['disabled_count']
                    cat_total = counts['total_count']
                    
                    enabled_pct = f"{(enabled_count / total_enabled * 100):.1f}%" if total_enabled > 0 else "0.0%"
                    disabled_pct = f"{(disabled_count / total_disabled * 100):.1f}%" if total_disabled > 0 else "0.0%"
                    total_pct = f"{(cat_total / total_count * 100):.1f}%" if total_count > 0 else "0.0%"
                    
                    computer_stats_ws.append([
                        category,
                        enabled_count,
                        enabled_pct,
                        disabled_count,
                        disabled_pct,
                        cat_total,
                        total_pct
                    ])
            
            # Order sheets according to SHEET_ORDER
            ordered_names = []
            for sheet in SHEET_ORDER:
                if sheet != 'AboutPyADRecon' and sheet in self.results and self.results[sheet]:
                    ordered_names.append(sheet)
            for name in self.results.keys():
                if name != 'AboutPyADRecon' and name not in ordered_names and self.results[name]:
                    ordered_names.append(name)
            
            # Process each result set
            total_sheets = len(ordered_names)
            current_sheet = 0

            for name in ordered_names:
                data = self.results[name]
                if not data:
                    continue

                current_sheet += 1
                record_count = len(data)
                logger.info(f"    [{current_sheet}/{total_sheets}] Writing {name} ({record_count:,} records)...")

                if data and len(data) > 0:
                    first_key = list(data[0].keys())[0]
                    try:
                        data = sorted(data, key=lambda x: str(x.get(first_key, '')).lower())
                    except Exception as e:
                        logger.debug(f"Could not sort {name} by {first_key}: {e}")

                display_name = SHEET_NAME_MAPPING.get(name, name)
                ws = wb.create_sheet(display_name[:31])
                sheet_name = display_name[:31]

                headers = list(data[0].keys())

                column_widths[sheet_name] = {}
                for idx, header in enumerate(headers):
                    column_widths[sheet_name][idx] = len(str(header))

                header_row = []
                for header in headers:
                    cell = WriteOnlyCell(ws, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = left_alignment
                    header_row.append(cell)
                ws.append(header_row)

                batch_size = 10000
                max_sample_rows = 1000
                
                for i, row_data in enumerate(data):
                    row = []
                    for col_idx, header in enumerate(headers):
                        value = row_data.get(header, '')
                        if hasattr(value, 'raw_values'):
                            value = value.value
                        if value is None:
                            value = ''
                        elif isinstance(value, bytes):
                            value = value.decode('utf-8', errors='replace')
                        elif isinstance(value, (list, dict)):
                            value = str(value)
                        elif not isinstance(value, (str, int, float, bool)):
                            value = str(value)
                        
                        if i < max_sample_rows:
                            val_len = len(str(value))
                            if val_len > column_widths[sheet_name][col_idx]:
                                column_widths[sheet_name][col_idx] = val_len
                        
                        row.append(value)
                    ws.append(row)

                    if record_count > batch_size and (i + 1) % batch_size == 0:
                        pct = ((i + 1) / record_count) * 100
                        logger.info(f"        {i + 1:,}/{record_count:,} rows ({pct:.0f}%)")

            # Save workbook
            filename = os.path.join(output_dir, f"{domain_name or 'ADRecon'}-Report.xlsx")
            logger.info(f"    Saving Excel file...")
            wb.save(filename)

            # Reopen in edit mode to set column widths and filters
            logger.info(f"    Auto-sizing columns and adding filters...")
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter
            
            wb = load_workbook(filename)
            
            for sheet_name, widths in column_widths.items():
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    
                    for col_idx, width in widths.items():
                        adjusted_width = min(width + 2, 100)
                        column_letter = get_column_letter(col_idx + 1)
                        ws.column_dimensions[column_letter].width = adjusted_width
                    
                    for row in ws.iter_rows():
                        for cell in row:
                            cell.alignment = Alignment(horizontal='left', vertical='top')
                    
                    if ws.max_row > 0 and ws.max_column > 0:
                        ws.auto_filter.ref = ws.dimensions
            
            # Auto-size special sheets
            for special_sheet in ["About PyADRecon", "Table of Contents", "User Stats", "Computer Stats"]:
                if special_sheet in wb.sheetnames:
                    ws = wb[special_sheet]
                    
                    if special_sheet == "About PyADRecon":
                        for row in ws.iter_rows():
                            for cell in row:
                                cell.alignment = Alignment(horizontal='left', vertical='top')
                    elif special_sheet == "Table of Contents":
                        for row in ws.iter_rows():
                            for col_idx, cell in enumerate(row, 1):
                                if col_idx == 1:
                                    cell.alignment = Alignment(horizontal='left', vertical='top')
                                else:
                                    cell.alignment = Alignment(horizontal='center', vertical='top')
                    elif special_sheet in ["User Stats", "Computer Stats"]:
                        for row in ws.iter_rows():
                            for col_idx, cell in enumerate(row, 1):
                                if col_idx == 1:
                                    cell.alignment = Alignment(horizontal='left', vertical='top')
                                else:
                                    cell.alignment = Alignment(horizontal='center', vertical='top')
                    
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if cell.value:
                                    max_length = max(max_length, len(str(cell.value)))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 100)
                        ws.column_dimensions[column_letter].width = adjusted_width
            
            # Apply security-based conditional formatting
            logger.info("    Applying security formatting...")
            self.apply_security_formatting(wb)
            
            # Apply striped row formatting
            self.apply_striped_formatting(wb)
            
            # Save final formatted workbook
            logger.info("    Saving formatted Excel file...")
            wb.save(filename)
            wb.close()

            duration = datetime.now() - start_time
            logger.info(f"[+] Excel Report saved to: {filename} (took {duration})")
            return filename

        except Exception as e:
            logger.error(f"Failed to create Excel report: {e}")
            import traceback
            traceback.print_exc()
            return None

    def close(self):
        """Close connection."""
        if self.conn:
            logger.info("[*] Closing ADWS connection")
            # ADWS connection doesn't need explicit close typically


def generate_excel_from_csv(csv_dir: str, output_file: str = None):
    """
    Standalone function to generate Excel report from CSV files.
    This is optimized for large datasets and can be run independently.

    Args:
        csv_dir: Path to directory containing CSV files
        output_file: Output Excel file path (optional, defaults to same directory)
    """
    if not OPENPYXL_AVAILABLE:
        logger.error("[!] openpyxl not available - install with: pip install openpyxl")
        return None

    if not os.path.isdir(csv_dir):
        logger.error(f"[!] CSV directory not found: {csv_dir}")
        return None

    logger.info(f"[*] Generating Excel Report from CSV files in: {csv_dir}")
    start_time = datetime.now()

    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.cell import WriteOnlyCell

        # Define sheet order to match ADRecon
        SHEET_ORDER = [
            'AboutPyADRecon', 'Users', 'UserSPNs',
            'krbtgt', 'ProtectedGroups', 'ASREPRoastable', 'Kerberoastable',
            'gMSA', 'dMSA', 'GroupMembers', 'Groups', 'OUs', 'Computers',
            'ComputerSPNs', 'LAPS', 'Printers', 'DNSZones', 'DNSRecords', 'gPLinks', 'GPOs',
            'DomainControllers', 'CertificateTemplates', 'CertificateAuthorities',
            'PasswordPolicy', 'FineGrainedPasswordPolicy',
            'SchemaHistory', 'Sites', 'Domain', 'Forest'
        ]
        
        # Friendly sheet names mapping
        SHEET_NAME_MAPPING = {
            'AboutPyADRecon': 'About PyADRecon'
        }

        # Find all CSV files
        all_csv_files = [f for f in os.listdir(csv_dir) if f.endswith('.csv')]
        if not all_csv_files:
            logger.error(f"[!] No CSV files found in: {csv_dir}")
            return None

        # Order CSV files according to SHEET_ORDER
        csv_files = []
        for sheet in SHEET_ORDER:
            csv_name = sheet + '.csv'
            if csv_name in all_csv_files:
                csv_files.append(csv_name)
        
        # Add any remaining CSV files not in the order
        for csv_file in sorted(all_csv_files):
            if csv_file not in csv_files:
                csv_files.append(csv_file)

        logger.info(f"    Found {len(csv_files)} CSV files")

        # Use write_only mode for performance
        wb = Workbook(write_only=True)

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        left_alignment = Alignment(horizontal='left', vertical='top')

        # First pass - count records for TOC
        file_counts = {}
        for csv_file in csv_files:
            csv_path = os.path.join(csv_dir, csv_file)
            with open(csv_path, 'r', encoding='utf-8', errors='replace') as f:
                count = sum(1 for _ in f) - 1  # Subtract header
                file_counts[csv_file] = max(0, count)

        # Create About PyADRecon sheet first if it exists
        about_csv = 'AboutPyADRecon.csv'
        if about_csv in csv_files:
            csv_path = os.path.join(csv_dir, about_csv)
            display_name = SHEET_NAME_MAPPING.get('AboutPyADRecon', 'AboutPyADRecon')
            ws = wb.create_sheet(display_name[:31])
            record_count = file_counts[about_csv]
            logger.info(f"    [1/1] Processing {about_csv} ({record_count:,} records)...")
            
            with open(csv_path, 'r', encoding='utf-8', errors='replace', newline='') as f:
                reader = csv.reader(f)
                try:
                    headers = next(reader)
                    header_row = []
                    for header in headers:
                        cell = WriteOnlyCell(ws, value=header)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = left_alignment
                        header_row.append(cell)
                    ws.append(header_row)
                except StopIteration:
                    pass
                
                # Write data rows
                for row in reader:
                    ws.append(row)

        # Build TOC data
        toc_data = [
            ["Sheet Name", "Record Count"],
        ]
        
        # Add User Stats and Computer Stats first (if they will be created)
        users_csv_path = os.path.join(csv_dir, 'Users.csv')
        computers_csv_path = os.path.join(csv_dir, 'Computers.csv')
        
        if os.path.exists(users_csv_path):
            toc_data.append(["User Stats", "Statistics"])
        if os.path.exists(computers_csv_path):
            toc_data.append(["Computer Stats", "Statistics"])
        
        # Add sheets to TOC (excluding AboutPyADRecon as it was already created)
        for csv_file in csv_files:
            if csv_file != about_csv:
                sheet_name = csv_file.replace('.csv', '')
                toc_data.append([sheet_name, file_counts[csv_file]])

        # Create TOC sheet second
        toc_ws = wb.create_sheet("Table of Contents")
        
        # Write header with styling
        header_row = []
        for header in toc_data[0]:
            cell = WriteOnlyCell(toc_ws, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = left_alignment
            header_row.append(cell)
        toc_ws.append(header_row)
        
        # Write data rows
        for row in toc_data[1:]:
            toc_ws.append(row)

        # Create User Stats tab if Users.csv exists
        if os.path.exists(users_csv_path):
            logger.info("    Creating User Stats tab from Users.csv...")
            # Read Users.csv into memory and detect thresholds from column names
            users_data = []
            password_age_days = 180  # Default
            dormant_days = 90  # Default
            
            with open(users_csv_path, 'r', encoding='utf-8', errors='replace', newline='') as f:
                reader = csv.DictReader(f)
                
                # Try to extract thresholds from column names
                if reader.fieldnames:
                    for fieldname in reader.fieldnames:
                        if 'Password Age (>' in fieldname and 'days)' in fieldname:
                            try:
                                match = re.search(r'> (\d+) days', fieldname)
                                if match:
                                    password_age_days = int(match.group(1))
                            except:
                                pass
                        elif 'Dormant (>' in fieldname and 'days)' in fieldname:
                            try:
                                match = re.search(r'> (\d+) days', fieldname)
                                if match:
                                    dormant_days = int(match.group(1))
                            except:
                                pass
                
                for row in reader:
                    # Convert string values to appropriate types
                    processed_row = {}
                    for key, value in row.items():
                        if value in ['True', 'TRUE', 'true']:
                            processed_row[key] = True
                        elif value in ['False', 'FALSE', 'false']:
                            processed_row[key] = False
                        elif value == '':
                            processed_row[key] = ''
                        else:
                            # Try to convert to number
                            try:
                                if '.' in value:
                                    processed_row[key] = float(value)
                                else:
                                    processed_row[key] = int(value)
                            except (ValueError, AttributeError):
                                processed_row[key] = value
                    users_data.append(processed_row)
            
            # Calculate statistics with detected thresholds
            user_stats = calculate_user_stats(users_data, password_age_days, dormant_days)
            user_stats_ws = wb.create_sheet("User Stats")
            
            # Add heading
            heading_cell = WriteOnlyCell(user_stats_ws, value="Status of User Accounts")
            heading_cell.font = Font(bold=True, size=14)
            user_stats_ws.append([heading_cell])
            user_stats_ws.append([])  # Blank row
            
            # Add enabled/disabled summary
            summary_header_row = []
            for header in ["Category", "Count"]:
                cell = WriteOnlyCell(user_stats_ws, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = left_alignment
                summary_header_row.append(cell)
            user_stats_ws.append(summary_header_row)
            
            user_stats_ws.append(["Enabled Users", user_stats['enabled_count']])
            user_stats_ws.append(["Disabled Users", user_stats['disabled_count']])
            user_stats_ws.append(["Total Users", user_stats['total_count']])
            user_stats_ws.append([])  # Blank row
            
            # Add statistics table header
            stats_headers = ["Category", "Enabled Count", "Enabled Percentage", 
                           "Disabled Count", "Disabled Percentage", "Total Count", "Total Percentage"]
            stats_header_row = []
            for header in stats_headers:
                cell = WriteOnlyCell(user_stats_ws, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = left_alignment
                stats_header_row.append(cell)
            user_stats_ws.append(stats_header_row)
            
            # Add statistics data
            total_enabled = user_stats['enabled_count']
            total_disabled = user_stats['disabled_count']
            total_count = user_stats['total_count']
            
            for category, counts in user_stats['categories'].items():
                enabled_count = counts['enabled_count']
                disabled_count = counts['disabled_count']
                cat_total = counts['total_count']
                
                # Calculate percentages
                enabled_pct = f"{(enabled_count / total_enabled * 100):.1f}%" if total_enabled > 0 else "0.0%"
                disabled_pct = f"{(disabled_count / total_disabled * 100):.1f}%" if total_disabled > 0 else "0.0%"
                total_pct = f"{(cat_total / total_count * 100):.1f}%" if total_count > 0 else "0.0%"
                
                user_stats_ws.append([
                    category,
                    enabled_count,
                    enabled_pct,
                    disabled_count,
                    disabled_pct,
                    cat_total,
                    total_pct
                ])

        # Create Computer Stats tab if Computers.csv exists
        computers_csv_path = os.path.join(csv_dir, 'Computers.csv')
        if os.path.exists(computers_csv_path):
            logger.info("    Creating Computer Stats tab from Computers.csv...")
            # Read Computers.csv into memory and detect thresholds from column names
            computers_data = []
            password_age_days = 30  # Default for computers
            dormant_days = 90  # Default
            
            with open(computers_csv_path, 'r', encoding='utf-8', errors='replace', newline='') as f:
                reader = csv.DictReader(f)
                
                # Try to extract thresholds from column names
                if reader.fieldnames:
                    for fieldname in reader.fieldnames:
                        if 'Password Age (>' in fieldname and 'days)' in fieldname:
                            try:
                                match = re.search(r'> (\d+) days', fieldname)
                                if match:
                                    password_age_days = int(match.group(1))
                            except:
                                pass
                        elif 'Dormant (>' in fieldname and 'days)' in fieldname:
                            try:
                                match = re.search(r'> (\d+) days', fieldname)
                                if match:
                                    dormant_days = int(match.group(1))
                            except:
                                pass
                
                for row in reader:
                    # Convert string values to appropriate types
                    processed_row = {}
                    for key, value in row.items():
                        if value in ['True', 'TRUE', 'true']:
                            processed_row[key] = True
                        elif value in ['False', 'FALSE', 'false']:
                            processed_row[key] = False
                        elif value == '':
                            processed_row[key] = ''
                        else:
                            # Try to convert to number
                            try:
                                if '.' in value:
                                    processed_row[key] = float(value)
                                else:
                                    processed_row[key] = int(value)
                            except (ValueError, AttributeError):
                                processed_row[key] = value
                    computers_data.append(processed_row)
            
            # Read LAPS.csv if it exists for LAPS detection
            laps_data = []
            laps_csv_path = os.path.join(csv_dir, 'LAPS.csv')
            if os.path.exists(laps_csv_path):
                with open(laps_csv_path, 'r', encoding='utf-8', errors='replace', newline='') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        processed_row = {}
                        for key, value in row.items():
                            if value in ['True', 'TRUE', 'true']:
                                processed_row[key] = True
                            elif value in ['False', 'FALSE', 'false']:
                                processed_row[key] = False
                            else:
                                processed_row[key] = value
                        laps_data.append(processed_row)
            
            # Calculate statistics with detected thresholds
            computer_stats = calculate_computer_stats(computers_data, laps_data, password_age_days, dormant_days)
            computer_stats_ws = wb.create_sheet("Computer Stats")
            
            # Add heading
            heading_cell = WriteOnlyCell(computer_stats_ws, value="Status of Computer Accounts")
            heading_cell.font = Font(bold=True, size=14)
            computer_stats_ws.append([heading_cell])
            computer_stats_ws.append([])  # Blank row
            
            # Add enabled/disabled summary
            summary_header_row = []
            for header in ["Category", "Count"]:
                cell = WriteOnlyCell(computer_stats_ws, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = left_alignment
                summary_header_row.append(cell)
            computer_stats_ws.append(summary_header_row)
            
            computer_stats_ws.append(["Enabled Computers", computer_stats['enabled_count']])
            computer_stats_ws.append(["Disabled Computers", computer_stats['disabled_count']])
            computer_stats_ws.append(["Total Computers", computer_stats['total_count']])
            computer_stats_ws.append([])  # Blank row
            
            # Add statistics table header
            stats_headers = ["Category", "Enabled Count", "Enabled Percentage", 
                           "Disabled Count", "Disabled Percentage", "Total Count", "Total Percentage"]
            stats_header_row = []
            for header in stats_headers:
                cell = WriteOnlyCell(computer_stats_ws, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = left_alignment
                stats_header_row.append(cell)
            computer_stats_ws.append(stats_header_row)
            
            # Add statistics data
            total_enabled = computer_stats['enabled_count']
            total_disabled = computer_stats['disabled_count']
            total_count = computer_stats['total_count']
            
            for category, counts in computer_stats['categories'].items():
                enabled_count = counts['enabled_count']
                disabled_count = counts['disabled_count']
                cat_total = counts['total_count']
                
                # Calculate percentages
                enabled_pct = f"{(enabled_count / total_enabled * 100):.1f}%" if total_enabled > 0 else "0.0%"
                disabled_pct = f"{(disabled_count / total_disabled * 100):.1f}%" if total_disabled > 0 else "0.0%"
                total_pct = f"{(cat_total / total_count * 100):.1f}%" if total_count > 0 else "0.0%"
                
                computer_stats_ws.append([
                    category,
                    enabled_count,
                    enabled_pct,
                    disabled_count,
                    disabled_pct,
                    cat_total,
                    total_pct
                ])

        # Process each CSV file (excluding AboutPyADRecon as it was already created)
        remaining_csv_files = [f for f in csv_files if f != 'AboutPyADRecon.csv']
        total_files = len(remaining_csv_files)
        for idx, csv_file in enumerate(remaining_csv_files, 1):
            original_name = csv_file.replace('.csv', '')
            # Use friendly name if available, otherwise use original name
            display_name = SHEET_NAME_MAPPING.get(original_name, original_name)
            sheet_name = display_name[:31]  # Excel limit
            record_count = file_counts[csv_file]

            logger.info(f"    [{idx}/{total_files}] Processing {csv_file} ({record_count:,} records)...")

            csv_path = os.path.join(csv_dir, csv_file)
            ws = wb.create_sheet(sheet_name)

            with open(csv_path, 'r', encoding='utf-8', errors='replace', newline='') as f:
                reader = csv.reader(f)

                # Write header with styling
                try:
                    headers = next(reader)
                    header_row = []
                    for header in headers:
                        cell = WriteOnlyCell(ws, value=header)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = left_alignment
                        header_row.append(cell)
                    ws.append(header_row)
                except StopIteration:
                    continue  # Empty file

                # Read all data rows into memory for sorting
                all_rows = []
                for row in reader:
                    # Convert empty strings and handle encoding
                    clean_row = []
                    for val in row:
                        if val == '':
                            clean_row.append('')
                        else:
                            # Try to convert to number if possible
                            try:
                                if '.' in val:
                                    clean_row.append(float(val))
                                else:
                                    clean_row.append(int(val))
                            except ValueError:
                                clean_row.append(val)
                    all_rows.append(clean_row)
                
                # Sort by first column (case-insensitive string comparison)
                try:
                    all_rows.sort(key=lambda x: str(x[0]).lower() if len(x) > 0 else '')
                except Exception as e:
                    logger.debug(f"Could not sort {csv_file}: {e}")
                
                # Write sorted data rows
                batch_size = 10000
                for row_num, clean_row in enumerate(all_rows, 1):
                    ws.append(clean_row)

                    # Progress for large files
                    if record_count > batch_size and row_num % batch_size == 0:
                        pct = (row_num / record_count) * 100
                        logger.info(f"        {row_num:,}/{record_count:,} rows ({pct:.0f}%)")

        # Determine output filename
        if output_file:
            filename = output_file
        else:
            # Put in parent directory of CSV-Files
            parent_dir = os.path.dirname(csv_dir.rstrip('/'))
            if os.path.basename(csv_dir) == 'CSV-Files':
                filename = os.path.join(parent_dir, "ADRecon-Report.xlsx")
            else:
                filename = os.path.join(csv_dir, "ADRecon-Report.xlsx")

        logger.info(f"    Saving Excel file...")
        wb.save(filename)
        
        # Reopen to add filters and auto-size columns
        logger.info(f"    Adding filters and auto-sizing columns...")
        
        wb = load_workbook(filename)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Apply alignment based on sheet type
            if sheet_name == "Table of Contents":
                # TOC: first column left-aligned, second column (Record Count) centered
                for row in ws.iter_rows():
                    for col_idx, cell in enumerate(row, 1):
                        if col_idx == 1:
                            cell.alignment = Alignment(horizontal='left', vertical='top')
                        else:
                            cell.alignment = Alignment(horizontal='center', vertical='top')
            elif sheet_name in ["User Stats", "Computer Stats"]:
                # Stats sheets: first column left-aligned, rest centered
                for row in ws.iter_rows():
                    for col_idx, cell in enumerate(row, 1):
                        if col_idx == 1:
                            cell.alignment = Alignment(horizontal='left', vertical='top')
                        else:
                            cell.alignment = Alignment(horizontal='center', vertical='top')
            else:
                # All other sheets: left-aligned
                for row in ws.iter_rows():
                    for cell in row:
                        cell.alignment = Alignment(horizontal='left', vertical='top')
            
            # Auto-size all columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 100)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Add filters (skip TOC and stats sheets)
            if sheet_name not in ["Table of Contents", "User Stats", "Computer Stats"]:
                if ws.max_row > 0 and ws.max_column > 0:
                    ws.auto_filter.ref = ws.dimensions
        
        wb.save(filename)
        wb.close()

        duration = datetime.now() - start_time
        logger.info(f"[+] Excel Report saved to: {filename}")
        logger.info(f"[+] Total time: {duration}")
        return filename

    except Exception as e:
        logger.error(f"Failed to create Excel report: {e}")
        import traceback
        traceback.print_exc()
        return None


def generate_dashboard_from_csv(csv_dir: str, output_file: str = None):
    """
    Standalone function to generate HTML dashboard from CSV files.
    
    Args:
        csv_dir: Path to directory containing CSV files
        output_file: Output HTML file path (optional, defaults to dashboard.html in same directory)
    
    Returns:
        Path to generated dashboard file on success, None on failure
    """
    if not DASHBOARD_AVAILABLE:
        logger.error("[!] Dashboard generator not available (dashboard_generator.py not found)")
        return None
    
    if not os.path.isdir(csv_dir):
        logger.error(f"[!] CSV directory not found: {csv_dir}")
        return None
    
    logger.info(f"[*] Generating HTML dashboard from CSV files in: {csv_dir}")
    start_time = datetime.now()
    
    try:
        # Determine output file path
        if output_file:
            if os.path.isdir(output_file):
                dashboard_file = os.path.join(output_file, 'dashboard.html')
            else:
                dashboard_file = output_file
        else:
            # Default to dashboard.html in the CSV directory's parent
            parent_dir = os.path.dirname(csv_dir) if os.path.dirname(csv_dir) else csv_dir
            dashboard_file = os.path.join(parent_dir, 'dashboard.html')
        
        # Generate the dashboard
        result = generate_dashboard(csv_dir, dashboard_file)
        
        if result:
            duration = datetime.now() - start_time
            logger.info(f"[+] Dashboard saved to: {os.path.abspath(dashboard_file)}")
            logger.info(f"[+] Total time: {duration}")
            return os.path.abspath(dashboard_file)
        else:
            logger.error("[!] Dashboard generation failed")
            return None
            
    except Exception as e:
        logger.error(f"[!] Failed to generate dashboard: {e}")
        import traceback
        traceback.print_exc()
        return None


def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(
        description='PyADRecon-ADWS # Active Directory Reconnaissance using ADWS',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Basic usage with NTLM authentication
  %(prog)s -dc 192.168.1.1 -u admin -p password123 -d DOMAIN.LOCAL

  # With Kerberos authentication (only works on Linux with gssapi atm)
  %(prog)s -dc dc01.domain.local -u admin -p password123 -d DOMAIN.LOCAL --auth kerberos

  # Only collect specific modules
  %(prog)s -dc 192.168.1.1 -u admin -p pass -d DOMAIN.LOCAL --collect users,groups,computers

  # Output to specific directory
  %(prog)s -dc 192.168.1.1 -u admin -p pass -d DOMAIN.LOCAL -o /tmp/adrecon_output

  # Generate Excel report from existing CSV files (standalone mode)
  %(prog)s --generate-excel-from /path/to/CSV-Files -o report.xlsx

  # Generate HTML dashboard from existing CSV files (standalone mode)
  %(prog)s --generate-dashboard-from /path/to/CSV-Files -o dashboard.html
        """
    )

    # Version
    parser.add_argument('--version', action='version', version=f'PyADRecon-ADWS {VERSION}')

    # Standalone generation modes
    parser.add_argument('--generate-excel-from', metavar='CSV_DIR',
                        help='Generate Excel report from existing CSV files (standalone mode)')
    parser.add_argument('--generate-dashboard-from', metavar='CSV_DIR',
                        help='Generate HTML dashboard from existing CSV files (standalone mode)')

    # Connection parameters
    parser.add_argument('-d', '--domain', help='Domain name (e.g., example.com)')
    parser.add_argument('-u', '--username', help='Username (DOMAIN\\user or user@domain.com)')
    parser.add_argument('-p', '--password', help='Password or LM:NTLM hash (will prompt if not provided)')
    parser.add_argument('-dc', '--domain-controller', help='Domain controller hostname or IP')
    parser.add_argument('--port', type=int, default=9389, help='ADWS port (default: 9389)')

    # Authentication options
    parser.add_argument('--auth', choices=['ntlm', 'kerberos'], default='ntlm',
                        help='Authentication method: ntlm or kerberos (default: ntlm)')
    parser.add_argument('--spn', type=str, default=None,
                        help='Service Principal Name override (default: HTTP/dc.fqdn)')
    parser.add_argument('--workstation', default='',
                        help='NTLM authentication workstation name (default: empty)')

    # Collection options  
    parser.add_argument('-c', '--collect', default='default',
                        help='Comma-separated modules to collect (default: all)')
    parser.add_argument('--only-enabled', action='store_true',
                        help='Only collect enabled users/computers')

    # Threshold options
    parser.add_argument('--page-size', type=int, default=256,
                        help='ADWS query page size (default: 256)')
    parser.add_argument('--dormant-days', type=int, default=90,
                        help='Users/Computers with lastLogon older than X days are dormant (default: 90)')
    parser.add_argument('--password-age', type=int, default=180,
                        help='Users with pwdLastSet older than X days have old passwords (default: 180)')

    # Output options
    parser.add_argument('-o', '--output', help='Output directory (default: PyADRecon-Report-<timestamp>)')
    parser.add_argument('--no-excel', action='store_true', help='Skip Excel export')
    parser.add_argument('--no-dashboard', action='store_true',
                        help='Skip interactive HTML dashboard generation')
    parser.add_argument('-v', '--verbose', action='store_true', help='Enable verbose output')

    args = parser.parse_args()

    # Standalone Excel generation mode
    if args.generate_excel_from:
        print(BANNER)
        if not OPENPYXL_AVAILABLE:
            logger.error("[!] openpyxl not available - install with: pip install openpyxl")
            sys.exit(1)
        
        logger.info(f"[*] Generating Excel report from: {args.generate_excel_from}")
        result = generate_excel_from_csv(args.generate_excel_from, args.output)
        if result:
            sys.exit(0)
        else:
            sys.exit(1)
    
    # Standalone Dashboard generation mode
    if args.generate_dashboard_from:
        print(BANNER)
        if not DASHBOARD_AVAILABLE:
            logger.error("[!] Dashboard generator not available (dashboard_generator.py not found)")
            sys.exit(1)
        
        logger.info(f"[*] Generating HTML dashboard from: {args.generate_dashboard_from}")
        result = generate_dashboard_from_csv(args.generate_dashboard_from, args.output)
        if result:
            logger.info(f"[+] Dashboard generated: {result}")
            sys.exit(0)
        else:
            sys.exit(1)

    # Print banner for normal mode
    print(BANNER)

    # Configure logging level
    if args.verbose:
        logging.basicConfig(level=logging.DEBUG, format='[%(levelname)s] %(message)s', force=True)
        logger.setLevel(logging.DEBUG)
    else:
        logging.basicConfig(level=logging.INFO, format='[%(levelname)s] %(message)s', force=True)
        logger.setLevel(logging.INFO)

    # Validate required arguments for normal mode
    if not args.domain or not args.username or not args.domain_controller:
        parser.error("the following arguments are required: -d/--domain, -u/--username, -dc/--domain-controller")

    # Kerberos authentication warnings
    if args.auth == 'kerberos':
        logger.info("[*] Using Kerberos authentication")
        if not args.password:
            logger.info("[*] No password provided - will use Kerberos credential cache")
        # Don't prompt for password if using Kerberos with ccache
        if not args.password:
            try:
                import subprocess
                result = subprocess.run(['klist'], capture_output=True, text=True, timeout=2)
                if result.returncode == 0:
                    logger.info("[*] Found Kerberos tickets in cache")
                    args.password = None  # Use ccache
                else:
                    logger.warning("[!] No Kerberos tickets found - you may need to run 'kinit user@DOMAIN.COM' first")
                    import getpass
                    args.password = getpass.getpass('Password: ')
            except (subprocess.TimeoutExpired, FileNotFoundError):
                logger.warning("[!] Could not check Kerberos ticket cache")
                import getpass
                args.password = getpass.getpass('Password: ')
    else:
        # Get password if not provided for NTLM
        if not args.password:
            import getpass
            args.password = getpass.getpass('Password: ')

    # Parse collection modules
    collect_modules = args.collect.lower().split(',')

    # Create config
    config = ADReconConfig(
        domain_controller=args.domain_controller,
        domain=args.domain,
        username=args.username,
        password=args.password,
        auth_method=args.auth.upper(),
        port=args.port,
        only_enabled=args.only_enabled,
        page_size=args.page_size,
        dormant_days=args.dormant_days,
        password_age_days=args.password_age,
        workstation=args.workstation if args.workstation else None,
        spn=args.spn if hasattr(args, 'spn') else None,
    )

    # Configure collection based on modules
    if 'default' not in collect_modules and 'all' not in collect_modules:
        # Disable all first, then enable only requested modules
        config.collect_forest = 'forest' in collect_modules
        config.collect_domain = 'domain' in collect_modules
        config.collect_trusts = 'trusts' in collect_modules
        config.collect_sites = 'sites' in collect_modules
        config.collect_subnets = 'subnets' in collect_modules
        config.collect_schema = 'schema' in collect_modules
        config.collect_password_policy = 'passwordpolicy' in collect_modules or 'password' in collect_modules
        config.collect_fgpp = 'fgpp' in collect_modules
        config.collect_dcs = 'dcs' in collect_modules or 'domaincontrollers' in collect_modules
        config.collect_users = 'users' in collect_modules
        config.collect_user_spns = 'userspns' in collect_modules
        config.collect_groups = 'groups' in collect_modules
        config.collect_group_members = 'groupmembers' in collect_modules
        config.collect_ous = 'ous' in collect_modules
        config.collect_gpos = 'gpos' in collect_modules
        config.collect_gplinks = 'gplinks' in collect_modules
        config.collect_dns_zones = 'dnszones' in collect_modules or 'dns' in collect_modules
        config.collect_dns_records = 'dnsrecords' in collect_modules or 'dns' in collect_modules
        config.collect_printers = 'printers' in collect_modules
        config.collect_computers = 'computers' in collect_modules
        config.collect_computer_spns = 'computerspns' in collect_modules
        config.collect_laps = 'laps' in collect_modules
        config.collect_bitlocker = 'bitlocker' in collect_modules
        config.collect_gmsa = 'gmsa' in collect_modules
        config.collect_dmsa = 'dmsa' in collect_modules
        config.collect_adcs = 'adcs' in collect_modules or 'certificates' in collect_modules
        config.collect_protected_groups = 'protectedgroups' in collect_modules or 'protected' in collect_modules
        config.collect_krbtgt = 'krbtgt' in collect_modules
        config.collect_kerberoastable = 'kerberoastable' in collect_modules
        config.collect_asrep_roastable = 'asreproastable' in collect_modules or 'asrep' in collect_modules

    # Create output directory
    if args.output:
        output_dir = args.output
    else:
        output_dir = f"PyADRecon-ADWS-Report-{datetime.now().strftime('%Y%m%d%H%M%S')}"

    os.makedirs(output_dir, exist_ok=True)
    config.output_dir = output_dir

    # Run reconnaissance
    recon = PyADRecon(config)

    try:
        if recon.run():
            # Export results
            csv_dir = recon.export_csv(output_dir)

            if not args.no_excel and OPENPYXL_AVAILABLE:
                domain_name = config.domain.replace('.', '_')
                recon.export_xlsx(output_dir, domain_name)

            # Generate interactive dashboard if requested
            if not args.no_dashboard:
                if 'default' in collect_modules or 'all' in collect_modules or not collect_modules:
                    if DASHBOARD_AVAILABLE:
                        logger.info("[*] Generating interactive HTML dashboard...")
                        dashboard_file = os.path.join(output_dir, 'dashboard.html')
                        if generate_dashboard(csv_dir, dashboard_file):
                            logger.info(f"[+] Dashboard: {os.path.abspath(dashboard_file)}")
                    else:
                        logger.warning("[!] Dashboard generation not available (dashboard_generator.py not found)")

            logger.info(f"[*] Output Directory: {os.path.abspath(output_dir)}")
            logger.info("[*] Completed.")
        else:
            logger.error("[!] Reconnaissance failed")
            sys.exit(1)

    except KeyboardInterrupt:
        logger.warning("\n[!] Interrupted by user")
        sys.exit(1)
    finally:
        recon.close()


if __name__ == "__main__":
    main()
